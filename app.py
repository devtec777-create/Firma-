from __future__ import annotations

import csv
import html
import os
import re
import secrets
import shutil
import smtplib
import sqlite3
import subprocess
import unicodedata
import zipfile
from datetime import datetime
from email.message import EmailMessage
from functools import wraps
from io import BytesIO
from io import StringIO
from pathlib import Path
from xml.etree import ElementTree as ET

from flask import (
    Flask,
    Response,
    abort,
    flash,
    g,
    redirect,
    render_template_string,
    request,
    send_file,
    session,
    url_for,
)
from pypdf import PdfReader, PdfWriter
from reportlab.lib.colors import HexColor
from reportlab.lib.pagesizes import letter
from reportlab.pdfgen import canvas
from werkzeug.security import check_password_hash, generate_password_hash
from werkzeug.utils import secure_filename


BASE_DIR = Path(__file__).resolve().parent
DATA_DIR = BASE_DIR / "data"
UPLOAD_DIR = DATA_DIR / "uploads"
SIGNED_DIR = DATA_DIR / "signed"
CONFIG_IMAGE_DIR = DATA_DIR / "config_images"
AUDIO_DIR = DATA_DIR / "audio"
DEFAULT_DB_PATH = DATA_DIR / "firma_digital.sqlite3"
LEGACY_DB_PATH = DATA_DIR / ".sqlite3"
DB_PATH = Path(
    os.environ.get(
        "ESIGNUM_DB_PATH",
        str(LEGACY_DB_PATH if LEGACY_DB_PATH.exists() and LEGACY_DB_PATH.stat().st_size > 0 and (not DEFAULT_DB_PATH.exists() or DEFAULT_DB_PATH.stat().st_size == 0) else DEFAULT_DB_PATH),
    )
)

ALLOWED_EXTENSIONS = {".pdf"}
ALLOWED_AUDIO_EXTENSIONS = {".mp3", ".wav", ".m4a", ".ogg", ".webm", ".mp4", ".mpeg", ".mpga"}
ALLOWED_CONFIG_IMAGE_EXTENSIONS = {".jpg", ".jpeg", ".png"}
WORD_MIMETYPE = "application/vnd.openxmlformats-officedocument.wordprocessingml.document"
DEFAULT_PASSWORD = "Cecom1115"
DEFAULT_ADMIN_RUT = "76.032.479-5"
RUT_PATTERN = re.compile(r"^\d{1,2}\.\d{3}\.\d{3}-[\dkK]$")
MENU_POSITIONS = {
    "top": "Superior",
    "bottom": "Inferior",
    "left": "Izquierda",
    "right": "Derecha",
}
BUTTON_COLORS = {
    "blue": ("Azul", "#1f5fbf"),
    "light_green": ("Verde claro", "#2f9e44"),
    "sky": ("Celeste", "#168aad"),
    "maroon": ("Granate", "#8f1d3f"),
    "gray": ("Gris", "#5b6775"),
    "dark_blue": ("Azul oscuro", "#0b2f6b"),
}
BUTTON_TEXT_COLORS = {
    "white": ("Blanco", "#ffffff"),
    "dark": ("Oscuro", "#16202a"),
}
FIELD_BORDER_COLORS = {
    "soft_gray": ("Gris claro", "#d8dee8"),
    "blue": ("Azul", "#1f5fbf"),
    "light_green": ("Verde claro", "#2f9e44"),
    "sky": ("Celeste", "#168aad"),
    "maroon": ("Granate", "#8f1d3f"),
    "gray": ("Gris", "#5b6775"),
    "dark_blue": ("Azul oscuro", "#0b2f6b"),
}
PAGE_BACKGROUNDS = {
    "soft_gray": ("Gris claro recomendado", "#f5f7fb"),
    "soft_blue": ("Celeste suave recomendado", "#eef7fb"),
    "soft_green": ("Verde suave recomendado", "#f1fbf4"),
    "soft_wine": ("Granate muy claro recomendado", "#fbf1f5"),
}
ACTA_FONT_FAMILIES = {
    "Helvetica": "Helvetica",
    "Times-Roman": "Times New Roman",
    "Courier": "Courier",
}
ACTA_TITLE_ALIGNMENTS = {
    "left": "Izquierda",
    "center": "Centrado",
    "right": "Derecha",
}
DEFAULT_SYSTEM_SETTINGS = {
    "menu_position": "top",
    "button_color": "blue",
    "button_text_color": "white",
    "field_border_color": "soft_gray",
    "page_background": "soft_gray",
    "screen_image_path": "",
    "screen_image_position": "top_right",
    "screen_image_size": "180",
    "screen_image_2_path": "",
    "screen_image_2_position": "top_left",
    "screen_image_2_size": "180",
    "screen_image_3_path": "",
    "screen_image_3_position": "bottom_right",
    "screen_image_3_size": "180",
    "smtp_host": "",
    "smtp_port": "587",
    "smtp_user": "",
    "smtp_password": "",
    "smtp_from": "",
    "smtp_tls": "1",
    "acta_title_line_1": "",
    "acta_title_line_2": "",
    "acta_title_alignment": "center",
    "acta_image_path": "",
    "acta_image_filename": "",
    "acta_image_enabled": "1",
    "acta_image_position": "top_right",
    "acta_font_family": "Helvetica",
    "acta_body_font_size": "9",
    "acta_title_font_size": "13",
    "acta_watermark_image_path": "",
    "acta_watermark_image_filename": "",
    "acta_watermark_enabled": "1",
    "acta_watermark_size": "260",
    "acta_watermark_opacity": "15",
}
SCREEN_IMAGE_POSITIONS = {
    "top_left": "Superior izquierda",
    "top_center": "Superior centro",
    "top_right": "Superior derecha",
    "center_left": "Centro izquierda",
    "bottom_left": "Inferior izquierda",
    "center_right": "Centro derecha",
    "bottom_right": "Inferior derecha",
    "bottom_center": "Inferior centro",
    "center": "Centro",
}

app = Flask(__name__)
app.secret_key = os.environ.get("FIRMA_SECRET_KEY", secrets.token_hex(32))


def now() -> str:
    return datetime.now().strftime("%Y-%m-%d %H:%M:%S")


def ensure_dirs() -> None:
    DATA_DIR.mkdir(exist_ok=True)
    UPLOAD_DIR.mkdir(exist_ok=True)
    SIGNED_DIR.mkdir(exist_ok=True)
    CONFIG_IMAGE_DIR.mkdir(exist_ok=True)
    AUDIO_DIR.mkdir(exist_ok=True)


def store_config_image(uploaded, prefix: str) -> str | None:
    if not uploaded or not uploaded.filename:
        return None
    extension = Path(uploaded.filename).suffix.lower()
    if extension not in ALLOWED_CONFIG_IMAGE_EXTENSIONS:
        raise ValueError("La imagen debe ser JPG o PNG.")
    filename = secure_filename(uploaded.filename)
    stored_path = CONFIG_IMAGE_DIR / f"{prefix}-{secrets.token_hex(8)}-{filename}"
    uploaded.save(stored_path)
    return str(stored_path)


def safe_display_filename(name: str) -> str:
    cleaned = re.sub(r'[<>:"/\\|?*\x00-\x1f]', " ", name).strip()
    cleaned = re.sub(r"\s+", " ", cleaned)
    return cleaned.rstrip(". ") or "documento"


def db() -> sqlite3.Connection:
    if "db" not in g:
        conn = sqlite3.connect(DB_PATH)
        conn.row_factory = sqlite3.Row
        conn.execute("PRAGMA foreign_keys = ON")
        g.db = conn
    return g.db


@app.teardown_appcontext
def close_db(_exc: object) -> None:
    conn = g.pop("db", None)
    if conn is not None:
        conn.close()


def audit(action: str, details: str = "", target_user_id: int | None = None, document_id: int | None = None) -> None:
    user_id = session.get("user_id")
    db().execute(
        """
        INSERT INTO audit_log(actor_user_id, target_user_id, document_id, action, details, ip_address, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?)
        """,
        (user_id, target_user_id, document_id, action, details, request.remote_addr if request else "", now()),
    )
    db().commit()


def send_email(to_email: str, subject: str, body: str) -> bool:
    settings = system_settings()
    host = settings.get("smtp_host", "").strip()
    try:
        port = int(settings.get("smtp_port", "587") or "587")
    except ValueError:
        port = 587
    username = settings.get("smtp_user", "").strip()
    password = settings.get("smtp_password", "")
    sender = settings.get("smtp_from", "").strip() or username or "sistema-firma@example.com"
    use_tls = settings.get("smtp_tls", "1") == "1"

    db().execute(
        """
        INSERT INTO email_log(to_email, subject, body, status, error, created_at)
        VALUES (?, ?, ?, ?, ?, ?)
        """,
        (to_email, subject, body, "pending", "", now()),
    )
    email_id = db().execute("SELECT last_insert_rowid() AS id").fetchone()["id"]
    db().commit()

    if not host:
        db().execute(
            "UPDATE email_log SET status = ?, error = ? WHERE id = ?",
            ("not_configured", "SMTP_HOST no configurado. El correo quedo registrado en la bandeja de salida.", email_id),
        )
        db().commit()
        return False

    message = EmailMessage()
    message["From"] = sender
    message["To"] = to_email
    message["Subject"] = subject
    message.set_content(body)

    try:
        with smtplib.SMTP(host, port, timeout=20) as smtp:
            if use_tls:
                smtp.starttls()
            if username and password:
                smtp.login(username, password)
            smtp.send_message(message)
        db().execute("UPDATE email_log SET status = ?, error = ? WHERE id = ?", ("sent", "", email_id))
        db().commit()
        return True
    except Exception as exc:
        db().execute("UPDATE email_log SET status = ?, error = ? WHERE id = ?", ("failed", str(exc), email_id))
        db().commit()
        return False


def notify_document_signers(document_id: int, signer_ids: list[int]) -> tuple[int, int]:
    document = db().execute(
        """
        SELECT d.title, dt.name AS type_name
        FROM documents d JOIN document_types dt ON dt.id = d.document_type_id
        WHERE d.id = ?
        """,
        (document_id,),
    ).fetchone()
    signers = db().execute(
        f"SELECT id, name, recovery_email FROM users WHERE id IN ({','.join('?' for _ in signer_ids)})",
        signer_ids,
    ).fetchall()
    sent = 0
    failed = 0
    link = request.host_url.rstrip("/") + url_for("document_detail", document_id=document_id)
    for signer in signers:
        ok = send_email(
            signer["recovery_email"],
            f"Documento pendiente de firma: {document['title']}",
            (
                f"Hola {signer['name']},\n\n"
                f"Tienes un documento pendiente de firma.\n\n"
                f"Documento: {document['title']}\n"
                f"Tipo: {document['type_name']}\n"
                f"Enlace: {link}\n\n"
                "Ingresa con tu RUT y clave para revisar el documento y firmar."
            ),
        )
        if ok:
            sent += 1
        else:
            failed += 1
    return sent, failed


def notify_admins_password_change_alert(user: sqlite3.Row) -> tuple[int, int]:
    admins = db().execute(
        """
        SELECT recovery_email
        FROM users
        WHERE active = 1
          AND recovery_email != ''
          AND (is_admin = 1 OR is_super_admin = 1)
        """
    ).fetchall()
    role = user["role_name"] if "role_name" in user.keys() else ""
    body = (
        "Alerta de seguridad\n\n"
        "Un usuario agotó los 4 intentos para cambiar su clave desde Configuración del sistema.\n\n"
        f"ID usuario: {user['id']}\n"
        f"Nombre: {user['name']}\n"
        f"Número de registro: {user['registration_number'] or 'Sin registrar'}\n"
        f"Cargo: {role or 'Sin cargo'}\n"
        f"Correo recuperación: {user['recovery_email']}\n"
        f"IP: {request.remote_addr if request else ''}\n"
        f"Fecha: {now()}\n"
    )
    sent = 0
    failed = 0
    for admin in admins:
        ok = send_email(admin["recovery_email"], "Alerta cambio de clave fallido", body)
        if ok:
            sent += 1
        else:
            failed += 1
    return sent, failed


def init_db() -> None:
    ensure_dirs()
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    cur = conn.cursor()
    cur.executescript(
        """
        CREATE TABLE IF NOT EXISTS roles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE
        );

        CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            registration_number TEXT NOT NULL DEFAULT '',
            fire_department TEXT NOT NULL DEFAULT '',
            company TEXT NOT NULL DEFAULT '',
            name TEXT NOT NULL,
            rut TEXT NOT NULL UNIQUE,
            recovery_email TEXT NOT NULL,
            internal_code TEXT NOT NULL UNIQUE,
            password_hash TEXT NOT NULL,
            role_id INTEGER NOT NULL,
            signature_order INTEGER NOT NULL DEFAULT 99,
            is_admin INTEGER NOT NULL DEFAULT 0,
            is_super_admin INTEGER NOT NULL DEFAULT 0,
            can_sign INTEGER NOT NULL DEFAULT 1,
            can_view INTEGER NOT NULL DEFAULT 1,
            can_print INTEGER NOT NULL DEFAULT 0,
            can_download INTEGER NOT NULL DEFAULT 0,
            can_view_user_history INTEGER NOT NULL DEFAULT 0,
            active INTEGER NOT NULL DEFAULT 1,
            created_at TEXT NOT NULL,
            FOREIGN KEY(role_id) REFERENCES roles(id)
        );

        CREATE TABLE IF NOT EXISTS admin_history (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            changed_by_user_id INTEGER NOT NULL,
            target_user_id INTEGER NOT NULL,
            action TEXT NOT NULL CHECK(action IN ('grant', 'revoke')),
            created_at TEXT NOT NULL,
            FOREIGN KEY(changed_by_user_id) REFERENCES users(id),
            FOREIGN KEY(target_user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS document_types (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL UNIQUE,
            required_signatures INTEGER NOT NULL DEFAULT 1,
            visible_to_completed_roles TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS document_type_signer_roles (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            document_type_id INTEGER NOT NULL,
            role_id INTEGER NOT NULL,
            UNIQUE(document_type_id, role_id),
            FOREIGN KEY(document_type_id) REFERENCES document_types(id),
            FOREIGN KEY(role_id) REFERENCES roles(id)
        );

        CREATE TABLE IF NOT EXISTS document_type_options (
            document_type_id INTEGER PRIMARY KEY,
            next_correlative INTEGER NOT NULL DEFAULT 1,
            FOREIGN KEY(document_type_id) REFERENCES document_types(id)
        );

        CREATE TABLE IF NOT EXISTS documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            title_key TEXT NOT NULL DEFAULT '',
            document_type_id INTEGER NOT NULL,
            original_filename TEXT NOT NULL,
            stored_path TEXT NOT NULL,
            signed_path TEXT,
            status TEXT NOT NULL DEFAULT 'pending' CHECK(status IN ('pending', 'observed', 'closed', 'annulled')),
            uploaded_by_user_id INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            closed_at TEXT,
            FOREIGN KEY(document_type_id) REFERENCES document_types(id),
            FOREIGN KEY(uploaded_by_user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS document_required_signers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            document_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            UNIQUE(document_id, user_id),
            FOREIGN KEY(document_id) REFERENCES documents(id),
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS signatures (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            document_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            signature_code TEXT NOT NULL,
            signed_at TEXT NOT NULL,
            UNIQUE(document_id, user_id),
            FOREIGN KEY(document_id) REFERENCES documents(id),
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS document_observations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            document_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            note TEXT NOT NULL,
            created_at TEXT NOT NULL,
            response TEXT,
            responded_by_user_id INTEGER,
            responded_at TEXT,
            FOREIGN KEY(document_id) REFERENCES documents(id),
            FOREIGN KEY(user_id) REFERENCES users(id),
            FOREIGN KEY(responded_by_user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS acta_documents (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            title TEXT NOT NULL,
            document_type_id INTEGER NOT NULL,
            acta_date TEXT NOT NULL,
            correlative TEXT NOT NULL,
            body TEXT NOT NULL,
            status TEXT NOT NULL DEFAULT 'draft' CHECK(status IN ('draft', 'closed')),
            pdf_path TEXT,
            created_by_user_id INTEGER NOT NULL,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL,
            closed_at TEXT,
            FOREIGN KEY(document_type_id) REFERENCES document_types(id),
            FOREIGN KEY(created_by_user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS acta_signature_links (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            acta_id INTEGER NOT NULL,
            document_id INTEGER NOT NULL,
            resend_reason TEXT NOT NULL DEFAULT '',
            sent_by_user_id INTEGER NOT NULL,
            sent_at TEXT NOT NULL,
            FOREIGN KEY(acta_id) REFERENCES acta_documents(id),
            FOREIGN KEY(document_id) REFERENCES documents(id),
            FOREIGN KEY(sent_by_user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            actor_user_id INTEGER,
            target_user_id INTEGER,
            document_id INTEGER,
            action TEXT NOT NULL,
            details TEXT NOT NULL,
            ip_address TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(actor_user_id) REFERENCES users(id),
            FOREIGN KEY(target_user_id) REFERENCES users(id),
            FOREIGN KEY(document_id) REFERENCES documents(id)
        );

        CREATE TABLE IF NOT EXISTS password_recovery (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER NOT NULL,
            token TEXT NOT NULL UNIQUE,
            used INTEGER NOT NULL DEFAULT 0,
            created_at TEXT NOT NULL,
            used_at TEXT,
            FOREIGN KEY(user_id) REFERENCES users(id)
        );

        CREATE TABLE IF NOT EXISTS email_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            to_email TEXT NOT NULL,
            subject TEXT NOT NULL,
            body TEXT NOT NULL,
            status TEXT NOT NULL,
            error TEXT NOT NULL DEFAULT '',
            created_at TEXT NOT NULL
        );

        CREATE TABLE IF NOT EXISTS system_settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        );
        """
    )

    cur.execute("SELECT sql FROM sqlite_master WHERE type = 'table' AND name = 'documents'")
    documents_sql = cur.fetchone()["sql"]
    if "observed" not in documents_sql or "annulled" not in documents_sql:
        cur.execute("PRAGMA foreign_keys = OFF")
        cur.execute("ALTER TABLE documents RENAME TO documents_old")
        cur.execute(
            """
            CREATE TABLE documents (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                title TEXT NOT NULL,
                title_key TEXT NOT NULL DEFAULT '',
                document_type_id INTEGER NOT NULL,
                original_filename TEXT NOT NULL,
                stored_path TEXT NOT NULL,
                signed_path TEXT,
                status TEXT NOT NULL DEFAULT 'pending' CHECK(status IN ('pending', 'observed', 'closed', 'annulled')),
                uploaded_by_user_id INTEGER NOT NULL,
                created_at TEXT NOT NULL,
                closed_at TEXT,
                FOREIGN KEY(document_type_id) REFERENCES document_types(id),
                FOREIGN KEY(uploaded_by_user_id) REFERENCES users(id)
            )
            """
        )
        cur.execute(
            """
            INSERT INTO documents(id, title, title_key, document_type_id, original_filename, stored_path, signed_path, status, uploaded_by_user_id, created_at, closed_at)
            SELECT id, title, '', document_type_id, original_filename, stored_path, signed_path, status, uploaded_by_user_id, created_at, closed_at
            FROM documents_old
            """
        )
        cur.execute("DROP TABLE documents_old")
        cur.execute("PRAGMA foreign_keys = ON")

    def repair_documents_foreign_key(table_name: str, create_sql: str, columns: list[str]) -> None:
        cur.execute("SELECT sql FROM sqlite_master WHERE type = 'table' AND name = ?", (table_name,))
        row = cur.fetchone()
        if not row or "documents_old" not in row["sql"]:
            return
        temp_name = f"{table_name}_old_fk"
        cur.execute("PRAGMA foreign_keys = OFF")
        cur.execute(f"ALTER TABLE {table_name} RENAME TO {temp_name}")
        cur.execute(create_sql)
        column_list = ", ".join(columns)
        cur.execute(f"INSERT INTO {table_name}({column_list}) SELECT {column_list} FROM {temp_name}")
        cur.execute(f"DROP TABLE {temp_name}")
        cur.execute("PRAGMA foreign_keys = ON")

    repair_documents_foreign_key(
        "document_required_signers",
        """
        CREATE TABLE document_required_signers (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            document_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            UNIQUE(document_id, user_id),
            FOREIGN KEY(document_id) REFERENCES documents(id),
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
        """,
        ["id", "document_id", "user_id"],
    )
    repair_documents_foreign_key(
        "signatures",
        """
        CREATE TABLE signatures (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            document_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            signature_code TEXT NOT NULL,
            signed_at TEXT NOT NULL,
            UNIQUE(document_id, user_id),
            FOREIGN KEY(document_id) REFERENCES documents(id),
            FOREIGN KEY(user_id) REFERENCES users(id)
        )
        """,
        ["id", "document_id", "user_id", "signature_code", "signed_at"],
    )
    repair_documents_foreign_key(
        "audit_log",
        """
        CREATE TABLE audit_log (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            actor_user_id INTEGER,
            target_user_id INTEGER,
            document_id INTEGER,
            action TEXT NOT NULL,
            details TEXT NOT NULL,
            ip_address TEXT,
            created_at TEXT NOT NULL,
            FOREIGN KEY(actor_user_id) REFERENCES users(id),
            FOREIGN KEY(target_user_id) REFERENCES users(id),
            FOREIGN KEY(document_id) REFERENCES documents(id)
        )
        """,
        ["id", "actor_user_id", "target_user_id", "document_id", "action", "details", "ip_address", "created_at"],
    )
    repair_documents_foreign_key(
        "document_observations",
        """
        CREATE TABLE document_observations (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            document_id INTEGER NOT NULL,
            user_id INTEGER NOT NULL,
            note TEXT NOT NULL,
            created_at TEXT NOT NULL,
            response TEXT,
            responded_by_user_id INTEGER,
            responded_at TEXT,
            FOREIGN KEY(document_id) REFERENCES documents(id),
            FOREIGN KEY(user_id) REFERENCES users(id),
            FOREIGN KEY(responded_by_user_id) REFERENCES users(id)
        )
        """,
        ["id", "document_id", "user_id", "note", "created_at"],
    )

    cur.execute("PRAGMA table_info(documents)")
    document_columns = {row["name"] for row in cur.fetchall()}
    if "title_key" not in document_columns:
        cur.execute("ALTER TABLE documents ADD COLUMN title_key TEXT NOT NULL DEFAULT ''")
    cur.execute("SELECT id, title FROM documents WHERE title_key = '' OR title_key IS NULL")
    for document_row in cur.fetchall():
        cur.execute(
            "UPDATE documents SET title_key = ? WHERE id = ?",
            (normalize_document_title(document_row["title"]), document_row["id"]),
        )

    cur.execute("PRAGMA table_info(users)")
    user_columns = {row["name"] for row in cur.fetchall()}
    if "registration_number" not in user_columns:
        cur.execute("ALTER TABLE users ADD COLUMN registration_number TEXT NOT NULL DEFAULT ''")
    if "fire_department" not in user_columns:
        cur.execute("ALTER TABLE users ADD COLUMN fire_department TEXT NOT NULL DEFAULT ''")
    if "company" not in user_columns:
        cur.execute("ALTER TABLE users ADD COLUMN company TEXT NOT NULL DEFAULT ''")
    if "signature_order" not in user_columns:
        cur.execute("ALTER TABLE users ADD COLUMN signature_order INTEGER NOT NULL DEFAULT 99")

    cur.execute("PRAGMA table_info(document_observations)")
    observation_columns = {row["name"] for row in cur.fetchall()}
    if "response" not in observation_columns:
        cur.execute("ALTER TABLE document_observations ADD COLUMN response TEXT")
    if "responded_by_user_id" not in observation_columns:
        cur.execute("ALTER TABLE document_observations ADD COLUMN responded_by_user_id INTEGER")
    if "responded_at" not in observation_columns:
        cur.execute("ALTER TABLE document_observations ADD COLUMN responded_at TEXT")

    roles = [
        "Bombero",
        "Capitan",
        "Director",
        "Secretario",
        "Tesorero",
        "Intendente",
        "Consejero de Administración",
        "Consejero de Disciplina",
    ]
    for role in roles:
        cur.execute("INSERT OR IGNORE INTO roles(name) VALUES (?)", (role,))

    cur.execute("SELECT id FROM roles WHERE name = ?", ("Secretario",))
    secretary_role_id = cur.fetchone()["id"]
    cur.execute("SELECT COUNT(*) AS total FROM users")
    if cur.fetchone()["total"] == 0:
        cur.execute(
            """
            INSERT INTO users(registration_number, name, rut, recovery_email, internal_code, password_hash, role_id,
                              is_admin, is_super_admin, can_sign, can_view, can_print,
                              can_download, can_view_user_history, active, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, 1, 1, 1, 1, 1, 1, 1, 1, ?)
            """,
            (
                "SUPER",
                "Super Administrador",
                DEFAULT_ADMIN_RUT,
                "admin@example.com",
                "SUPER",
                generate_password_hash(DEFAULT_PASSWORD),
                secretary_role_id,
                now(),
            ),
        )
    else:
        cur.execute(
            "UPDATE users SET rut = ? WHERE rut = 'SUPER-ADMIN' AND is_super_admin = 1",
            (DEFAULT_ADMIN_RUT,),
        )

    def role_id(name: str) -> int:
        cur.execute("SELECT id FROM roles WHERE name = ?", (name,))
        return cur.fetchone()["id"]

    default_types = [
        (
            "Acta de Reuniones de Compania",
            2,
            ["Bombero", "Capitan", "Director", "Secretario", "Tesorero", "Intendente", "Consejero de Administración", "Consejero de Disciplina"],
            ["Director", "Secretario"],
        ),
        (
            "Acta Consejo de Administración",
            11,
            ["Consejero de Administración", "Capitan", "Director", "Secretario", "Tesorero", "Intendente"],
            ["Consejero de Administración", "Capitan", "Director", "Secretario", "Tesorero", "Intendente"],
        ),
        (
            "Consejo de Administración Disciplinario",
            10,
            ["Consejero de Administración", "Capitan", "Secretario", "Tesorero", "Intendente"],
            ["Consejero de Administración", "Capitan", "Secretario", "Tesorero", "Intendente"],
        ),
        (
            "Consejo de Disciplina",
            10,
            ["Consejero de Disciplina", "Director", "Secretario", "Tesorero"],
            ["Consejero de Disciplina", "Director", "Secretario", "Tesorero"],
        ),
    ]
    for name, required, visible_roles, signer_roles in default_types:
        cur.execute(
            """
            INSERT OR IGNORE INTO document_types(name, required_signatures, visible_to_completed_roles, created_at)
            VALUES (?, ?, ?, ?)
            """,
            (name, required, ",".join(visible_roles), now()),
        )
        cur.execute("SELECT id FROM document_types WHERE name = ?", (name,))
        doc_type_id = cur.fetchone()["id"]
        for signer_role in signer_roles:
            cur.execute(
                "INSERT OR IGNORE INTO document_type_signer_roles(document_type_id, role_id) VALUES (?, ?)",
                (doc_type_id, role_id(signer_role)),
            )
        cur.execute(
            "INSERT OR IGNORE INTO document_type_options(document_type_id, next_correlative) VALUES (?, 1)",
            (doc_type_id,),
        )

    for key, value in DEFAULT_SYSTEM_SETTINGS.items():
        cur.execute("INSERT OR IGNORE INTO system_settings(key, value) VALUES (?, ?)", (key, value))

    conn.commit()
    conn.close()


def current_user() -> sqlite3.Row | None:
    if "user_id" not in session:
        return None
    return db().execute(
        """
        SELECT u.*, CASE WHEN u.is_super_admin = 1 THEN '' ELSE r.name END AS role_name
        FROM users u LEFT JOIN roles r ON r.id = u.role_id
        WHERE u.id = ? AND u.active = 1
        """,
        (session["user_id"],),
    ).fetchone()


@app.before_request
def load_user() -> None:
    g.user = current_user()


def login_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not g.user:
            return redirect(url_for("login"))
        return view(*args, **kwargs)

    return wrapped


def admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not g.user or not (g.user["is_admin"] or g.user["is_super_admin"]):
            abort(403)
        return view(*args, **kwargs)

    return wrapped


def super_admin_required(view):
    @wraps(view)
    def wrapped(*args, **kwargs):
        if not g.user or not g.user["is_super_admin"]:
            abort(403)
        return view(*args, **kwargs)

    return wrapped


def can_manage_admins() -> bool:
    return bool(g.user and (g.user["is_super_admin"] or g.user["is_admin"]))


def random_internal_code() -> str:
    return secrets.token_hex(4).upper()


def random_signature_code() -> str:
    return "".join(str(secrets.randbelow(10)) for _ in range(10))


def random_recovery_code() -> str:
    return "".join(str(secrets.randbelow(10)) for _ in range(6))


def normalize_document_title(title: str) -> str:
    without_accents = "".join(
        char for char in unicodedata.normalize("NFKD", title.strip().lower())
        if not unicodedata.combining(char)
    )
    return re.sub(r"[^a-z0-9]+", " ", without_accents).strip()


def strip_accents(value: str) -> str:
    return "".join(
        char for char in unicodedata.normalize("NFKD", value)
        if not unicodedata.combining(char)
    )


def normalize_rut(rut: str) -> str:
    clean = rut.strip().replace(".", "").replace("-", "").upper()
    if len(clean) < 2 or not clean[:-1].isdigit() or not re.fullmatch(r"[\dK]", clean[-1]):
        return rut.strip()
    number = f"{int(clean[:-1]):,}".replace(",", ".")
    return f"{number}-{clean[-1]}"


def valid_rut(rut: str) -> bool:
    normalized = normalize_rut(rut)
    if not RUT_PATTERN.fullmatch(normalized):
        return False
    body, verifier = normalized.replace(".", "").split("-")
    factors = [2, 3, 4, 5, 6, 7]
    total = 0
    for index, digit in enumerate(reversed(body)):
        total += int(digit) * factors[index % len(factors)]
    expected_value = 11 - (total % 11)
    if expected_value == 11:
        expected = "0"
    elif expected_value == 10:
        expected = "K"
    else:
        expected = str(expected_value)
    return verifier.upper() == expected


def valid_password(password: str) -> bool:
    return 7 <= len(password) <= 15 and not any(char.isspace() for char in password)


def password_help() -> str:
    return "Clave no cumple: debe tener 7 caracteres mínimo y 15 máximo; puede tener símbolos."


def excel_bool(value: object, default: bool = False) -> bool:
    if value is None or str(value).strip() == "":
        return default
    normalized = strip_accents(str(value).strip().lower())
    return normalized in {"1", "si", "s", "true", "verdadero", "x", "activo", "activa"}


def excel_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if text.endswith(".0") and text[:-2].isdigit():
        return text[:-2]
    return text


def valid_reason_note(reason: str) -> bool:
    return 10 <= len(reason) <= 80 and bool(re.fullmatch(r"[A-Za-z0-9ÁÉÍÓÚÜÑáéíóúüñ ]+", reason))


def valid_observation_text(text: str) -> bool:
    return 10 <= len(text) <= 300


def document_status_label(status: str) -> str:
    return {
        "pending": "Pendiente",
        "observed": "Observado",
        "closed": "Cerrado",
        "annulled": "Anulado",
    }.get(status, status)


def is_signed_pending_close(status: str, required_count: int = 0, signed_count: int = 0) -> bool:
    return status == "pending" and required_count > 0 and required_count == signed_count


def document_status_display(status: str, observations_count: int, required_count: int = 0, signed_count: int = 0) -> str:
    has_observations = observations_count > 0
    if is_signed_pending_close(status, required_count, signed_count):
        return "Firmado pendiente de cierre"
    if status == "pending" and has_observations:
        return "Pendiente con observaciónes"
    if status == "closed" and has_observations:
        return "Cerrado con observaciónes"
    return document_status_label(status)


def document_status_class(status: str, observations_count: int, required_count: int = 0, signed_count: int = 0) -> str:
    has_observations = observations_count > 0
    if is_signed_pending_close(status, required_count, signed_count):
        return "status-signed-pending-close"
    if status == "pending" and not has_observations:
        return "status-pending"
    if status in {"pending", "observed"} and has_observations:
        return "status-observed"
    if status == "closed" and has_observations:
        return "status-closed-observed"
    if status == "closed":
        return "status-closed"
    if status == "annulled":
        return "status-annulled"
    return ""


def acta_signature_status(row: sqlite3.Row) -> str:
    if not row["signature_document_id"]:
        return "No enviada a firma"
    status = document_status_display(
        row["signature_status"],
        row["signature_observations_count"] or 0,
        row["signature_required_count"] or 0,
        row["signature_signed_count"] or 0,
    )
    return f"Enviada a firma: {status}"


def document_filter_labels(selected_type: str, selected_status: str) -> tuple[str, str]:
    type_label = "Todos"
    if selected_type.isdigit():
        row = db().execute("SELECT name FROM document_types WHERE id = ?", (int(selected_type),)).fetchone()
        if row:
            type_label = row["name"]
    status_label = {
        "": "Todos",
        "pending": "Pendiente",
        "pending_and_observed": "Pendientes y observados",
        "signed_pending_close": "Firmado pendiente de cierre",
        "pending_observed": "Pendiente con observaciónes",
        "observed": "Observado",
        "closed": "Cerrado",
        "closed_observed": "Cerrado con observaciónes",
        "annulled": "Anulado",
    }.get(selected_status, "Todos")
    return type_label, status_label


def eligible_signers_for_document_type(document_type_id: int, exclude_document_id: int | None = None) -> list[sqlite3.Row]:
    params: list[object] = [document_type_id]
    exclude_sql = ""
    if exclude_document_id is not None:
        exclude_sql = """
          AND NOT EXISTS(
            SELECT 1 FROM document_required_signers rs
            WHERE rs.document_id = ? AND rs.user_id = u.id
          )
        """
        params.append(exclude_document_id)
    return db().execute(
        f"""
        SELECT u.*, r.name AS role_name
        FROM users u
        JOIN roles r ON r.id = u.role_id
        WHERE u.active = 1
          AND u.can_sign = 1
          AND u.is_super_admin = 0
          AND EXISTS(
            SELECT 1 FROM document_type_signer_roles dtsr
            WHERE dtsr.document_type_id = ? AND dtsr.role_id = u.role_id
          )
          {exclude_sql}
        ORDER BY r.name, u.name
        """,
        params,
    ).fetchall()


def acta_related_users() -> list[sqlite3.Row]:
    return db().execute(
        """
        SELECT DISTINCT u.id, u.name, u.signature_order, r.name AS role_name, dtsr.document_type_id
        FROM users u
        JOIN roles r ON r.id = u.role_id
        JOIN document_type_signer_roles dtsr ON dtsr.role_id = u.role_id
        WHERE u.active = 1
          AND u.can_sign = 1
          AND u.is_super_admin = 0
        ORDER BY u.signature_order, r.name, u.name
        """
    ).fetchall()


def acta_title(acta_date: str, correlative: str, type_name: str) -> str:
    clean_date = " ".join(acta_date.split("-"))
    clean_correlative = correlative.strip()
    return "{} N{} {}".format(clean_date, "\u00b0", f"{clean_correlative} {type_name}".strip()).strip()


def transcribe_audio_file(audio_path: Path) -> tuple[bool, str]:
    model_name = os.environ.get("ESIGNUM_WHISPER_MODEL", "base")
    try:
        from faster_whisper import WhisperModel

        model = WhisperModel(model_name, device="cpu", compute_type="int8")
        segments, _ = model.transcribe(str(audio_path), language="es")
        text = " ".join(segment.text.strip() for segment in segments if segment.text.strip()).strip()
        return bool(text), text or "No se detecto texto en el audio."
    except ModuleNotFoundError:
        pass
    except Exception as exc:
        return False, f"No se pudo transcribir con faster-whisper: {exc}"

    try:
        import whisper

        model = whisper.load_model(model_name)
        result = model.transcribe(str(audio_path), language="es")
        text = str(result.get("text", "")).strip()
        return bool(text), text or "No se detecto texto en el audio."
    except ModuleNotFoundError:
        pass
    except Exception as exc:
        return False, f"No se pudo transcribir con whisper: {exc}"

    whisper_cli = shutil.which("whisper")
    if whisper_cli:
        try:
            output_dir = AUDIO_DIR / f"transcription-{secrets.token_hex(6)}"
            output_dir.mkdir(exist_ok=True)
            subprocess.run(
                [
                    whisper_cli,
                    str(audio_path),
                    "--language",
                    "Spanish",
                    "--model",
                    model_name,
                    "--output_format",
                    "txt",
                    "--output_dir",
                    str(output_dir),
                ],
                check=True,
                capture_output=True,
                text=True,
                timeout=900,
            )
            text_files = list(output_dir.glob("*.txt"))
            text = text_files[0].read_text(encoding="utf-8").strip() if text_files else ""
            return bool(text), text or "No se detecto texto en el audio."
        except Exception as exc:
            return False, f"No se pudo transcribir con Whisper CLI: {exc}"

    return (
        False,
        "No hay motor de transcripción instalado. Instala faster-whisper o whisper en el entorno del sistema para usar esta función.",
    )


def ensure_document_type_options() -> None:
    db().execute(
        """
        CREATE TABLE IF NOT EXISTS document_type_options (
            document_type_id INTEGER PRIMARY KEY,
            next_correlative INTEGER NOT NULL DEFAULT 1,
            FOREIGN KEY(document_type_id) REFERENCES document_types(id)
        )
        """
    )
    db().execute(
        """
        INSERT OR IGNORE INTO document_type_options(document_type_id, next_correlative)
        SELECT id, 1 FROM document_types
        """
    )
    db().commit()


def document_type_options_map() -> dict[int, int]:
    ensure_document_type_options()
    rows = db().execute("SELECT document_type_id, next_correlative FROM document_type_options").fetchall()
    return {int(row["document_type_id"]): int(row["next_correlative"]) for row in rows}


def advance_document_correlative(document_type_id: int, used_correlative: str) -> None:
    match = re.search(r"\d+", used_correlative)
    if not match:
        return
    used_number = int(match.group())
    ensure_document_type_options()
    current = db().execute(
        "SELECT next_correlative FROM document_type_options WHERE document_type_id = ?",
        (document_type_id,),
    ).fetchone()
    next_number = int(current["next_correlative"]) if current else 1
    if used_number >= next_number:
        db().execute(
            """
            INSERT INTO document_type_options(document_type_id, next_correlative)
            VALUES (?, ?)
            ON CONFLICT(document_type_id) DO UPDATE SET next_correlative = excluded.next_correlative
            """,
            (document_type_id, used_number + 1),
        )


def find_user_by_rut(rut: str, active_only: bool = False) -> sqlite3.Row | None:
    normalized = normalize_rut(rut)
    compact = normalized.replace(".", "")
    raw = rut.strip().upper()
    where = "rut IN (?, ?, ?)"
    params: list[object] = [normalized, compact, raw]
    if active_only:
        where += " AND active = 1"
    return db().execute(f"SELECT * FROM users WHERE {where}", params).fetchone()


def ensure_system_settings_table() -> None:
    db().execute(
        """
        CREATE TABLE IF NOT EXISTS system_settings (
            key TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
        """
    )
    for key, value in DEFAULT_SYSTEM_SETTINGS.items():
        db().execute("INSERT OR IGNORE INTO system_settings(key, value) VALUES (?, ?)", (key, value))
    db().commit()


def system_settings() -> dict[str, str]:
    settings = dict(DEFAULT_SYSTEM_SETTINGS)
    ensure_system_settings_table()
    rows = db().execute("SELECT key, value FROM system_settings").fetchall()
    settings.update({row["key"]: row["value"] for row in rows})
    if settings.get("menu_position") not in MENU_POSITIONS:
        settings["menu_position"] = DEFAULT_SYSTEM_SETTINGS["menu_position"]
    if settings.get("button_color") not in BUTTON_COLORS:
        settings["button_color"] = DEFAULT_SYSTEM_SETTINGS["button_color"]
    if settings.get("button_text_color") not in BUTTON_TEXT_COLORS:
        settings["button_text_color"] = DEFAULT_SYSTEM_SETTINGS["button_text_color"]
    if settings.get("field_border_color") not in FIELD_BORDER_COLORS:
        settings["field_border_color"] = DEFAULT_SYSTEM_SETTINGS["field_border_color"]
    if settings.get("page_background") not in PAGE_BACKGROUNDS:
        settings["page_background"] = DEFAULT_SYSTEM_SETTINGS["page_background"]
    if settings.get("screen_image_position") not in SCREEN_IMAGE_POSITIONS:
        settings["screen_image_position"] = DEFAULT_SYSTEM_SETTINGS["screen_image_position"]
    if settings.get("acta_title_alignment") not in ACTA_TITLE_ALIGNMENTS:
        settings["acta_title_alignment"] = DEFAULT_SYSTEM_SETTINGS["acta_title_alignment"]
    if settings.get("acta_image_position") not in SCREEN_IMAGE_POSITIONS:
        settings["acta_image_position"] = DEFAULT_SYSTEM_SETTINGS["acta_image_position"]
    if settings.get("acta_font_family") not in ACTA_FONT_FAMILIES:
        settings["acta_font_family"] = DEFAULT_SYSTEM_SETTINGS["acta_font_family"]
    if settings.get("acta_image_enabled") not in {"0", "1"}:
        settings["acta_image_enabled"] = DEFAULT_SYSTEM_SETTINGS["acta_image_enabled"]
    if settings.get("acta_watermark_enabled") not in {"0", "1"}:
        settings["acta_watermark_enabled"] = DEFAULT_SYSTEM_SETTINGS["acta_watermark_enabled"]
    for position_key, default_value in (
        ("screen_image_position", "top_right"),
        ("screen_image_2_position", "top_left"),
        ("acta_image_position", "top_right"),
    ):
        if settings.get(position_key) not in SCREEN_IMAGE_POSITIONS:
            settings[position_key] = default_value
    for size_key, default_value in (
        ("screen_image_size", "180"),
        ("screen_image_2_size", "180"),
        ("acta_body_font_size", "9"),
        ("acta_title_font_size", "13"),
        ("acta_watermark_size", "260"),
        ("acta_watermark_opacity", "15"),
    ):
        if not str(settings.get(size_key, "")).isdigit():
            settings[size_key] = default_value
        if size_key == "acta_watermark_opacity":
            upper_limit = 100
            lower_limit = 1
        else:
            upper_limit = 600 if "image" in size_key or "watermark" in size_key else 28
            lower_limit = 40 if "image" in size_key or "watermark" in size_key else 7
        settings[size_key] = str(min(upper_limit, max(lower_limit, int(settings[size_key]))))
    screen_image_path = settings.get("screen_image_path", "")
    settings["screen_image_url"] = url_for("system_screen_image") if screen_image_path and Path(screen_image_path).exists() else ""
    for image_index in (2,):
        image_path = settings.get(f"screen_image_{image_index}_path", "")
        settings[f"screen_image_{image_index}_url"] = (
            url_for("system_extra_screen_image", image_index=image_index)
            if image_path and Path(image_path).exists()
            else ""
        )
    acta_image_path = settings.get("acta_image_path", "")
    settings["acta_image_url"] = url_for("acta_config_image") if acta_image_path and Path(acta_image_path).exists() else ""
    if not settings.get("acta_image_filename") and acta_image_path:
        settings["acta_image_filename"] = Path(acta_image_path).name
    watermark_path = settings.get("acta_watermark_image_path", "")
    settings["acta_watermark_image_url"] = url_for("acta_watermark_config_image") if watermark_path and Path(watermark_path).exists() else ""
    if not settings.get("acta_watermark_image_filename") and watermark_path:
        settings["acta_watermark_image_filename"] = Path(watermark_path).name
    settings["button_color_hex"] = BUTTON_COLORS[settings["button_color"]][1]
    settings["button_color_label"] = BUTTON_COLORS[settings["button_color"]][0]
    settings["button_text_color_hex"] = BUTTON_TEXT_COLORS[settings["button_text_color"]][1]
    settings["button_text_color_label"] = BUTTON_TEXT_COLORS[settings["button_text_color"]][0]
    settings["field_border_color_hex"] = FIELD_BORDER_COLORS[settings["field_border_color"]][1]
    settings["field_border_color_label"] = FIELD_BORDER_COLORS[settings["field_border_color"]][0]
    settings["page_background_hex"] = PAGE_BACKGROUNDS[settings["page_background"]][1]
    settings["page_background_label"] = PAGE_BACKGROUNDS[settings["page_background"]][0]
    settings["menu_position_label"] = MENU_POSITIONS[settings["menu_position"]]
    settings["screen_image_position_label"] = SCREEN_IMAGE_POSITIONS[settings["screen_image_position"]]
    settings["screen_image_2_position_label"] = SCREEN_IMAGE_POSITIONS[settings["screen_image_2_position"]]
    settings["acta_title_alignment_label"] = ACTA_TITLE_ALIGNMENTS[settings["acta_title_alignment"]]
    settings["acta_image_position_label"] = SCREEN_IMAGE_POSITIONS[settings["acta_image_position"]]
    return settings


def html_page(title: str, content: str) -> str:
    settings = system_settings()
    return render_template_string(
        """
        <!doctype html>
        <html lang="es">
        <head>
          <meta charset="utf-8">
          <meta name="viewport" content="width=device-width, initial-scale=1">
          <title>{{ title }}</title>
          <style>
            :root { --bg:{{ settings.page_background_hex }}; --ink:#16202a; --muted:#5b6775; --line:#d8dee8; --field-border:{{ settings.field_border_color_hex }}; --brand:{{ settings.button_color_hex }}; --button-text:{{ settings.button_text_color_hex }}; --danger:#a93636; }
            * { box-sizing:border-box; }
            body { margin:0; font-family:Arial, sans-serif; background:var(--bg); color:var(--ink); }
            header { background:#fff; border-bottom:1px solid var(--line); padding:14px 24px; display:flex; gap:18px; align-items:center; justify-content:space-between; }
            nav a { margin-right:12px; color:var(--brand); text-decoration:none; font-weight:700; }
            nav .user-name { color:var(--muted); font-weight:700; }
            main { max-width:1180px; margin:18px auto; padding:0 18px; position:relative; z-index:1; }
            .screen-image { position:fixed; z-index:5; width:var(--screen-image-size, 180px); max-width:42vw; max-height:42vh; object-fit:contain; opacity:.95; pointer-events:none; }
            .screen-image.top_left { top:82px; left:24px; }
            .screen-image.top_center { top:82px; left:50%; transform:translateX(-50%); }
            .screen-image.top_right { top:82px; right:24px; }
            .screen-image.center_left { top:50%; left:24px; transform:translateY(-50%); }
            .screen-image.center_right { top:50%; right:24px; transform:translateY(-50%); }
            .screen-image.bottom_left { bottom:24px; left:24px; }
            .screen-image.bottom_center { bottom:24px; left:50%; transform:translateX(-50%); }
            .screen-image.bottom_right { bottom:24px; right:24px; }
            .screen-image.center { top:50%; left:50%; transform:translate(-50%, -50%); max-width:260px; max-height:260px; }
            body.menu-bottom { padding-bottom:74px; }
            body.menu-bottom header { position:fixed; left:0; right:0; bottom:0; top:auto; border-top:1px solid var(--line); border-bottom:0; z-index:10; }
            body.menu-left, body.menu-right { min-height:100vh; }
            body.menu-left header, body.menu-right header { position:fixed; top:0; bottom:0; width:250px; align-items:flex-start; justify-content:flex-start; flex-direction:column; z-index:10; border-bottom:0; overflow:auto; }
            body.menu-left header { left:0; border-right:1px solid var(--line); }
            body.menu-right header { right:0; border-left:1px solid var(--line); }
            body.menu-left nav, body.menu-right nav { display:flex; flex-direction:column; gap:10px; width:100%; }
            body.menu-left nav a, body.menu-right nav a { margin-right:0; padding:9px 10px; border-radius:6px; }
            body.menu-left nav a:hover, body.menu-right nav a:hover { background:#eef4f8; }
            body.menu-left main { margin-left:276px; }
            body.menu-right main { margin-right:276px; }
            h1 { font-size:24px; margin:0 0 12px; }
            h2 { font-size:18px; margin:18px 0 9px; }
            .panel { background:#fff; border:1px solid var(--line); border-radius:8px; padding:12px; margin-bottom:12px; }
            table { width:100%; border-collapse:collapse; background:#fff; }
            th, td { border-bottom:1px solid var(--line); padding:7px 9px; text-align:left; vertical-align:top; }
            th { color:var(--muted); font-size:13px; }
            input, select, button { font:inherit; padding:7px 9px; border:1px solid var(--field-border); border-radius:6px; background:#fff; }
            input[type="checkbox"] { appearance:none; width:16px; height:16px; min-width:16px; padding:0; border:2px solid var(--danger); border-radius:3px; background:#fff; cursor:pointer; vertical-align:middle; }
            input[type="checkbox"]:checked { background:var(--brand); border-color:var(--brand); position:relative; }
            input[type="checkbox"]:checked::after { content:""; display:block; width:4px; height:8px; border:solid var(--button-text); border-width:0 2px 2px 0; transform:rotate(45deg); margin:1px 0 0 4px; }
            input[type="checkbox"]:disabled { opacity:.65; cursor:not-allowed; }
            label { display:block; margin:7px 0 3px; font-weight:700; }
            button, .btn { background:var(--brand); color:var(--button-text); border:0; text-decoration:none; display:inline-block; padding:7px 10px; border-radius:6px; cursor:pointer; }
            .danger { background:var(--danger); }
            .muted { color:var(--muted); }
            .grid { display:grid; grid-template-columns:repeat(auto-fit, minmax(240px, 1fr)); gap:10px; }
            .config-fields-grid { display:grid; grid-template-columns:repeat(2, minmax(220px, 1fr)); gap:0 8px; align-items:end; }
            .config-fields-grid input, .config-fields-grid select { width:30ch; max-width:100%; }
            .config-fields-grid .full-row { grid-column:1 / -1; }
            .config-inline-pair { display:flex; align-items:end; gap:8px; flex-wrap:wrap; }
            .config-inline-pair > div { min-width:0; }
            .config-align-top { align-self:start; }
            .title-line-input { width:70ch !important; max-width:100%; }
            .document-options-modal input[type="checkbox"] { width:12px; height:12px; min-width:12px; }
            .document-options-modal input[type="checkbox"]:checked::after { width:3px; height:6px; margin:0 0 0 3px; }
            .image-size-input { width:7ch !important; }
            .correlative-input { width:10ch !important; }
            .correlative-row { display:flex; align-items:center; gap:8px; margin:7px 0 3px; }
            .correlative-row label { margin:0; font-weight:700; }
            .inline-actions { display:flex; align-items:end; gap:8px; flex-wrap:wrap; }
            .push-right { margin-left:auto; }
            .inline-sign { display:flex; align-items:end; gap:8px; flex-wrap:wrap; margin:0; }
            .inline-sign label { margin:0; }
            .inline-sign input { width:15ch; }
            .detail-actions { margin-top:12px; }
            .acta-user { display:flex; align-items:center; gap:8px; flex-wrap:wrap; }
            .acta-user .acta-excuse-control, .acta-user .acta-include-control { font-weight:400; display:inline-flex; align-items:center; gap:4px; }
            .acta-header-fields { display:grid; grid-template-columns:minmax(50ch, 50ch) 170px 9ch; gap:10px; align-items:end; }
            .acta-header-fields input, .acta-header-fields select { width:100%; }
            .acta-text-actions { display:flex; align-items:center; gap:8px; flex-wrap:wrap; }
            .acta-text-actions .close-document-action { margin-left:auto; }
            .acta-audio-tools { display:flex; align-items:center; gap:8px; flex-wrap:wrap; margin:8px 0 10px; }
            .acta-audio-tools input[type=file] { max-width:330px; }
            .flash { background:#fff6d8; border:1px solid #e0c36f; padding:10px; border-radius:6px; margin-bottom:12px; }
            .status { font-weight:700; }
            .status-pill { display:inline-block; padding:5px 9px; border-radius:6px; color:#fff; font-weight:700; }
            .status-pending { background:#16803c; }
            .status-signed-pending-close { background:#0f766e; }
            .status-observed { background:#d97800; }
            .status-closed-observed { background:#a93636; }
            .status-closed { background:#1f5fbf; }
            .status-annulled { background:#5b6775; }
            .acta-sent-row td { background:#ffe4e4; color:#7a1f1f; }
            .modal-backdrop { position:fixed; inset:0; background:rgba(0,0,0,.45); display:none; align-items:center; justify-content:center; padding:18px; z-index:20; }
            .modal-backdrop.active { display:flex; }
            .modal { background:#fff; border-radius:8px; padding:18px; width:min(560px, 100%); border:1px solid var(--line); }
            .modal-scroll { max-height:92vh; overflow:hidden; display:flex; flex-direction:column; }
            .modal-scroll form { min-height:0; overflow:hidden; display:flex; flex-direction:column; }
            .modal-scroll-body { overflow:auto; padding-right:6px; }
            .modal-actions-sticky { position:sticky; bottom:0; background:#fff; border-top:1px solid var(--line); padding-top:10px; margin:10px 0 0; }
            .user-modal { width:min(980px,96vw); }
            .user-form-grid { display:grid; grid-template-columns:minmax(280px, 1fr) minmax(260px, .85fr); gap:18px; align-items:start; }
            .user-name-input { width:50ch; max-width:100%; }
            .user-fields-grid { display:grid; grid-template-columns:minmax(0, 1fr) minmax(0, 1fr); gap:0 10px; align-items:end; }
            .user-fields-grid .full-row { grid-column:1 / -1; }
            .field-registry { width:10ch; max-width:100%; }
            .field-company { width:7ch; max-width:100%; }
            .field-signature-order { width:4ch; max-width:100%; }
            .field-email { width:30ch; max-width:100%; }
            .field-role { width:50ch; max-width:100%; }
            .uppercase-input { text-transform:uppercase; }
            .lowercase-input { text-transform:lowercase; }
            .user-permissions { border-left:1px solid var(--line); padding-left:18px; }
            .user-permissions h2 { margin-top:0; }
            .pdf-modal { width:min(1100px, 96vw); height:92vh; display:flex; flex-direction:column; gap:12px; }
            .pdf-modal-header { display:flex; justify-content:space-between; align-items:center; gap:12px; }
            .pdf-modal-header h2 { margin:0; }
            .pdf-modal-actions { display:flex; align-items:center; gap:8px; flex-wrap:wrap; }
            .pdf-frame { width:100%; flex:1; border:1px solid var(--line); border-radius:6px; background:#fff; }
            .pdf-observe-form { display:none; border:1px solid var(--line); border-radius:6px; padding:12px; margin:0; }
            .pdf-observe-form.active { display:block; }
            .pdf-observe-form textarea { min-height:76px; }
            .message-copy-text { min-height:90px; background:#f8fafc; }
            .progress-track { height:14px; background:#e8edf5; border-radius:999px; overflow:hidden; border:1px solid var(--line); }
            .progress-fill { height:100%; width:12%; background:var(--brand); border-radius:999px; transition:width .35s ease; }
            textarea { width:100%; min-height:110px; font:inherit; padding:9px 10px; border:1px solid var(--field-border); border-radius:6px; background:#fff; }
            .login-title { text-align:center; }
            .login-panel { max-width:380px; margin:0 auto 18px; text-align:center; }
            .login-panel input { width:17ch; max-width:100%; text-align:center; }
            .login-panel .actions { display:flex; flex-direction:column; align-items:center; gap:10px; margin:16px 0 0; }
            .login-panel .recover-link { color:var(--brand); font-weight:700; text-decoration:none; }
            .print-only { display:none; }
            @media (max-width: 860px) {
              body.menu-bottom, body.menu-left, body.menu-right { padding-bottom:0; }
              body.menu-left header, body.menu-right header, body.menu-bottom header { position:static; width:auto; border-left:0; border-right:0; border-bottom:1px solid var(--line); }
              body.menu-left main, body.menu-right main { margin-left:auto; margin-right:auto; }
              .user-form-grid { grid-template-columns:1fr; }
              .user-fields-grid { grid-template-columns:1fr; }
              .config-fields-grid { grid-template-columns:1fr; }
              .config-fields-grid input, .config-fields-grid select, .title-line-input { width:100% !important; max-width:100%; }
              .config-inline-pair { display:block; }
              .modal { width:min(100%, 96vw) !important; max-height:92vh; overflow:auto; }
              .user-permissions { border-left:0; padding-left:0; border-top:1px solid var(--line); padding-top:12px; }
              header { align-items:flex-start; flex-direction:column; padding:12px 16px; }
              main { margin:18px auto; padding:0 14px; }
              h1 { font-size:22px; }
              .grid { grid-template-columns:1fr; }
              table { display:block; overflow-x:auto; white-space:nowrap; }
              input, select, textarea { max-width:100%; }
              .btn, button { min-height:40px; }
            }
            @media print {
              header, .no-print, form, .btn, button { display:none !important; }
              body { background:#fff; }
              main { max-width:none; margin:0; padding:0; }
              .panel { border:0; padding:0; margin:0 0 12px; }
              .print-only { display:block !important; }
              table { font-size:11px; }
              th, td { padding:6px; }
            }
          </style>
        </head>
        <body class="menu-{{ settings.menu_position }}">
          <header>
            <strong>E-Signum Firma Digital para Documentos</strong>
            <nav>
              {% if g.user %}
                <a href="{{ url_for('home') }}">Inicio</a>
                <a href="{{ url_for('profile') }}">Perfil</a>
                <span class="user-name">{{ g.user.name }}</span>
                <a href="{{ url_for('logout_choice') }}">Salir</a>
              {% endif %}
            </nav>
          </header>
          {% if g.user and settings.screen_image_url %}<img class="screen-image {{ settings.screen_image_position }}" style="--screen-image-size:{{ settings.screen_image_size }}px" src="{{ settings.screen_image_url }}" alt="">{% endif %}
          {% if g.user and settings.screen_image_2_url %}<img class="screen-image {{ settings.screen_image_2_position }}" style="--screen-image-size:{{ settings.screen_image_2_size }}px" src="{{ settings.screen_image_2_url }}" alt="">{% endif %}
          <main>
            {% for message in get_flashed_messages() %}<div class="flash">{{ message }}</div>{% endfor %}
            {{ content|safe }}
          </main>
          {% if g.user %}
            <div class="modal-backdrop" id="pdf_viewer_modal">
              <div class="modal pdf-modal">
                <div class="pdf-modal-header">
                  <h2>Ver PDF</h2>
                  <div class="pdf-modal-actions">
                    <button type="button" class="danger" id="pdf_observe_open" style="display:none;">Observar documento</button>
                    <button type="button" class="danger" id="pdf_viewer_close">Cerrar</button>
                  </div>
                </div>
                <form class="pdf-observe-form" id="pdf_observe_form" method="post">
                  <label>Observación</label>
                  <textarea name="note" minlength="10" maxlength="300" required placeholder="Observación de 10 a 300 caracteres"></textarea>
                  <p><button class="danger">Guardar observación</button></p>
                </form>
                <iframe class="pdf-frame" id="pdf_viewer_frame" title="Vista PDF"></iframe>
              </div>
            </div>
          {% endif %}
          <div class="modal-backdrop" id="copyable_message_modal">
            <div class="modal">
              <h2>Mensaje</h2>
              <textarea class="message-copy-text" id="copyable_message_text" readonly></textarea>
              <p>
                <button type="button" id="copyable_message_copy">Copiar mensaje</button>
                <button type="button" class="danger" id="copyable_message_close">Cerrar</button>
              </p>
            </div>
          </div>
          <script>
            const nativeAlert = window.alert.bind(window);
            function showCopyableMessage(message) {
              const modal = document.getElementById("copyable_message_modal");
              const text = document.getElementById("copyable_message_text");
              if (!modal || !text) {
                nativeAlert(message);
                return;
              }
              text.value = String(message || "");
              modal.classList.add("active");
              text.focus();
              text.select();
            }
            window.alert = showCopyableMessage;
            document.getElementById("copyable_message_close")?.addEventListener("click", () => {
              document.getElementById("copyable_message_modal").classList.remove("active");
            });
            document.getElementById("copyable_message_copy")?.addEventListener("click", async () => {
              const text = document.getElementById("copyable_message_text");
              text.focus();
              text.select();
              try {
                await navigator.clipboard.writeText(text.value);
              } catch (error) {
                document.execCommand("copy");
              }
            });
            document.addEventListener("click", (event) => {
              const label = event.target.closest?.("label");
              if (!label) return;
              const hasCheckbox = !!label.querySelector('input[type="checkbox"]');
              if (!hasCheckbox) return;
              if (!event.target.matches?.('input[type="checkbox"]')) {
                event.preventDefault();
                event.stopPropagation();
              }
            }, true);
            function formatRutValue(value) {
              const clean = value.replace(/[.\\-\\s]/g, "").toUpperCase();
              if (clean.length < 2) return clean;
              const body = clean.slice(0, -1).replace(/\\D/g, "");
              const dv = clean.slice(-1).replace(/[^0-9K]/g, "");
              if (!body || !dv) return value;
              return Number(body).toLocaleString("es-CL") + "-" + dv;
            }
            document.querySelectorAll("[data-rut-input]").forEach((input) => {
              input.addEventListener("blur", () => {
                input.value = formatRutValue(input.value);
              });
            });
            function normalizeCasedInputs(root=document) {
              root.querySelectorAll(".uppercase-input").forEach((input) => {
                input.value = input.value.toUpperCase();
              });
              root.querySelectorAll(".lowercase-input").forEach((input) => {
                input.value = input.value.toLowerCase();
              });
            }
            document.querySelectorAll(".uppercase-input").forEach((input) => {
              input.addEventListener("input", () => { input.value = input.value.toUpperCase(); });
            });
            document.querySelectorAll(".lowercase-input").forEach((input) => {
              input.addEventListener("input", () => { input.value = input.value.toLowerCase(); });
            });
            document.querySelectorAll("form").forEach((form) => {
              form.addEventListener("submit", () => normalizeCasedInputs(form));
            });
            normalizeCasedInputs();
            const defaultButtonTitles = new Map([
              ["Ingresar", "Entrar al sistema con RUT y clave"],
              ["Recuperar clave", "Iniciar recuperación de clave"],
              ["Nueva acta", "Abrir la ventana para redactar una nueva acta"],
              ["Guardar", "Guardar el documento sin cerrar la ventana"],
              ["Cerrar documento", "Cerrar definitivamente y generar el PDF"],
              ["Salir", "Cerrar la ventana de texto"],
              ["Agregar o quitar usuarios", "Volver para seleccionar usuarios relacionados"],
              ["Editar", "Abrir el documento para modificarlo si aún no está cerrado"],
              ["Ver PDF", "Abrir el PDF asociado"],
              ["Volver", "Regresar a la pantalla anterior"],
            ]);
            function buttonLabelKey(element) {
              return "button-label:" + (element.id || element.getAttribute("href") || element.name || element.dataset.openConfigModal || element.textContent.trim());
            }
            document.querySelectorAll("button, a.btn, a.recover-link").forEach((element) => {
              const original = element.textContent.trim();
              const stableKey = buttonLabelKey(element);
              const saved = localStorage.getItem(stableKey);
              if (saved) element.textContent = saved;
              element.title = element.title || defaultButtonTitles.get(original) || element.getAttribute("aria-label") || original;
              element.addEventListener("contextmenu", (event) => {
                event.preventDefault();
                const newLabel = prompt("Nuevo nombre del boton:", element.textContent.trim());
                if (newLabel && newLabel.trim()) {
                  element.textContent = newLabel.trim().slice(0, 40);
                  localStorage.setItem(stableKey, element.textContent);
                }
              });
            });
            const pdfViewerModal = document.getElementById("pdf_viewer_modal");
            const pdfViewerFrame = document.getElementById("pdf_viewer_frame");
            const pdfViewerClose = document.getElementById("pdf_viewer_close");
            const pdfObserveOpen = document.getElementById("pdf_observe_open");
            const pdfObserveForm = document.getElementById("pdf_observe_form");
            document.querySelectorAll("[data-pdf-modal]").forEach((link) => {
              link.addEventListener("click", (event) => {
                if (!pdfViewerModal || !pdfViewerFrame) return;
                event.preventDefault();
                pdfViewerFrame.src = link.href;
                if (pdfObserveForm) {
                  pdfObserveForm.classList.remove("active");
                  pdfObserveForm.reset();
                  pdfObserveForm.action = link.dataset.observeUrl || "";
                }
                if (pdfObserveOpen) {
                  pdfObserveOpen.style.display = link.dataset.observeUrl ? "inline-block" : "none";
                }
                pdfViewerModal.classList.add("active");
              });
            });
            if (pdfObserveOpen && pdfObserveForm) {
              pdfObserveOpen.addEventListener("click", () => {
                pdfObserveForm.classList.add("active");
                const note = pdfObserveForm.querySelector('textarea[name="note"]');
                if (note) note.focus();
              });
            }
            if (pdfViewerClose) {
              pdfViewerClose.addEventListener("click", () => {
                pdfViewerModal.classList.remove("active");
                pdfViewerFrame.src = "";
                if (pdfObserveForm) pdfObserveForm.classList.remove("active");
              });
            }
            if (pdfViewerModal) {
              pdfViewerModal.addEventListener("click", (event) => {
                if (event.target === pdfViewerModal) {
                  pdfViewerModal.classList.remove("active");
                  pdfViewerFrame.src = "";
                  if (pdfObserveForm) pdfObserveForm.classList.remove("active");
                }
              });
            }
          </script>
        </body>
        </html>
        """,
        title=title,
        content=content,
        settings=settings,
    )


@app.route("/", methods=["GET"])
def index():
    return redirect(url_for("home") if g.user else url_for("login"))


@app.route("/login", methods=["GET", "POST"])
def login():
    if request.method == "POST":
        rut = normalize_rut(request.form["rut"])
        password = request.form["password"]
        user = find_user_by_rut(rut, active_only=True)
        if user and check_password_hash(user["password_hash"], password):
            session.clear()
            session["user_id"] = user["id"]
            audit("login", "Ingreso correcto", user["id"])
            return redirect(url_for("home"))
        flash("RUT o clave incorrectos.")
    return html_page(
        "Ingreso",
        """
        <h1 class="login-title">Ingreso</h1>
        <form class="panel login-panel" method="post" autocomplete="off" autocorrect="off" spellcheck="false">
          <label>RUT</label><input name="rut" data-rut-input value="" autocomplete="off" autocapitalize="off" spellcheck="false" required>
          <label>Clave</label><input name="password" type="password" value="" autocomplete="off" autocapitalize="off" autocorrect="off" spellcheck="false" data-lpignore="true" data-1p-ignore="true" readonly onfocus="this.removeAttribute('readonly')" required>
          <p class="actions"><button>Ingresar</button> <a class="recover-link" href="/recover">Recuperar clave</a></p>
        </form>
        <script>
          function clearLoginFields() {
            const rut = document.querySelector('input[name="rut"]');
            const password = document.querySelector('input[name="password"]');
            if (rut) rut.value = "";
            if (password) password.value = "";
          }
          window.addEventListener("pageshow", clearLoginFields);
          document.addEventListener("DOMContentLoaded", clearLoginFields);
        </script>
        """,
    )


@app.route("/salir")
@login_required
def logout_choice():
    return html_page(
        "Salir",
        """
        <h1>Salir</h1>
        <div class="panel">
          <p>¿Desea salir o cambiar de usuario?</p>
          <p>
            <a class="btn danger" href="/logout">Salir</a>
            <a class="btn" href="/logout?next=login">Cambio de usuario</a>
            <a class="btn" href="/home">Cancelar</a>
          </p>
        </div>
        """,
    )


@app.route("/logout")
def logout():
    if g.user:
        audit("logout", "Salida del sistema", g.user["id"])
    session.clear()
    return redirect(url_for("login"))


@app.route("/recover", methods=["GET", "POST"])
def recover():
    if request.method == "POST":
        action = request.form.get("action", "request_code")
        if action == "request_code":
            rut = normalize_rut(request.form["rut"])
            email = request.form["email"].strip()
            user = find_user_by_rut(rut, active_only=True)
            if user and user["recovery_email"].lower() != email.lower():
                user = None
            if not user:
                flash("No se encontró una cuenta con esos datos.")
                return redirect(url_for("recover"))

            code = random_recovery_code()
            while db().execute("SELECT 1 FROM password_recovery WHERE token = ? AND used = 0", (code,)).fetchone():
                code = random_recovery_code()
            db().execute(
                "INSERT INTO password_recovery(user_id, token, created_at) VALUES (?, ?, ?)",
                (user["id"], code, now()),
            )
            db().commit()
            audit("password_recovery_requested", "Código de recuperación generado", user["id"])
            ok = send_email(
                user["recovery_email"],
                "Recuperación de clave",
                f"Hola {user['name']},\n\nTu código para recuperar la clave es: {code}\n\nIngresa este código en el sistema para crear una nueva clave.\n\nSi no lo pediste, ignora este mensaje.",
            )
            if ok:
                flash("Se envió un código de recuperación al correo.")
            else:
                flash("No se pudo enviar el correo porque SMTP no está configurado o falló. Revisa la bandeja de salida.")
            return redirect(url_for("recover"))

        if action == "reset_with_code":
            code = request.form.get("code", "").strip()
            password = request.form.get("password", "")
            confirm_password = request.form.get("confirm_password", "")
            record = db().execute(
                """
                SELECT pr.*, u.id AS user_id
                FROM password_recovery pr
                JOIN users u ON u.id = pr.user_id
                WHERE pr.token = ? AND pr.used = 0 AND u.active = 1
                """,
                (code,),
            ).fetchone()
            if not record:
                flash("Código de recuperación inválido o ya utilizado.")
                return redirect(url_for("recover"))
            if password != confirm_password:
                flash("Claves no coinciden")
                return redirect(url_for("recover"))
            if not valid_password(password):
                flash(password_help())
                return redirect(url_for("recover"))
            db().execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(password), record["user_id"]))
            db().execute("UPDATE password_recovery SET used = 1, used_at = ? WHERE id = ?", (now(), record["id"]))
            db().commit()
            audit("password_changed", "Clave cambiada mediante código de recuperación", record["user_id"])
            flash("Clave cambiada. Ahora puedes ingresar.")
            return redirect(url_for("login"))

        flash("Acción no válida.")
        return redirect(url_for("recover"))

    return html_page(
        "Recuperar clave",
        render_template_string(
            """
            <h1>Recuperar clave</h1>
            <form class="panel" method="post" id="send_acta_form">
              <input type="hidden" name="action" value="request_code">
              <h2>Solicitar código</h2>
              <label>RUT</label><input name="rut" data-rut-input required>
              <label>Correo de recuperación</label><input name="email" type="email" required>
              <p><button>Enviar código</button></p>
            </form>
            <form class="panel" method="post" id="edit_user_form">
              <input type="hidden" name="action" value="reset_with_code">
              <h2>Cambiar clave con código</h2>
              <label>Código recibido</label><input name="code" inputmode="numeric" pattern="\\d{6}" maxlength="6" required>
              <label>Nueva clave</label><input name="password" type="password" minlength="7" maxlength="15" required autocomplete="new-password">
              <label>Confirmar nueva clave</label><input name="confirm_password" type="password" minlength="7" maxlength="15" required autocomplete="new-password">
              <p><button>Cambiar clave</button></p>
            </form>
            """,
        ),
    )


@app.route("/reset/<token>", methods=["GET", "POST"])
def reset_password(token: str):
    record = db().execute(
        "SELECT pr.*, u.rut FROM password_recovery pr JOIN users u ON u.id = pr.user_id WHERE token = ? AND used = 0",
        (token,),
    ).fetchone()
    if not record:
        abort(404)
    if request.method == "POST":
        password = request.form["password"]
        if not valid_password(password):
            flash(password_help())
            return redirect(url_for("reset_password", token=token))
        db().execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(password), record["user_id"]))
        db().execute("UPDATE password_recovery SET used = 1, used_at = ? WHERE id = ?", (now(), record["id"]))
        db().commit()
        audit("password_changed", "Clave cambiada mediante recuperación", record["user_id"])
        flash("Clave cambiada. Ahora puedes ingresar.")
        return redirect(url_for("login"))
    return html_page(
        "Cambiar clave",
        """
        <h1>Cambiar clave</h1>
        <form class="panel" method="post">
          <label>Nueva clave</label><input name="password" type="password" minlength="7" maxlength="15" title="Debe tener entre 7 y 15 caracteres. Puede incluir letras, números y símbolos." required>
          <p><button>Guardar clave</button></p>
        </form>
        """,
    )


@app.route("/home")
@login_required
def home():
    return html_page(
        "Inicio",
        render_template_string(
            """
            <h1>Inicio</h1>
            <div class="grid">
              <a class="btn" href="{{ url_for('admin_home') }}">Administración</a>
              <a class="btn" href="{{ url_for('documents_home') }}">Documentos</a>
            </div>
            """,
        ),
    )


def document_access_sql() -> tuple[str, tuple]:
    if g.user["is_admin"] or g.user["is_super_admin"]:
        return "1=1", ()
    return (
        """
        (
          d.status IN ('pending', 'observed')
          AND EXISTS(SELECT 1 FROM document_required_signers rs WHERE rs.document_id = d.id AND rs.user_id = ?)
        )
        """,
        (g.user["id"],),
    )


@app.route("/dashboard")
@login_required
def dashboard():
    return redirect(url_for("documents_home"))


@app.route("/documents")
@login_required
def documents_home():
    return html_page(
        "Documentos",
        render_template_string(
            """
            <h1>Documentos</h1>
            <div class="grid">
              {% if g.user.is_admin or g.user.is_super_admin %}
                <a class="btn" href="{{ url_for('upload_document') }}">Subir archivo</a>
                <a class="btn" href="{{ url_for('actas_document') }}">Actas guardadas</a>
                <a class="btn" href="{{ url_for('actas_document', view='closed') }}">Actas cerradas</a>
              {% endif %}
              <a class="btn" href="{{ url_for('list_documents', searched=1, status='pending_and_observed') }}">Listar documentos</a>
              <a class="btn" href="{{ url_for('verify_signature') }}">Verificar firma</a>
            </div>
            """
        ),
    )


@app.route("/documents/actas")
@admin_required
def actas_document():
    doc_types = db().execute("SELECT * FROM document_types ORDER BY name").fetchall()
    correlative_options = document_type_options_map()
    users = acta_related_users()
    view_status = request.args.get("view", "saved")
    show_closed = view_status == "closed"
    acta_status_filter = "closed" if show_closed else "draft"
    acta_id = request.args.get("id", "")
    acta = None
    if acta_id.isdigit():
        acta = db().execute("SELECT * FROM acta_documents WHERE id = ?", (int(acta_id),)).fetchone()
    actas = db().execute(
        """
        SELECT a.*, dt.name AS type_name, u.name AS creator_name,
          l.document_id AS signature_document_id,
          d.status AS signature_status,
          (SELECT COUNT(*) FROM document_required_signers WHERE document_id = l.document_id) AS signature_required_count,
          (SELECT COUNT(*) FROM signatures WHERE document_id = l.document_id) AS signature_signed_count,
          (SELECT COUNT(*) FROM document_observations WHERE document_id = l.document_id) AS signature_observations_count
        FROM acta_documents a
        JOIN document_types dt ON dt.id = a.document_type_id
        JOIN users u ON u.id = a.created_by_user_id
        LEFT JOIN acta_signature_links l ON l.id = (
          SELECT id FROM acta_signature_links
          WHERE acta_id = a.id
          ORDER BY sent_at DESC, id DESC
          LIMIT 1
        )
        LEFT JOIN documents d ON d.id = l.document_id
        WHERE a.status = ?
        ORDER BY a.created_at DESC
        LIMIT 100
        """
        ,
        (acta_status_filter,),
    ).fetchall()
    return html_page(
        "Actas",
        render_template_string(
            """
            <h1>Actas</h1>
            <form class="panel" method="post" action="{{ url_for('save_acta_document') }}" id="acta_form">
              <input type="hidden" name="acta_id" id="acta_id" value="{{ acta.id if acta else '' }}">
              <input type="hidden" name="action" id="acta_action" value="save">
              <div class="acta-header-fields">
                <div>
                  <label>Tipo de documento</label>
                  <select id="acta_type" name="document_type_id" required {% if acta and acta.status == "closed" %}disabled{% endif %}>
                    {% for t in doc_types %}
                      <option value="{{ t.id }}" data-next-correlative="{{ correlative_options.get(t.id, 1) }}" {% if acta and acta.document_type_id == t.id %}selected{% endif %}>{{ t.name }}</option>
                    {% endfor %}
                  </select>
                </div>
                <div>
                  <label>Fecha</label><input id="acta_date" name="acta_date" type="date" value="{{ acta.acta_date if acta else today }}" required {% if acta and acta.status == "closed" %}readonly{% endif %}>
                </div>
                <div>
                  <label>N°</label><input id="acta_correlative" name="correlative" inputmode="numeric" maxlength="4" placeholder="1" value="{{ acta.correlative if acta else '' }}" required readonly>
                </div>
              </div>
              <h2>Usuarios relacionados</h2>
              <div id="acta_users">
                {% for u in users %}
                  <label class="acta-user" data-document-type-id="{{ u.document_type_id }}" data-role-name="{{ u.role_name }}" data-order="{{ u.signature_order or 99 }}" data-name="{{ u.name }}">
                    <span class="acta-include-control"><input class="acta-user-include" type="checkbox" checked {% if acta and acta.status == "closed" %}disabled{% endif %}> Habilitar</span>
                    <span class="acta-excuse-control"><input class="acta-user-excuse" type="checkbox" {% if acta and acta.status == "closed" %}disabled{% endif %}> Excusa</span>
                    {{ u.name }} - {{ u.role_name }}
                  </label>
                {% endfor %}
              </div>
              <p><button type="button" id="open_acta_text">Nueva acta</button> <a class="btn" href="{{ url_for('documents_home') }}">Volver</a></p>

              <div class="modal-backdrop" id="acta_text_modal">
                <div class="modal" style="width:min(980px,96vw); max-height:92vh; overflow:auto;">
                  <h2>Texto</h2>
                  <div class="acta-audio-tools">
                    <input id="acta_audio_file" type="file" accept="audio/*,.mp3,.wav,.m4a,.ogg,.webm,.mp4,.mpeg,.mpga" {% if acta and acta.status == "closed" %}disabled{% endif %}>
                    <button type="button" id="transcribe_acta_audio" title="Subir audio y agregar la transcripción al texto" {% if acta and acta.status == "closed" %}disabled{% endif %}>Transcribir audio</button>
                    <span class="muted" id="acta_audio_status"></span>
                  </div>
                  <textarea id="acta_text" name="body" minlength="5" style="min-height:62vh;" placeholder="Escribe aqui el texto del acta. Este campo permite mas de 10.000 caracteres." {% if acta and acta.status == "closed" %}readonly{% endif %}>{{ acta.body if acta else '' }}</textarea>
                  <p class="acta-text-actions">
                    <button type="button" id="save_acta_text" title="Guardar el documento sin cerrar la ventana" {% if acta and acta.status == "closed" %}disabled{% endif %}>Guardar</button>
                    <button type="button" id="close_for_users">Volver</button>
                    <button type="submit" class="danger close-document-action" title="Cerrar definitivamente y generar el PDF" data-acta-action="close" {% if acta and acta.status == "closed" %}disabled{% endif %}>Cerrar documento</button>
                  </p>
                </div>
              </div>
            </form>

            <p>
              {% if show_closed %}
                <a class="btn" href="{{ url_for('actas_document') }}">Ver guardadas</a>
              {% else %}
                <a class="btn" href="{{ url_for('actas_document', view='closed') }}">Ver cerradas</a>
              {% endif %}
            </p>
            <section id="actas_saved_section" {% if not actas and not show_closed %}style="display:none;"{% endif %}>
              <h2>{{ "Actas cerradas" if show_closed else "Actas guardadas" }}</h2>
              <table id="actas_saved_table">
                <thead><tr><th>Nombre</th><th>Tipo</th><th>Estado</th><th>Firma</th><th>Creado por</th><th>Acciones</th></tr></thead>
                <tbody id="actas_saved_body">
                  {% for a in actas %}
                  <tr data-acta-row-id="{{ a.id }}" class="{% if a.signature_document_id %}acta-sent-row{% endif %}">
                    <td>{{ a.title }}</td>
                    <td>{{ a.type_name }}</td>
                    <td>{{ "Cerrada" if a.status == "closed" else "Borrador" }}</td>
                    <td>
                      {{ acta_signature_status(a) }}
                      {% if a.signature_document_id %}
                        <br><a href="{{ url_for('document_detail', document_id=a.signature_document_id) }}">Ver estado</a>
                      {% endif %}
                    </td>
                    <td>{{ a.creator_name }}</td>
                    <td>
                      {% if a.status != "closed" %}
                        <a class="btn" href="{{ url_for('actas_document', id=a.id, edit=1) }}">Editar</a>
                      {% endif %}
                      {% if show_closed and a.status == "closed" and a.pdf_path %}
                        <a class="btn" href="{{ url_for('view_acta_pdf', acta_id=a.id) }}" data-pdf-modal>Ver PDF</a>
                        {% if g.user.is_admin or g.user.is_super_admin %}
                          <a class="btn" href="{{ url_for('send_acta_for_signature', acta_id=a.id) }}">{{ "Reenviar a firma" if a.signature_document_id else "Enviar a firma" }}</a>
                        {% endif %}
                      {% endif %}
                    </td>
                  </tr>
                  {% endfor %}
                </tbody>
              </table>
            </section>

            <script>
              const actaType = document.getElementById("acta_type");
              const actaForm = document.getElementById("acta_form");
              const actaSaveUrl = "{{ url_for('save_acta_document') }}";
              const actaTranscribeUrl = "{{ url_for('transcribe_acta_audio') }}";
              const actaAction = document.getElementById("acta_action");
              const actaUsers = Array.from(document.querySelectorAll(".acta-user"));
              const actaTextModal = document.getElementById("acta_text_modal");
              const actaText = document.getElementById("acta_text");
              const actaDate = document.getElementById("acta_date");
              const actaCorrelative = document.getElementById("acta_correlative");
              const loadedActa = {{ 'true' if acta else 'false' }};
              const openEditOnLoad = {{ 'true' if request.args.get('edit') == '1' and acta and acta.status != 'closed' else 'false' }};
              const loadedActaBody = {{ (acta.body if acta else '')|tojson }};
              const currentUserName = {{ g.user.name|tojson }};

              function currentTypeName() {
                return actaType.options[actaType.selectedIndex]?.text || "";
              }

              function visibleUsers() {
                return actaUsers
                  .filter((label) => label.dataset.documentTypeId === actaType.value)
                  .sort((a, b) => Number(a.dataset.order) - Number(b.dataset.order) || a.dataset.name.localeCompare(b.dataset.name));
              }

              function filterActaUsers() {
                visibleUsers();
                actaUsers.forEach((label) => {
                  const visible = label.dataset.documentTypeId === actaType.value;
                  label.style.display = visible ? "block" : "none";
                  const checkbox = label.querySelector(".acta-user-include");
                  const excuse = label.querySelector(".acta-user-excuse");
                  if (checkbox) {
                    checkbox.disabled = !!(excuse && excuse.checked) || ({{ 'true' if acta and acta.status == "closed" else 'false' }});
                    checkbox.checked = visible && !(excuse && excuse.checked);
                  }
                  if (excuse && !visible) excuse.checked = false;
                });
              }

              function selectedUsers() {
                return visibleUsers().filter((label) => label.querySelector(".acta-user-include")?.checked);
              }

              function excusedUsers() {
                return selectedUsers().filter((label) => label.querySelector(".acta-user-excuse")?.checked);
              }

              function syncExcuseCheckboxes() {
                actaUsers.forEach((label) => {
                  const include = label.querySelector(".acta-user-include");
                  const excuse = label.querySelector(".acta-user-excuse");
                  if (!include || !excuse) return;
                  const updateFromExcuse = () => {
                    if (excuse.checked) {
                      include.checked = false;
                      include.disabled = true;
                    } else {
                      include.disabled = false;
                      include.checked = label.dataset.documentTypeId === actaType.value;
                    }
                  };
                  excuse.addEventListener("change", updateFromExcuse);
                  updateFromExcuse();
                });
              }

              function applyAutomaticCorrelative() {
                if (loadedActa) return;
                const nextValue = actaType.options[actaType.selectedIndex]?.dataset.nextCorrelative || "1";
                actaCorrelative.value = String(Number(nextValue) || 1);
              }

              function todayValue() {
                const date = new Date();
                const month = String(date.getMonth() + 1).padStart(2, "0");
                const day = String(date.getDate()).padStart(2, "0");
                return `${date.getFullYear()}-${month}-${day}`;
              }

              function ensureActaDefaults() {
                if (!actaDate.value) actaDate.value = todayValue();
                if (!actaCorrelative.value) applyAutomaticCorrelative();
                if (!actaText.value.trim()) actaText.value = actaBaseText();
              }

              function missingActaFields() {
                const missing = [];
                if (!actaType.value) missing.push("tipo de documento");
                if (!actaDate.value) missing.push("fecha");
                if (!actaCorrelative.value) missing.push("número correlativo");
                if (!actaText.value.trim()) missing.push("texto del acta");
                else if (actaText.value.trim().length < 5) missing.push("texto del acta con mínimo 5 caracteres");
                return missing;
              }

              function actaTitleText() {
                return (actaDate.value || "").replaceAll("-", " ") + " N\u00b0 " + (actaCorrelative.value || "") + " " + currentTypeName();
              }

              function escapeHtml(value) {
                return String(value).replace(/[&<>"']/g, (character) => ({
                  "&": "&amp;",
                  "<": "&lt;",
                  ">": "&gt;",
                  '"': "&quot;",
                  "'": "&#039;",
                }[character]));
              }

              function upsertSavedActaRow(id) {
                const section = document.getElementById("actas_saved_section");
                const body = document.getElementById("actas_saved_body");
                if (!section || !body || !id) return;
                section.style.display = "";
                let row = body.querySelector(`[data-acta-row-id="${id}"]`);
                if (!row) {
                  row = document.createElement("tr");
                  row.dataset.actaRowId = id;
                  body.prepend(row);
                }
                row.innerHTML = `
                  <td>${escapeHtml(actaTitleText())}</td>
                  <td>${escapeHtml(currentTypeName())}</td>
                  <td>Borrador</td>
                  <td>${escapeHtml(currentUserName)}</td>
                  <td><a class="btn" href="/documents/actas?id=${id}&edit=1">Editar</a></td>
                `;
              }

              function actaBaseText() {
                const selected = selectedUsers();
                const excusas = visibleUsers().filter((label) => label.querySelector(".acta-user-excuse")?.checked);
                const oficiales = selected.filter((label) => !label.dataset.roleName.toLowerCase().includes("consejero"));
                const consejeros = selected.filter((label) => label.dataset.roleName.toLowerCase().includes("consejero"));
                const lines = [
                  currentTypeName(),
                  (actaDate.value || "").replaceAll("-", " ") + " N\u00b0 " + (actaCorrelative.value || ""),
                  "",
                  "OFICIALES",
                ];
                oficiales.forEach((label) => lines.push(label.dataset.name + " - " + label.dataset.roleName));
                lines.push("", "", "CONSEJEROS");
                consejeros.forEach((label) => lines.push(label.dataset.name + " - " + label.dataset.roleName));
                lines.push("", "", "EXCUSAS", "");
                excusas.forEach((label) => lines.push(label.dataset.name + " - " + label.dataset.roleName));
                lines.push("", "");
                return lines.join("\\n");
              }

              actaType.addEventListener("change", () => {
                filterActaUsers();
                applyAutomaticCorrelative();
              });
              document.getElementById("open_acta_text").addEventListener("click", () => {
                ensureActaDefaults();
                actaText.value = actaBaseText();
                actaTextModal.classList.add("active");
                actaText.focus();
              });
              document.getElementById("close_for_users").addEventListener("click", () => {
                actaTextModal.classList.remove("active");
              });
              document.getElementById("transcribe_acta_audio")?.addEventListener("click", async () => {
                const audioInput = document.getElementById("acta_audio_file");
                const status = document.getElementById("acta_audio_status");
                const button = document.getElementById("transcribe_acta_audio");
                const file = audioInput?.files?.[0];
                if (!file) {
                  alert("Debes seleccionar un archivo de audio.");
                  return;
                }
                const formData = new FormData();
                formData.append("audio", file);
                if (status) status.textContent = "Transcribiendo audio...";
                if (button) button.disabled = true;
                try {
                  const response = await fetch(actaTranscribeUrl, {
                    method: "POST",
                    body: formData,
                    headers: {"X-Requested-With": "XMLHttpRequest"}
                  });
                  const result = await response.json();
                  if (!result.ok) {
                    alert(result.message || "No se pudo transcribir el audio.");
                    return;
                  }
                  const transcription = (result.text || "").trim();
                  if (transcription) {
                    actaText.value = actaText.value.trim()
                      ? `${actaText.value.trim()}\\n\\n${transcription}`
                      : transcription;
                  }
                  if (status) status.textContent = result.message || "Audio transcrito.";
                } catch (error) {
                  alert("No se pudo transcribir el audio. Detalle: " + error.message);
                } finally {
                  if (button) button.disabled = false;
                }
              });
              document.getElementById("save_acta_text").addEventListener("click", async () => {
                actaAction.value = "save";
                ensureActaDefaults();
                const missing = missingActaFields();
                if (missing.length) {
                  alert("Falta completar: " + missing.join(", ") + ".");
                  return;
                }
                const formData = new FormData(actaForm);
                formData.set("action", "save");
                formData.set("acta_id", document.getElementById("acta_id").value || "");
                formData.set("document_type_id", actaType.value || "");
                formData.set("acta_date", actaDate.value || "");
                formData.set("correlative", actaCorrelative.value || "");
                formData.set("body", actaText.value || "");
                try {
                  const response = await fetch(actaSaveUrl, {
                    method: "POST",
                    body: formData,
                    headers: {"X-Requested-With": "XMLHttpRequest"}
                  });
                  const responseText = await response.text();
                  let result;
                  try {
                    result = JSON.parse(responseText);
                  } catch (parseError) {
                    alert("No se pudo guardar. El servidor respondio: " + response.status + ". " + responseText.slice(0, 180));
                    return;
                  }
                  if (result.ok) {
                    document.getElementById("acta_id").value = result.id;
                    upsertSavedActaRow(result.id);
                    alert("Se guard\u00f3 con \u00e9xito.");
                  } else {
                    alert(result.message || "No se pudo guardar el acta.");
                  }
                } catch (error) {
                  const missingNow = missingActaFields();
                  if (missingNow.length) {
                    alert("Falta completar: " + missingNow.join(", ") + ".");
                  } else {
                    alert("No se pudo guardar el acta. Detalle: " + error.message);
                  }
                }
              });
              document.querySelectorAll("[data-acta-action]").forEach((button) => {
                button.addEventListener("click", (event) => {
                  const action = button.dataset.actaAction;
                  if (action === "close" && !document.getElementById("acta_id").value) {
                    event.preventDefault();
                    alert("No se puede cerrar mientras no sea guardado.");
                    return;
                  }
                  if (action === "close" && !confirm("ESTA SEGURO QUE QUIERE CERRAR ESTE DOCUMENTO NO SE PODR\u00c1 VOLVER A EDITAR")) {
                    event.preventDefault();
                    return;
                  }
                  actaAction.value = action;
                });
              });
              filterActaUsers();
              syncExcuseCheckboxes();
              applyAutomaticCorrelative();
              if (openEditOnLoad) {
                actaText.value = loadedActaBody;
                actaTextModal.classList.add("active");
                actaText.focus();
              }
            </script>
            """,
            doc_types=doc_types,
            correlative_options=correlative_options,
            users=users,
            acta=acta,
            actas=actas,
            show_closed=show_closed,
            acta_signature_status=acta_signature_status,
            today=datetime.now().strftime("%Y-%m-%d"),
        ),
    )


@app.post("/documents/actas/save")
@admin_required
def save_acta_document():
    wants_json = request.headers.get("X-Requested-With") == "XMLHttpRequest"

    def acta_response(ok: bool, message: str, acta_id_value: int | None = None, redirect_id: str | int | None = None):
        if wants_json:
            return {"ok": ok, "message": message, "id": acta_id_value}
        if message:
            flash(message)
        return redirect(url_for("actas_document", id=redirect_id) if redirect_id else url_for("actas_document"))

    action = request.form.get("action", "save")
    if action not in {"save", "close"}:
        return acta_response(False, "Acción no válida.")
    acta_id = request.form.get("acta_id", "")
    if action == "close" and not acta_id.isdigit():
        return acta_response(False, "No se puede cerrar mientras no sea guardado.")
    document_type_raw = request.form.get("document_type_id", "").strip()
    if not document_type_raw.isdigit():
        return acta_response(False, "Falta completar: tipo de documento.", redirect_id=acta_id)
    document_type_id = int(document_type_raw)
    acta_date = request.form.get("acta_date", "").strip() or datetime.now().strftime("%Y-%m-%d")
    correlative = request.form.get("correlative", "").strip()
    body = request.form.get("body", "")
    doc_type = db().execute("SELECT * FROM document_types WHERE id = ?", (document_type_id,)).fetchone()
    if not doc_type or not acta_date or not correlative or not body.strip() or len(body.strip()) < 5:
        missing_fields = []
        if not doc_type:
            missing_fields.append("tipo de documento")
        if not acta_date:
            missing_fields.append("fecha")
        if not correlative:
            missing_fields.append("correlativo")
        if not body.strip():
            missing_fields.append("texto")
        elif len(body.strip()) < 5:
            missing_fields.append("texto con mínimo 5 caracteres")
        return acta_response(False, "Falta completar: " + ", ".join(missing_fields) + ".", redirect_id=acta_id)
    title = acta_title(acta_date, correlative, doc_type["name"])
    if acta_id.isdigit():
        existing = db().execute("SELECT * FROM acta_documents WHERE id = ?", (int(acta_id),)).fetchone()
        if not existing:
            abort(404)
        if existing["status"] == "closed":
            return acta_response(False, "El acta ya esta cerrada y no se puede modificar.", redirect_id=acta_id)
        db().execute(
            """
            UPDATE acta_documents
            SET title = ?, document_type_id = ?, acta_date = ?, correlative = ?, body = ?, updated_at = ?
            WHERE id = ?
            """,
            (title, document_type_id, acta_date, correlative, body, now(), int(acta_id)),
        )
        saved_id = int(acta_id)
        audit("acta_updated", f"Acta actualizada: {title}")
    else:
        cur = db().execute(
            """
            INSERT INTO acta_documents(title, document_type_id, acta_date, correlative, body, created_by_user_id, created_at, updated_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """,
            (title, document_type_id, acta_date, correlative, body, g.user["id"], now(), now()),
        )
        saved_id = cur.lastrowid
        audit("acta_created", f"Acta guardada: {title}")
    advance_document_correlative(document_type_id, correlative)

    if action == "close":
        pdf_path = create_acta_pdf(saved_id)
        db().execute(
            "UPDATE acta_documents SET status = 'closed', pdf_path = ?, closed_at = ?, updated_at = ? WHERE id = ?",
            (pdf_path, now(), now(), saved_id),
        )
        audit("acta_closed", f"Acta cerrada y convertida a PDF: {title}")
        message = "Acta cerrada y PDF generado."
    else:
        message = "EL DOCUMENTO SE GUARD\u00d3 EXITOSAMENTE"
    db().commit()
    return acta_response(True, message, saved_id, saved_id)


@app.post("/documents/actas/transcribe-audio")
@admin_required
def transcribe_acta_audio():
    uploaded = request.files.get("audio")
    if not uploaded or not uploaded.filename:
        return {"ok": False, "message": "Debes seleccionar un archivo de audio."}, 400
    extension = Path(uploaded.filename).suffix.lower()
    if extension not in ALLOWED_AUDIO_EXTENSIONS:
        return {"ok": False, "message": "Formato de audio no permitido. Usa MP3, WAV, M4A, OGG, WEBM o MP4."}, 400
    filename = secure_filename(uploaded.filename)
    stored_path = AUDIO_DIR / f"{secrets.token_hex(8)}-{filename}"
    uploaded.save(stored_path)
    ok, message = transcribe_audio_file(stored_path)
    audit(
        "acta_audio_transcribed" if ok else "acta_audio_transcription_failed",
        f"Archivo de audio: {Path(uploaded.filename).name}",
    )
    db().commit()
    if ok:
        return {"ok": True, "text": message, "message": "Audio transcrito correctamente."}
    return {"ok": False, "message": message}, 503


@app.route("/documents/actas/<int:acta_id>/pdf")
@admin_required
def view_acta_pdf(acta_id: int):
    acta = db().execute("SELECT * FROM acta_documents WHERE id = ?", (acta_id,)).fetchone()
    if not acta or not acta["pdf_path"]:
        abort(404)
    audit("acta_pdf_viewed", f"Visualizó PDF de acta: {acta['title']}")
    return send_file(acta["pdf_path"], mimetype="application/pdf", as_attachment=False)


@app.route("/documents/actas/<int:acta_id>/send-for-signature", methods=["GET", "POST"])
@admin_required
def send_acta_for_signature(acta_id: int):
    acta = db().execute(
        """
        SELECT a.*, dt.name AS type_name, dt.required_signatures
        FROM acta_documents a
        JOIN document_types dt ON dt.id = a.document_type_id
        WHERE a.id = ?
        """,
        (acta_id,),
    ).fetchone()
    if not acta:
        abort(404)
    pdf_path = Path(acta["pdf_path"] or "")
    if acta["status"] != "closed" or not acta["pdf_path"] or not pdf_path.exists():
        flash("Solo se pueden enviar a firma actas cerradas con PDF generado.")
        return redirect(url_for("actas_document"))
    previous_link = db().execute(
        """
        SELECT l.*, d.status AS document_status,
          (SELECT COUNT(*) FROM document_required_signers WHERE document_id = l.document_id) AS required_count,
          (SELECT COUNT(*) FROM signatures WHERE document_id = l.document_id) AS signed_count,
          (SELECT COUNT(*) FROM document_observations WHERE document_id = l.document_id) AS observations_count
        FROM acta_signature_links l
        JOIN documents d ON d.id = l.document_id
        WHERE l.acta_id = ?
        ORDER BY l.sent_at DESC, l.id DESC
        LIMIT 1
        """,
        (acta_id,),
    ).fetchone()

    users = db().execute(
        """
        SELECT u.id, u.name, u.role_id, u.active, u.can_sign, u.is_super_admin, r.name AS role_name
        FROM users u
        JOIN roles r ON r.id = u.role_id
        WHERE u.active = 1
          AND u.can_sign = 1
          AND u.is_super_admin = 0
          AND EXISTS(
            SELECT 1 FROM document_type_signer_roles dtsr
            WHERE dtsr.document_type_id = ? AND dtsr.role_id = u.role_id
          )
        ORDER BY u.signature_order, u.name
        """,
        (acta["document_type_id"],),
    ).fetchall()

    if request.method == "POST":
        signer_ids = [int(x) for x in request.form.getlist("signers") if x.isdigit()]
        resend_reason = request.form.get("resend_reason", "").strip()
        if previous_link and not valid_reason_note(resend_reason):
            flash("Para reenviar debes escribir un motivo alfanumerico de entre 10 y 80 caracteres.")
            return redirect(url_for("send_acta_for_signature", acta_id=acta_id))
        title = acta["title"]
        title_key = normalize_document_title(title)
        duplicate = db().execute(
            """
            SELECT id, title
            FROM documents
            WHERE document_type_id = ?
              AND title_key = ?
              AND status != 'annulled'
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (acta["document_type_id"], title_key),
        ).fetchone()
        if duplicate and not previous_link:
            flash(f"Ya existe un documento enviado a firma con este título: {duplicate['title']}.")
            return redirect(url_for("send_acta_for_signature", acta_id=acta_id))
        if not signer_ids:
            flash("Debes seleccionar al menos un firmante.")
            return redirect(url_for("send_acta_for_signature", acta_id=acta_id))
        if len(signer_ids) > acta["required_signatures"]:
            flash(f"Este tipo de documento permite máximo {acta['required_signatures']} firmantes.")
            return redirect(url_for("send_acta_for_signature", acta_id=acta_id))
        allowed_user_ids = {user["id"] for user in users}
        if any(signer_id not in allowed_user_ids for signer_id in signer_ids):
            flash("Uno o más firmantes seleccionados no están autorizados para este tipo de documento.")
            return redirect(url_for("send_acta_for_signature", acta_id=acta_id))

        filename = safe_display_filename(pdf_path.name)
        stored_path = UPLOAD_DIR / f"{secrets.token_hex(8)}-{filename}"
        shutil.copy2(pdf_path, stored_path)
        cur = db().execute(
            """
            INSERT INTO documents(title, title_key, document_type_id, original_filename, stored_path, uploaded_by_user_id, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (title, title_key, acta["document_type_id"], filename, str(stored_path), g.user["id"], now()),
        )
        document_id = cur.lastrowid
        for signer_id in signer_ids:
            db().execute("INSERT INTO document_required_signers(document_id, user_id) VALUES (?, ?)", (document_id, signer_id))
        db().execute(
            """
            INSERT INTO acta_signature_links(acta_id, document_id, resend_reason, sent_by_user_id, sent_at)
            VALUES (?, ?, ?, ?, ?)
            """,
            (acta_id, document_id, resend_reason if previous_link else "", g.user["id"], now()),
        )
        db().commit()
        audit(
            "acta_sent_for_signature",
            f"Acta {'reenviada' if previous_link else 'enviada'} a firma: {title}. Motivo: {resend_reason if previous_link else 'primer envío'}",
            document_id=document_id,
        )
        sent, failed = notify_document_signers(document_id, signer_ids)
        audit("document_signers_notified", f"Correos enviados: {sent}. Pendientes/fallidos: {failed}", document_id=document_id)
        if failed:
            flash(f"Acta enviada a firma. Correos enviados: {sent}. Correos no enviados o sin SMTP: {failed}.")
        else:
            flash(f"Acta enviada a firma y correos enviados a {sent} firmantes.")
        return redirect(url_for("document_detail", document_id=document_id))

    return html_page(
        "Enviar acta a firma",
        render_template_string(
            """
            <h1>{{ "Reenviar acta a firma" if previous_link else "Enviar acta a firma" }}</h1>
            <form class="panel" method="post" id="edit_user_form">
              <p><strong>Documento:</strong> {{ acta.title }}</p>
              <p><strong>Tipo:</strong> {{ acta.type_name }}</p>
              {% if previous_link %}
                <div class="flash">
                  Esta acta ya fue enviada a firma.
                  Estado actual: {{ document_status_display(previous_link.document_status, previous_link.observations_count or 0, previous_link.required_count or 0, previous_link.signed_count or 0) }}.
                  <a href="{{ url_for('document_detail', document_id=previous_link.document_id) }}">Ver documento enviado</a>
                </div>
                <label>Motivo del reenvío</label>
                <input name="resend_reason" minlength="10" maxlength="80" pattern="[A-Za-z0-9ÁÉÍÓÚÜÑáéíóúüñ ]+" required placeholder="Motivo de 10 a 80 caracteres">
              {% endif %}
              <p><a class="btn" href="{{ url_for('view_acta_pdf', acta_id=acta.id) }}" data-pdf-modal>Ver PDF</a></p>
              <h2>Firmantes</h2>
              <p class="muted">{{ users|length }} usuarios disponibles. Máximo {{ acta.required_signatures }} firmantes.</p>
              {% for u in users %}
                <label>
                  <input type="checkbox" name="signers" value="{{ u.id }}"> {{ u.name }} - {{ u.role_name }}
                </label>
              {% endfor %}
              <p><button id="send_acta_button">{{ "Reenviar a firma" if previous_link else "Enviar a firma" }}</button> <a class="btn" href="{{ url_for('actas_document') }}">Volver</a></p>
            </form>
            <div class="modal-backdrop" id="email_progress_modal">
              <div class="modal">
                <h2>Envío de correos</h2>
                <p id="email_progress_text">Preparando envío...</p>
                <div class="progress-track"><div class="progress-fill" id="email_progress_fill"></div></div>
              </div>
            </div>
            <script>
              const sendActaForm = document.getElementById("send_acta_form");
              const sendActaButton = document.getElementById("send_acta_button");
              const progressModal = document.getElementById("email_progress_modal");
              const progressFill = document.getElementById("email_progress_fill");
              const progressText = document.getElementById("email_progress_text");
              sendActaForm?.addEventListener("submit", (event) => {
                if (sendActaForm.dataset.submitting === "1") return;
                event.preventDefault();
                sendActaForm.dataset.submitting = "1";
                if (sendActaButton) {
                  sendActaButton.disabled = true;
                  sendActaButton.textContent = "Enviando...";
                }
                progressModal?.classList.add("active");
                if (progressFill) progressFill.style.width = "12%";
                if (progressText) progressText.textContent = "Preparando envío...";
                const steps = [
                  ["Preparando documento...", "25%"],
                  ["Registrando firmantes...", "50%"],
                  ["Enviando correos a firmantes...", "75%"],
                  ["Finalizando envío...", "92%"],
                ];
                let index = 0;
                const timer = setInterval(() => {
                  const step = steps[Math.min(index, steps.length - 1)];
                  if (progressText) progressText.textContent = step[0];
                  if (progressFill) progressFill.style.width = step[1];
                  index += 1;
                  if (index >= steps.length) clearInterval(timer);
                }, 500);
                setTimeout(() => {
                  HTMLFormElement.prototype.submit.call(sendActaForm);
                }, 350);
              });
            </script>
            """,
            acta=acta,
            users=users,
            previous_link=previous_link,
            document_status_display=document_status_display,
        ),
    )


@app.route("/documents/list")
@login_required
def list_documents():
    doc_types = db().execute("SELECT * FROM document_types ORDER BY name").fetchall()
    selected_type, selected_status, filter_sql, query_params = document_list_filter_data()
    type_label, status_label = document_filter_labels(selected_type, selected_status)
    searched = request.args.get("searched") == "1"
    documents = document_list_rows(filter_sql, query_params) if searched else []
    return html_page(
        "Listar documentos",
        render_template_string(
            """
            <h1>Listar documentos</h1>
            <form class="panel no-print" method="get">
              <input type="hidden" name="searched" value="1">
              <label>Tipo de documento</label>
              <select name="document_type_id">
                <option value="">Todos</option>
                {% for t in doc_types %}
                  <option value="{{ t.id }}" {% if selected_type|int == t.id %}selected{% endif %}>{{ t.name }}</option>
                {% endfor %}
              </select>
              <label>Estado</label>
              <select name="status">
                <option value="" {% if selected_status == "" %}selected{% endif %}>Todos</option>
                <option value="pending" {% if selected_status == "pending" %}selected{% endif %}>Pendiente</option>
                <option value="pending_and_observed" {% if selected_status == "pending_and_observed" %}selected{% endif %}>Pendientes y observados</option>
                <option value="signed_pending_close" {% if selected_status == "signed_pending_close" %}selected{% endif %}>Firmado pendiente de cierre</option>
                <option value="pending_observed" {% if selected_status == "pending_observed" %}selected{% endif %}>Pendiente con observaciónes</option>
                <option value="observed" {% if selected_status == "observed" %}selected{% endif %}>Observado</option>
                <option value="closed" {% if selected_status == "closed" %}selected{% endif %}>Cerrado</option>
                <option value="closed_observed" {% if selected_status == "closed_observed" %}selected{% endif %}>Cerrado con observaciónes</option>
                <option value="annulled" {% if selected_status == "annulled" %}selected{% endif %}>Anulado</option>
              </select>
              <p>
                <button>Buscar</button>
                <a class="btn" href="{{ url_for('list_documents') }}">Limpiar</a>
                <a class="btn" href="{{ url_for('download_document_list_pdf', searched=1, document_type_id=selected_type, status=selected_status) }}" download="listado-documentos.pdf">Descargar</a>
                <button type="button" onclick="window.print()">Imprimir</button>
                <a class="btn" href="{{ url_for('home') }}">Volver</a>
              </p>
            </form>
            {% if searched %}
              <div class="panel print-only" id="print_header" style="text-align:center;">
                <h2>Listado de Documentos</h2>
                <p><strong>Tipo de documento:</strong> {{ type_label }}</p>
                <p><strong>Estado:</strong> {{ status_label }}</p>
              </div>
              <table id="print_table">
                <tr><th>Título</th><th>Tipo</th><th>Estado</th><th>Firmas</th><th class="no-print">Acciones</th></tr>
                {% for d in documents %}
                  <tr>
                    <td>{{ d.title }}</td>
                    <td>{{ d.type_name }}</td>
                    <td><span class="status-pill {{ status_class(d.status, d.observations_count, d.required_count, d.signed_count) }}">{{ status_display(d.status, d.observations_count, d.required_count, d.signed_count) }}</span></td>
                    <td>{{ d.signed_count }}/{{ d.required_count }}</td>
                    <td class="no-print"><a class="btn" href="{{ url_for('document_detail', document_id=d.id) }}">Ver</a></td>
                  </tr>
                {% endfor %}
              </table>
            {% endif %}
            """,
            documents=documents,
            doc_types=doc_types,
            selected_type=selected_type,
            selected_status=selected_status,
            type_label=type_label,
            status_label=status_label,
            searched=searched,
            status_display=document_status_display,
            status_class=document_status_class,
        ),
    )


@app.route("/documents/verify-signature", methods=["GET", "POST"])
@login_required
def verify_signature():
    result = None
    searched = False
    code = ""
    if request.method == "POST":
        searched = True
        code = request.form.get("signature_code", "").strip()
        if code:
            where, params = document_access_sql()
            result = db().execute(
                f"""
                SELECT s.signature_code, s.signed_at, u.name AS signer_name, r.name AS role_name,
                       d.id AS document_id, d.title AS document_title, d.status, dt.name AS type_name
                FROM signatures s
                JOIN users u ON u.id = s.user_id
                JOIN roles r ON r.id = u.role_id
                JOIN documents d ON d.id = s.document_id
                JOIN document_types dt ON dt.id = d.document_type_id
                WHERE s.signature_code = ? AND {where}
                """,
                (code, *params),
            ).fetchone()
    return html_page(
        "Verificar firma",
        render_template_string(
            """
            <h1>Verificar firma</h1>
            <form class="panel" method="post">
              <label>Código de firma</label>
              <input name="signature_code" value="{{ code }}" inputmode="numeric" pattern="\\d+" required>
              <p><button>Verificar</button> <a class="btn" href="{{ url_for('documents_home') }}">Volver</a></p>
            </form>
            {% if searched %}
              {% if result %}
                <div class="panel">
                  <h2>Firma encontrada</h2>
                  <p><strong>Código:</strong> {{ result.signature_code }}</p>
                  <p><strong>Firmante:</strong> {{ result.signer_name }}</p>
                  <p><strong>Cargo:</strong> {{ result.role_name }}</p>
                  <p><strong>Documento:</strong> {{ result.document_title }}</p>
                  <p><strong>Tipo:</strong> {{ result.type_name }}</p>
                  <p><strong>Fecha firma:</strong> {{ result.signed_at }}</p>
                  <p><strong>Estado documento:</strong> {{ result.status }}</p>
                  <p><a class="btn" href="{{ url_for('view_pdf', document_id=result.document_id) }}" data-pdf-modal>Ver PDF</a></p>
                </div>
              {% else %}
                <div class="flash">No se encontro una firma visible para ese código.</div>
              {% endif %}
            {% endif %}
            """,
            code=code,
            searched=searched,
            result=result,
        ),
    )


def document_list_filter_data() -> tuple[str, str, str, list[object]]:
    where, params = document_access_sql()
    filters: list[str] = [where]
    query_params = list(params)
    selected_type = request.args.get("document_type_id", "")
    selected_status = request.args.get("status", "")
    searched = request.args.get("searched") == "1"
    if selected_type.isdigit():
        filters.append("d.document_type_id = ?")
        query_params.append(int(selected_type))
    if selected_status == "pending_and_observed":
        filters.append("d.status IN ('pending', 'observed')")
    elif selected_status == "pending":
        filters.append("d.status = 'pending'")
        filters.append("NOT EXISTS(SELECT 1 FROM document_observations o WHERE o.document_id = d.id AND o.responded_at IS NULL)")
        filters.append("(SELECT COUNT(*) FROM document_required_signers WHERE document_id = d.id) != (SELECT COUNT(*) FROM signatures WHERE document_id = d.id)")
    elif selected_status == "signed_pending_close":
        filters.append("d.status = 'pending'")
        filters.append("(SELECT COUNT(*) FROM document_required_signers WHERE document_id = d.id) > 0")
        filters.append("(SELECT COUNT(*) FROM document_required_signers WHERE document_id = d.id) = (SELECT COUNT(*) FROM signatures WHERE document_id = d.id)")
        filters.append("NOT EXISTS(SELECT 1 FROM document_observations o WHERE o.document_id = d.id AND o.responded_at IS NULL)")
    elif selected_status == "pending_observed":
        filters.append("d.status = 'pending'")
        filters.append("EXISTS(SELECT 1 FROM document_observations o WHERE o.document_id = d.id AND o.responded_at IS NULL)")
    elif selected_status == "observed":
        filters.append("d.status = 'observed'")
    elif selected_status == "closed":
        filters.append("d.status = 'closed'")
        filters.append("NOT EXISTS(SELECT 1 FROM document_observations o WHERE o.document_id = d.id)")
    elif selected_status == "closed_observed":
        filters.append("d.status = 'closed'")
        filters.append("EXISTS(SELECT 1 FROM document_observations o WHERE o.document_id = d.id)")
    elif selected_status == "annulled":
        filters.append("d.status = 'annulled'")
    filter_sql = " AND ".join(f"({item})" for item in filters)
    return selected_type, selected_status, filter_sql, query_params


def document_list_rows(filter_sql: str, query_params: list[object]) -> list[sqlite3.Row]:
    return db().execute(
        f"""
        SELECT d.*, dt.name AS type_name,
          (SELECT COUNT(*) FROM document_required_signers WHERE document_id = d.id) AS required_count,
          (SELECT COUNT(*) FROM signatures WHERE document_id = d.id) AS signed_count,
          CASE
            WHEN d.status = 'closed' THEN (SELECT COUNT(*) FROM document_observations WHERE document_id = d.id)
            ELSE (SELECT COUNT(*) FROM document_observations WHERE document_id = d.id AND responded_at IS NULL)
          END AS observations_count,
          CASE WHEN d.status = 'pending'
             AND ? = 1
             AND EXISTS(SELECT 1 FROM document_required_signers rs WHERE rs.document_id = d.id AND rs.user_id = ?)
             AND NOT EXISTS(SELECT 1 FROM signatures s WHERE s.document_id = d.id AND s.user_id = ?)
          THEN 1 ELSE 0 END AS can_current_user_sign
        FROM documents d
        JOIN document_types dt ON dt.id = d.document_type_id
        WHERE {filter_sql}
        ORDER BY d.created_at DESC
        """,
        [g.user["can_sign"], g.user["id"], g.user["id"], *query_params],
    ).fetchall()


@app.route("/documents/list/pdf")
@login_required
def download_document_list_pdf():
    selected_type, selected_status, filter_sql, query_params = document_list_filter_data()
    type_label, status_label = document_filter_labels(selected_type, selected_status)
    documents = document_list_rows(filter_sql, query_params)
    packet = BytesIO()
    c = canvas.Canvas(packet, pagesize=letter)
    c.setFont("Helvetica-Bold", 14)
    c.drawCentredString(306, 750, "Listado de Documentos")
    c.setFont("Helvetica", 9)
    c.drawCentredString(306, 732, f"Tipo de documento: {type_label}")
    c.drawCentredString(306, 718, f"Estado: {status_label}")
    c.drawCentredString(306, 704, f"Fecha: {now()}")
    y = 676
    c.setFont("Helvetica-Bold", 8)
    c.drawString(42, y, "Título")
    c.drawString(205, y, "Tipo")
    c.drawString(390, y, "Estado")
    c.drawString(515, y, "Firmas")
    y -= 12
    c.setFont("Helvetica", 8)
    for doc in documents:
        if y < 45:
            c.showPage()
            y = 750
            c.setFont("Helvetica", 8)
        c.drawString(42, y, doc["title"][:30])
        c.drawString(205, y, doc["type_name"][:32])
        c.drawString(390, y, document_status_display(doc["status"], doc["observations_count"], doc["required_count"], doc["signed_count"])[:24])
        c.drawString(515, y, f"{doc['signed_count']}/{doc['required_count']}")
        y -= 12
    c.save()
    packet.seek(0)
    audit("document_list_downloaded", "Descargó listado de documentos en PDF")
    return send_file(
        packet,
        mimetype="application/pdf",
        as_attachment=True,
        download_name="listado-documentos.pdf",
    )


@app.route("/profile")
@login_required
def profile():
    return html_page(
        "Perfil",
        render_template_string(
            """
            <h1>Perfil</h1>
            <div class="panel">
              <p><strong>Nombre:</strong> {{ g.user.name }}</p>
              <p><strong>Número de registro:</strong> {{ g.user.registration_number or "Sin registrar" }}</p>
              <p><strong>Cuerpo de Bomberos:</strong> {{ g.user.fire_department or "Sin registrar" }}</p>
              <p><strong>Compañía:</strong> {{ g.user.company or "Sin registrar" }}</p>
              <p><strong>RUT:</strong> {{ g.user.rut }}</p>
              {% if not g.user.is_super_admin %}<p><strong>Cargo:</strong> {{ g.user.role_name }}</p>{% endif %}
              <p><strong>Código interno:</strong> {{ g.user.internal_code }}</p>
              <p><strong>Correo recuperación:</strong> {{ g.user.recovery_email }}</p>
            </div>
            """,
        ),
    )


@app.route("/documents/<int:document_id>")
@login_required
def document_detail(document_id: int):
    where, params = document_access_sql()
    document = db().execute(
        f"""
        SELECT d.*, dt.name AS type_name
        FROM documents d JOIN document_types dt ON dt.id = d.document_type_id
        WHERE d.id = ? AND {where}
        """,
        (document_id, *params),
    ).fetchone()
    if not document:
        abort(404)
    signers = db().execute(
        """
        SELECT u.id, u.name, r.name AS role_name, s.signature_code, s.signed_at
        FROM document_required_signers rs
        JOIN users u ON u.id = rs.user_id
        JOIN roles r ON r.id = u.role_id
        LEFT JOIN signatures s ON s.document_id = rs.document_id AND s.user_id = u.id
        WHERE rs.document_id = ?
        ORDER BY r.name, u.name
        """,
        (document_id,),
    ).fetchall()
    can_sign_doc = any(s["id"] == g.user["id"] and not s["signed_at"] for s in signers) and document["status"] == "pending" and g.user["can_sign"]
    can_observe_doc = can_sign_doc
    signed_file = bool(document["signed_path"])
    can_manage_signers = bool((g.user["is_admin"] or g.user["is_super_admin"]) and document["status"] == "pending")
    available_signers = eligible_signers_for_document_type(document["document_type_id"], document_id) if can_manage_signers else []
    max_signers = db().execute("SELECT required_signatures FROM document_types WHERE id = ?", (document["document_type_id"],)).fetchone()["required_signatures"]
    signature_counts = document_signature_counts(document_id)
    can_close_document = bool(
        (g.user["is_admin"] or g.user["is_super_admin"])
        and document["status"] == "pending"
        and signature_counts["required_count"] > 0
        and signature_counts["required_count"] == signature_counts["signed_count"]
    )
    can_force_close_document = bool(
        (g.user["is_admin"] or g.user["is_super_admin"])
        and document["status"] == "pending"
        and signature_counts["required_count"] > 0
        and signature_counts["signed_count"] > 0
        and signature_counts["signed_count"] < signature_counts["required_count"]
    )
    unresolved_observations_count = db().execute(
        "SELECT COUNT(*) AS total FROM document_observations WHERE document_id = ? AND responded_at IS NULL",
        (document_id,),
    ).fetchone()["total"]
    can_resolve_observed = bool((g.user["is_admin"] or g.user["is_super_admin"]) and document["status"] == "observed" and unresolved_observations_count > 0)
    observations = db().execute(
        """
        SELECT o.*, u.name AS user_name, r.name AS role_name, responder.name AS responder_name
        FROM document_observations o
        JOIN users u ON u.id = o.user_id
        JOIN roles r ON r.id = u.role_id
        LEFT JOIN users responder ON responder.id = o.responded_by_user_id
        WHERE o.document_id = ?
        ORDER BY o.created_at DESC
        """,
        (document_id,),
    ).fetchall()
    observations_count = len(observations)
    status_observations_count = observations_count if document["status"] == "closed" else unresolved_observations_count
    document_format = document_file_format(document)
    view_label = "Ver PDF" if signed_file else ("Ver Word" if document_format == "word" else "Ver PDF")
    download_label = "Descargar PDF firmado"
    review_note = "El documento para revisar y firmar incluye las observaciónes registradas."
    close_note = "Todas las firmas asignadas estan completas. Solo un administrador puede cerrar el documento y generar el archivo firmado final."
    return html_page(
        document["title"],
        render_template_string(
            """
            <h1>{{ document.title }}</h1>
            <div class="panel">
              <p><strong>Tipo:</strong> {{ document.type_name }}</p>
              <p><strong>Estado:</strong> <span class="status-pill {{ status_class(document.status, status_observations_count, signature_counts.required_count, signature_counts.signed_count) }}">{{ status_display(document.status, status_observations_count, signature_counts.required_count, signature_counts.signed_count) }}</span></p>
              <div class="inline-actions detail-actions">
                <a class="btn" href="{{ url_for('view_pdf', document_id=document.id) }}" data-pdf-modal {% if can_observe_doc %}data-observe-url="{{ url_for('observe_document', document_id=document.id) }}"{% endif %}>{{ view_label }}</a>
                {% if signed_file and g.user.can_download %}<a class="btn" href="{{ url_for('download_pdf', document_id=document.id) }}">{{ download_label }}</a>{% endif %}
                {% if signed_file and g.user.can_print %}<a class="btn" href="{{ url_for('view_pdf', document_id=document.id, print=1) }}">Imprimir</a>{% endif %}
                {% if can_sign_doc %}<button type="button" id="open_sign_modal">Firmar</button>{% endif %}
                <a class="btn push-right" href="{{ url_for('list_documents', searched=1, status='pending') }}">Volver</a>
              </div>
              {% if observations and document.status != "closed" %}
                <p class="muted">{{ review_note }}</p>
              {% endif %}
            </div>
            {% if observations %}
              <h2>Observaciónes del documento</h2>
              <table>
                <tr><th>Usuario</th><th>Cargo</th><th>Observación</th><th>Fecha</th><th>Respuesta administrador</th><th>Respondio</th></tr>
                {% for o in observations %}
                  <tr>
                    <td>{{ o.user_name }}</td><td>{{ o.role_name }}</td><td>{{ o.note }}</td><td>{{ o.created_at }}</td>
                    <td>{{ o.response or "Pendiente de respuesta" }}</td>
                    <td>{% if o.responded_at %}{{ o.responder_name }}<br>{{ o.responded_at }}{% else %}Pendiente{% endif %}</td>
                  </tr>
                {% endfor %}
              </table>
            {% endif %}
            {% if can_close_document %}
              <form class="panel" method="post" action="{{ url_for('close_document_manually', document_id=document.id) }}" onsubmit="return confirm('Cerrar este documento? Despues no podras modificar firmantes ni firmas.');">
                <h2>Cierre de documento</h2>
                <p class="muted">{{ close_note }}</p>
                <p><button>Cerrar documento</button></p>
              </form>
            {% elif can_force_close_document %}
              <div class="panel">
                <h2>Cierre de documento</h2>
                <p class="muted">Aun faltan firmas. El administrador puede forzar el cierre dejando una nota del motivo en la auditoria.</p>
                <form method="post" action="{{ url_for('force_close_document', document_id=document.id) }}" class="reason-after-action" data-reason-label="forzar el cierre de este documento">
                  <input type="hidden" name="reason">
                  <p><button class="danger">Forzar cierre</button></p>
                </form>
              </div>
            {% elif document.status == "pending" and signature_counts.required_count > 0 and signature_counts.required_count == signature_counts.signed_count %}
              <div class="panel">
                <p class="muted">Todas las firmas estan completas. El documento queda pendiente de cierre por administrador.</p>
              </div>
            {% endif %}
            {% if can_resolve_observed %}
              <div class="panel">
                <h2>Responder observación</h2>
                <p class="muted">El administrador debe dar conforme y responder la observación. Al guardar la respuesta, el documento vuelve a pendiente y queda listo para firmar.</p>
                <form method="post" action="{{ url_for('answer_observed_document', document_id=document.id) }}">
                  <label>Respuesta</label><textarea name="response" minlength="10" maxlength="300" required placeholder="Respuesta del administrador para dejar trazabilidad"></textarea>
                  <p><button>Guardar respuesta y dejar pendiente</button></p>
                </form>
                <form method="post" action="{{ url_for('annul_document', document_id=document.id) }}" class="reason-after-action" data-reason-label="anular este documento" style="display:inline">
                  <input type="hidden" name="reason">
                  <button class="danger">Anular documento</button>
                </form>
              </div>
            {% endif %}
            {% if can_manage_signers %}
              <div class="panel">
                <h2>Administrar firmantes</h2>
                <p class="muted">Puedes agregar firmantes hasta el máximo permitido para este tipo de documento. Solo se pueden quitar firmantes que aún no han firmado.</p>
                <form method="post" action="{{ url_for('add_document_signer', document_id=document.id) }}" class="reason-after-action" data-reason-label="habilitar este firmante">
                  <label>Agregar firmante</label>
                  <select name="user_id" required>
                    <option value="">Seleccionar</option>
                    {% for u in available_signers %}
                      <option value="{{ u.id }}">{{ u.name }} - {{ u.role_name }}</option>
                    {% endfor %}
                  </select>
                  <input type="hidden" name="reason">
                  <p><button>Habilitar firmante</button></p>
                </form>
                <p class="muted">Firmantes actuales: {{ signers|length }}/{{ max_signers }}</p>
              </div>
            {% endif %}
            <h2>Estado de firmas</h2>
            <table>
              <tr><th>Firmante</th><th>Cargo</th><th>Estado</th><th>Código</th><th>Fecha</th>{% if can_manage_signers %}<th>Acciones</th>{% endif %}</tr>
              {% for s in signers %}
                <tr>
                  <td>{{ s.name }}</td><td>{{ s.role_name }}</td>
                  <td>{{ "Firmado" if s.signed_at else "Pendiente" }}</td>
                  <td>{{ s.signature_code or "" }}</td><td>{{ s.signed_at or "" }}</td>
                  {% if can_manage_signers %}
                    <td>
                      {% if not s.signed_at %}
                        <form method="post" action="{{ url_for('remove_document_signer', document_id=document.id, user_id=s.id) }}" class="reason-after-action" data-reason-label="deshabilitar este firmante">
                          <input type="hidden" name="reason">
                          <button class="danger">Deshabilitar</button>
                        </form>
                      {% else %}
                        <span class="muted">No se puede quitar</span>
                      {% endif %}
                    </td>
                  {% endif %}
                </tr>
              {% endfor %}
            </table>
            {% if can_manage_signers or can_resolve_observed %}
              <div class="modal-backdrop" id="reason_modal">
                <div class="modal">
                  <h2 id="reason_title">Nota del motivo</h2>
                  <p class="muted">Debes escribir una nota alfanumerica de 10 a 80 caracteres para dejar registro en la auditoria.</p>
                  <textarea id="reason_text" maxlength="80" placeholder="Escribe aqui el motivo de la acción..."></textarea>
                  <p id="reason_error" class="flash" style="display:none; border-color:#a93636;background:#ffe4e4;color:#7a1f1f;">La nota debe tener entre 10 y 80 caracteres alfanumericos.</p>
                  <p>
                    <button type="button" id="reason_confirm">Guardar nota y continuar</button>
                    <button type="button" class="danger" id="reason_cancel">Cancelar</button>
                  </p>
                </div>
              </div>
            {% endif %}
            {% if can_sign_doc %}
              <div class="modal-backdrop" id="sign_modal">
                <div class="modal">
                  <h2>Firmar documento</h2>
                  <form method="post" action="{{ url_for('sign_document', document_id=document.id) }}">
                    <label>Clave</label><input name="password" type="password" required autocomplete="new-password">
                    <p>
                      <button>Firmar</button>
                      <button type="button" class="danger" id="cancel_sign_modal">Cancelar</button>
                    </p>
                  </form>
                </div>
              </div>
            {% endif %}
            <script>
              const signModal = document.getElementById("sign_modal");
              const openSignModal = document.getElementById("open_sign_modal");
              const cancelSignModal = document.getElementById("cancel_sign_modal");
              if (openSignModal && signModal) {
                openSignModal.addEventListener("click", () => {
                  signModal.classList.add("active");
                  const password = signModal.querySelector('input[name="password"]');
                  if (password) {
                    password.value = "";
                    password.focus();
                  }
                });
              }
              if (cancelSignModal && signModal) {
                cancelSignModal.addEventListener("click", () => {
                  signModal.classList.remove("active");
                });
              }
              const reasonModal = document.getElementById("reason_modal");
              const reasonTitle = document.getElementById("reason_title");
              const reasonText = document.getElementById("reason_text");
              const reasonError = document.getElementById("reason_error");
              const reasonConfirm = document.getElementById("reason_confirm");
              const reasonCancel = document.getElementById("reason_cancel");
              let pendingReasonForm = null;

              document.querySelectorAll(".reason-after-action").forEach((form) => {
                form.addEventListener("submit", (event) => {
                  const reasonInput = form.querySelector('input[name="reason"]');
                  if (/^[A-Za-z0-9ÁÉÍÓÚÜÑáéíóúüñ ]{10,80}$/.test(reasonInput.value.trim())) {
                    return;
                  }
                  event.preventDefault();
                  pendingReasonForm = form;
                  reasonTitle.textContent = "Nota para " + (form.dataset.reasonLabel || "realizar esta acción");
                  reasonText.value = "";
                  reasonError.style.display = "none";
                  reasonModal.classList.add("active");
                  reasonText.focus();
                });
              });

              if (reasonConfirm) {
                reasonConfirm.addEventListener("click", () => {
                  const reason = reasonText.value.trim();
                  if (!/^[A-Za-z0-9ÁÉÍÓÚÜÑáéíóúüñ ]{10,80}$/.test(reason)) {
                    reasonError.style.display = "block";
                    reasonText.focus();
                    return;
                  }
                  pendingReasonForm.querySelector('input[name="reason"]').value = reason;
                  reasonModal.classList.remove("active");
                  pendingReasonForm.submit();
                });
              }

              if (reasonCancel) {
                reasonCancel.addEventListener("click", () => {
                  pendingReasonForm = null;
                  reasonModal.classList.remove("active");
                });
              }
            </script>
            """,
            document=document,
            signers=signers,
            can_sign_doc=can_sign_doc,
            signed_file=signed_file,
            can_manage_signers=can_manage_signers,
            available_signers=available_signers,
            max_signers=max_signers,
            signature_counts=signature_counts,
            can_close_document=can_close_document,
            can_force_close_document=can_force_close_document,
            can_observe_doc=can_observe_doc,
            can_resolve_observed=can_resolve_observed,
            observations=observations,
            observations_count=observations_count,
            status_observations_count=status_observations_count,
            view_label=view_label,
            download_label=download_label,
            review_note=review_note,
            close_note=close_note,
            status_display=document_status_display,
            status_class=document_status_class,
        ),
    )


@app.route("/documents/<int:document_id>/pdf")
@login_required
def view_pdf(document_id: int):
    document = get_accessible_document(document_id)
    if document["status"] == "closed" and document["signed_path"] and Path(document["signed_path"]).suffix.lower() == ".pdf":
        audit("document_viewed", f"Visualizó PDF firmado {document['title']}", document_id=document_id)
        return send_file(document["signed_path"], mimetype="application/pdf", as_attachment=False)
    if document_file_format(document) == "word":
        source_path = document["signed_path"] if document["status"] == "closed" and document["signed_path"] else document["stored_path"]
        audit("document_viewed", f"Visualizó Word {document['title']}", document_id=document_id)
        return word_preview_html(document, source_path, request.args.get("print") == "1")
    if document["status"] == "closed" and document["signed_path"]:
        path = document["signed_path"]
        audit("document_viewed", f"Visualizó PDF {document['title']}", document_id=document_id)
        return send_file(path, mimetype="application/pdf", as_attachment=False)
    if document_has_observations(document_id):
        audit("document_viewed", f"Visualizó PDF con observaciónes {document['title']}", document_id=document_id)
        return send_file(
            BytesIO(pdf_with_observations(document_id)),
            mimetype="application/pdf",
            as_attachment=False,
            download_name=f"observado-{document['original_filename']}",
        )
    audit("document_viewed", f"Visualizó PDF {document['title']}", document_id=document_id)
    return send_file(document["stored_path"], mimetype="application/pdf", as_attachment=False)


@app.route("/documents/<int:document_id>/download")
@login_required
def download_pdf(document_id: int):
    if not g.user["can_download"] and not (g.user["is_admin"] or g.user["is_super_admin"]):
        abort(403)
    document = get_accessible_document(document_id)
    if document["status"] != "closed" or not document["signed_path"]:
        abort(404)
    audit("document_downloaded", f"Descargó documento firmado {document['title']}", document_id=document_id)
    return send_file(
        document["signed_path"],
        mimetype=document_mimetype(document),
        as_attachment=True,
        download_name=signed_download_name(document),
    )


def get_accessible_document(document_id: int) -> sqlite3.Row:
    where, params = document_access_sql()
    document = db().execute(
        f"SELECT d.*, dt.name AS type_name FROM documents d JOIN document_types dt ON dt.id = d.document_type_id WHERE d.id = ? AND {where}",
        (document_id, *params),
    ).fetchone()
    if not document:
        abort(404)
    return document


def get_admin_document(document_id: int) -> sqlite3.Row:
    document = db().execute(
        """
        SELECT d.*, dt.name AS type_name, dt.required_signatures AS max_signers
        FROM documents d
        JOIN document_types dt ON dt.id = d.document_type_id
        WHERE d.id = ?
        """,
        (document_id,),
    ).fetchone()
    if not document:
        abort(404)
    return document


@app.post("/admin/documents/<int:document_id>/signers/add")
@admin_required
def add_document_signer(document_id: int):
    document = get_admin_document(document_id)
    if document["status"] == "closed":
        flash("No se pueden modificar firmantes de un documento cerrado.")
        return redirect(url_for("document_detail", document_id=document_id))
    user_id = int(request.form["user_id"])
    reason = request.form.get("reason", "").strip()
    if not valid_reason_note(reason):
        flash("Debes ingresar una nota alfanumerica de entre 10 y 80 caracteres.")
        return redirect(url_for("document_detail", document_id=document_id))
    current_count = db().execute(
        "SELECT COUNT(*) AS total FROM document_required_signers WHERE document_id = ?",
        (document_id,),
    ).fetchone()["total"]
    if current_count >= document["max_signers"]:
        flash(f"Este documento ya tiene el máximo de {document['max_signers']} firmantes.")
        return redirect(url_for("document_detail", document_id=document_id))
    eligible = db().execute(
        """
        SELECT u.*, r.name AS role_name
        FROM users u
        JOIN roles r ON r.id = u.role_id
        WHERE u.id = ?
          AND u.active = 1
          AND u.can_sign = 1
          AND u.is_super_admin = 0
          AND EXISTS(
            SELECT 1 FROM document_type_signer_roles dtsr
            WHERE dtsr.document_type_id = ? AND dtsr.role_id = u.role_id
          )
          AND NOT EXISTS(
            SELECT 1 FROM document_required_signers rs
            WHERE rs.document_id = ? AND rs.user_id = u.id
          )
        """,
        (user_id, document["document_type_id"], document_id),
    ).fetchone()
    if not eligible:
        flash("El usuario no está autorizado, ya está asignado o no puede firmar este documento.")
        return redirect(url_for("document_detail", document_id=document_id))
    db().execute("INSERT INTO document_required_signers(document_id, user_id) VALUES (?, ?)", (document_id, user_id))
    db().commit()
    audit("document_signer_added", f"Firmante habilitado: {eligible['name']}. Motivo: {reason}", user_id, document_id)
    sent, failed = notify_document_signers(document_id, [user_id])
    audit("document_signers_notified", f"Nuevo firmante notificado. Correos enviados: {sent}. Pendientes/fallidos: {failed}", document_id=document_id)
    flash("Firmante habilitado para este documento.")
    return redirect(url_for("document_detail", document_id=document_id))


@app.post("/admin/documents/<int:document_id>/signers/<int:user_id>/remove")
@admin_required
def remove_document_signer(document_id: int, user_id: int):
    document = get_admin_document(document_id)
    if document["status"] == "closed":
        flash("No se pueden modificar firmantes de un documento cerrado.")
        return redirect(url_for("document_detail", document_id=document_id))
    reason = request.form.get("reason", "").strip()
    if not valid_reason_note(reason):
        flash("Debes ingresar una nota alfanumerica de entre 10 y 80 caracteres.")
        return redirect(url_for("document_detail", document_id=document_id))
    signature = db().execute(
        "SELECT 1 FROM signatures WHERE document_id = ? AND user_id = ?",
        (document_id, user_id),
    ).fetchone()
    if signature:
        flash("No se puede deshabilitar un firmante que ya firmo. La trazabilidad se conserva.")
        return redirect(url_for("document_detail", document_id=document_id))
    current_count = db().execute(
        "SELECT COUNT(*) AS total FROM document_required_signers WHERE document_id = ?",
        (document_id,),
    ).fetchone()["total"]
    if current_count <= 1:
        flash("El documento debe mantener al menos un firmante.")
        return redirect(url_for("document_detail", document_id=document_id))
    signer = db().execute("SELECT name FROM users WHERE id = ?", (user_id,)).fetchone()
    result = db().execute(
        "DELETE FROM document_required_signers WHERE document_id = ? AND user_id = ?",
        (document_id, user_id),
    )
    db().commit()
    if result.rowcount:
        audit("document_signer_removed", f"Firmante deshabilitado: {signer['name'] if signer else user_id}. Motivo: {reason}", user_id, document_id)
        flash("Firmante deshabilitado para este documento.")
    else:
        flash("El firmante no estaba asignado a este documento.")
    return redirect(url_for("document_detail", document_id=document_id))


@app.post("/documents/<int:document_id>/observe")
@login_required
def observe_document(document_id: int):
    document = get_accessible_document(document_id)
    if document["status"] != "pending" or not g.user["can_sign"]:
        abort(403)
    is_required = db().execute(
        "SELECT 1 FROM document_required_signers WHERE document_id = ? AND user_id = ?",
        (document_id, g.user["id"]),
    ).fetchone()
    already_signed = db().execute(
        "SELECT 1 FROM signatures WHERE document_id = ? AND user_id = ?",
        (document_id, g.user["id"]),
    ).fetchone()
    if not is_required or already_signed:
        abort(403)
    note = request.form.get("note", "").strip()
    if not valid_observation_text(note):
        flash("La observación debe tener entre 10 y 300 caracteres.")
        return redirect(url_for("document_detail", document_id=document_id))
    db().execute(
        "INSERT INTO document_observations(document_id, user_id, note, created_at) VALUES (?, ?, ?, ?)",
        (document_id, g.user["id"], note, now()),
    )
    db().execute("UPDATE documents SET status = 'observed' WHERE id = ?", (document_id,))
    db().commit()
    audit("document_observed", f"Documento observado. Motivo: {note}", g.user["id"], document_id)
    flash("Documento observado. Queda pendiente de resolucion administrativa.")
    return redirect(url_for("document_detail", document_id=document_id))


@app.post("/documents/<int:document_id>/sign")
@login_required
def sign_document(document_id: int):
    document = get_accessible_document(document_id)
    next_url = request.form.get("next", "").strip()
    redirect_target = next_url if next_url.startswith("/") and not next_url.startswith("//") else url_for("document_detail", document_id=document_id)
    if document["status"] == "closed" or not g.user["can_sign"]:
        abort(403)
    is_required = db().execute(
        "SELECT 1 FROM document_required_signers WHERE document_id = ? AND user_id = ?",
        (document_id, g.user["id"]),
    ).fetchone()
    already_signed = db().execute(
        "SELECT 1 FROM signatures WHERE document_id = ? AND user_id = ?",
        (document_id, g.user["id"]),
    ).fetchone()
    if not is_required or already_signed:
        abort(403)
    if not check_password_hash(g.user["password_hash"], request.form["password"]):
        flash("Clave incorrecta.")
        return redirect(redirect_target)
    code = random_signature_code()
    db().execute(
        "INSERT INTO signatures(document_id, user_id, signature_code, signed_at) VALUES (?, ?, ?, ?)",
        (document_id, g.user["id"], code, now()),
    )
    db().commit()
    audit("document_signed", f"Firma registrada con código {code}", g.user["id"], document_id)
    flash("Firma registrada correctamente.")
    return redirect(redirect_target)


def document_signature_counts(document_id: int) -> sqlite3.Row:
    return db().execute(
        """
        SELECT
          (SELECT COUNT(*) FROM document_required_signers WHERE document_id = ?) AS required_count,
          (SELECT COUNT(*) FROM signatures WHERE document_id = ?) AS signed_count
        """,
        (document_id, document_id),
    ).fetchone()


def document_has_observations(document_id: int) -> bool:
    return bool(
        db().execute(
            "SELECT 1 FROM document_observations WHERE document_id = ? LIMIT 1",
            (document_id,),
        ).fetchone()
    )


def document_file_format(document: sqlite3.Row) -> str:
    suffix = Path(document["original_filename"]).suffix.lower()
    return "word" if suffix == ".docx" else "pdf"


def document_mimetype(document: sqlite3.Row) -> str:
    path = document["signed_path"] or document["stored_path"]
    return WORD_MIMETYPE if Path(path).suffix.lower() == ".docx" else "application/pdf"


def signed_download_name(document: sqlite3.Row) -> str:
    original = Path(document["original_filename"])
    suffix = Path(document["signed_path"] or original.name).suffix.lower()
    return f"firmado-{original.stem}{suffix}"


def docx_paragraphs(path: str) -> list[str]:
    try:
        with zipfile.ZipFile(path) as archive:
            xml_bytes = archive.read("word/document.xml")
    except (KeyError, zipfile.BadZipFile):
        return ["No se pudo leer el contenido del documento Word."]

    namespace = {"w": "http://schemas.openxmlformats.org/wordprocessingml/2006/main"}
    root = ET.fromstring(xml_bytes)
    paragraphs: list[str] = []
    for paragraph in root.findall(".//w:p", namespace):
        text = "".join(node.text or "" for node in paragraph.findall(".//w:t", namespace)).strip()
        if text:
            paragraphs.append(text)
    return paragraphs or ["Documento Word sin texto visible para previsualizar."]


def word_preview_html(document: sqlite3.Row, source_path: str, print_now: bool = False) -> str:
    observations = db().execute(
        """
        SELECT u.name, r.name AS role_name, o.note, o.created_at, o.response, o.responded_at, responder.name AS responder_name
        FROM document_observations o
        JOIN users u ON u.id = o.user_id
        JOIN roles r ON r.id = u.role_id
        LEFT JOIN users responder ON responder.id = o.responded_by_user_id
        WHERE o.document_id = ?
        ORDER BY o.created_at
        """,
        (document["id"],),
    ).fetchall()
    paragraphs = docx_paragraphs(source_path)
    body = "\n".join(f"<p>{html.escape(text)}</p>" for text in paragraphs)
    observations_html = ""
    if observations:
        rows = "\n".join(
            "<tr>"
            f"<td>{html.escape(obs['name'])}</td>"
            f"<td>{html.escape(obs['role_name'])}</td>"
            f"<td>{html.escape(obs['note'])}</td>"
            f"<td>{html.escape(obs['created_at'])}</td>"
            f"<td>{html.escape(obs['response'] or 'Pendiente de respuesta')}</td>"
            f"<td>{html.escape((obs['responder_name'] or '') + (' | ' + obs['responded_at'] if obs['responded_at'] else ''))}</td>"
            "</tr>"
            for obs in observations
        )
        observations_html = f"""
        <h2>Observaciónes asociadas</h2>
        <table>
          <tr><th>Usuario</th><th>Cargo</th><th>Observación</th><th>Fecha</th><th>Respuesta</th><th>Respondio</th></tr>
          {rows}
        </table>
        """
    print_script = "<script>window.addEventListener('load', () => window.print());</script>" if print_now else ""
    return html_page(
        document["title"],
        f"""
        <h1>{html.escape(document['title'])}</h1>
        <div class="panel">
          <p class="muted">Vista de solo lectura del documento Word. El archivo original no se edita hasta que un administrador cierre el documento.</p>
        </div>
        <div class="panel word-preview">
          {body}
        </div>
        {observations_html}
        {print_script}
        """,
    )


@app.post("/admin/documents/<int:document_id>/close")
@admin_required
def close_document_manually(document_id: int):
    document = get_admin_document(document_id)
    if document["status"] == "closed":
        flash("El documento ya esta cerrado.")
        return redirect(url_for("document_detail", document_id=document_id))
    if document["status"] != "pending":
        flash("Este cierre solo aplica a documentos pendientes con todas sus firmas.")
        return redirect(url_for("document_detail", document_id=document_id))
    counts = document_signature_counts(document_id)
    if counts["required_count"] == 0:
        flash("El documento debe tener al menos un firmante para poder cerrarse.")
        return redirect(url_for("document_detail", document_id=document_id))
    if counts["required_count"] != counts["signed_count"]:
        flash("No se puede cerrar. Aun faltan firmas pendientes.")
        return redirect(url_for("document_detail", document_id=document_id))
    signed_path = stamp_document(document_id)
    db().execute(
        "UPDATE documents SET status = 'closed', signed_path = ?, closed_at = ? WHERE id = ?",
        (signed_path, now(), document_id),
    )
    db().commit()
    audit("document_closed", "Documento cerrado manualmente por administrador con todas las firmas", document_id=document_id)
    flash("Documento cerrado correctamente por administrador.")
    return redirect(url_for("document_detail", document_id=document_id))


@app.post("/admin/documents/<int:document_id>/force-close")
@admin_required
def force_close_document(document_id: int):
    document = get_admin_document(document_id)
    if document["status"] == "closed":
        flash("El documento ya esta cerrado.")
        return redirect(url_for("document_detail", document_id=document_id))
    if document["status"] != "pending":
        flash("Solo se puede forzar el cierre de documentos pendientes.")
        return redirect(url_for("document_detail", document_id=document_id))
    unresolved = db().execute(
        "SELECT COUNT(*) AS total FROM document_observations WHERE document_id = ? AND responded_at IS NULL",
        (document_id,),
    ).fetchone()["total"]
    if unresolved:
        flash("Primero debes responder la observación antes de forzar el cierre.")
        return redirect(url_for("document_detail", document_id=document_id))
    counts = document_signature_counts(document_id)
    if counts["required_count"] == 0 or counts["signed_count"] == 0:
        flash("El documento debe tener al menos una firma registrada para forzar el cierre.")
        return redirect(url_for("document_detail", document_id=document_id))
    if counts["signed_count"] >= counts["required_count"]:
        flash("Todas las firmas estan completas. Usa Cerrar documento.")
        return redirect(url_for("document_detail", document_id=document_id))
    reason = request.form.get("reason", "").strip()
    if not valid_reason_note(reason):
        flash("La nota debe tener entre 10 y 80 caracteres alfanumericos.")
        return redirect(url_for("document_detail", document_id=document_id))
    signed_path = stamp_document(document_id)
    db().execute(
        "UPDATE documents SET status = 'closed', signed_path = ?, closed_at = ? WHERE id = ?",
        (signed_path, now(), document_id),
    )
    db().commit()
    audit(
        "document_force_closed",
        f"Documento cerrado forzadamente por administrador. Firmas: {counts['signed_count']}/{counts['required_count']}. Motivo: {reason}",
        document_id=document_id,
    )
    flash("Documento cerrado forzadamente. La nota quedo registrada en auditoria.")
    return redirect(url_for("document_detail", document_id=document_id))


@app.post("/admin/documents/<int:document_id>/close-observed")
@admin_required
def close_observed_document(document_id: int):
    document = get_admin_document(document_id)
    if document["status"] != "observed":
        flash("Solo se pueden cerrar como observados los documentos en estado observado.")
        return redirect(url_for("document_detail", document_id=document_id))
    unresolved = db().execute(
        "SELECT COUNT(*) AS total FROM document_observations WHERE document_id = ? AND responded_at IS NULL",
        (document_id,),
    ).fetchone()["total"]
    if unresolved:
        flash("Primero debes responder la observación antes de cerrar o continuar el documento.")
        return redirect(url_for("document_detail", document_id=document_id))
    signed_path = stamp_document(document_id)
    db().execute(
        "UPDATE documents SET status = 'closed', signed_path = ?, closed_at = ? WHERE id = ?",
        (signed_path, now(), document_id),
    )
    db().commit()
    audit("document_closed_observed", "Documento observado cerrado administrativamente con observaciónes", document_id=document_id)
    flash("Documento observado cerrado administrativamente.")
    return redirect(url_for("document_detail", document_id=document_id))


@app.post("/admin/documents/<int:document_id>/answer-observation")
@admin_required
def answer_observed_document(document_id: int):
    document = get_admin_document(document_id)
    if document["status"] != "observed":
        flash("Solo se pueden responder observaciónes de documentos observados.")
        return redirect(url_for("document_detail", document_id=document_id))
    response = request.form.get("response", "").strip()
    if not valid_observation_text(response):
        flash("La respuesta debe tener entre 10 y 300 caracteres.")
        return redirect(url_for("document_detail", document_id=document_id))
    pending = db().execute(
        "SELECT COUNT(*) AS total FROM document_observations WHERE document_id = ? AND responded_at IS NULL",
        (document_id,),
    ).fetchone()["total"]
    if pending == 0:
        flash("Este documento no tiene observaciónes pendientes de respuesta.")
        return redirect(url_for("document_detail", document_id=document_id))
    db().execute(
        """
        UPDATE document_observations
        SET response = ?, responded_by_user_id = ?, responded_at = ?
        WHERE document_id = ? AND responded_at IS NULL
        """,
        (response, g.user["id"], now(), document_id),
    )
    db().execute("UPDATE documents SET status = 'pending' WHERE id = ?", (document_id,))
    db().commit()
    audit("document_observation_answered", f"Observación respondida por administrador. Respuesta: {response}", g.user["id"], document_id)
    flash("Respuesta registrada. El documento vuelve a pendiente y queda listo para firmar.")
    return redirect(url_for("document_detail", document_id=document_id))


@app.post("/admin/documents/<int:document_id>/reopen")
@admin_required
def reopen_document(document_id: int):
    document = get_admin_document(document_id)
    if document["status"] != "observed":
        flash("Solo se pueden reabrir documentos observados.")
        return redirect(url_for("document_detail", document_id=document_id))
    unresolved = db().execute(
        "SELECT COUNT(*) AS total FROM document_observations WHERE document_id = ? AND responded_at IS NULL",
        (document_id,),
    ).fetchone()["total"]
    if unresolved:
        flash("Para volver a pendiente debes responder la observación desde el formulario administrativo.")
        return redirect(url_for("document_detail", document_id=document_id))
    reason = request.form.get("reason", "").strip()
    if not valid_reason_note(reason):
        flash("Debes ingresar una nota alfanumerica de entre 10 y 80 caracteres.")
        return redirect(url_for("document_detail", document_id=document_id))
    db().execute("UPDATE documents SET status = 'pending' WHERE id = ?", (document_id,))
    db().commit()
    audit("document_reopened", f"Documento observado reabierto para firma. Motivo: {reason}", document_id=document_id)
    flash("Documento reabierto para firma.")
    return redirect(url_for("document_detail", document_id=document_id))


@app.post("/admin/documents/<int:document_id>/annul")
@admin_required
def annul_document(document_id: int):
    document = get_admin_document(document_id)
    if document["status"] not in {"pending", "observed"}:
        flash("Solo se pueden anular documentos pendientes u observados.")
        return redirect(url_for("document_detail", document_id=document_id))
    reason = request.form.get("reason", "").strip()
    if not valid_reason_note(reason):
        flash("Debes ingresar una nota alfanumerica de entre 10 y 80 caracteres.")
        return redirect(url_for("document_detail", document_id=document_id))
    db().execute("UPDATE documents SET status = 'annulled', closed_at = ? WHERE id = ?", (now(), document_id))
    db().commit()
    audit("document_annulled", f"Documento anulado administrativamente. Motivo: {reason}", document_id=document_id)
    flash("Documento anulado administrativamente.")
    return redirect(url_for("document_detail", document_id=document_id))


def pdf_with_observations(document_id: int) -> bytes:
    document = db().execute("SELECT * FROM documents WHERE id = ?", (document_id,)).fetchone()
    observations = db().execute(
        """
        SELECT u.name, r.name AS role_name, o.note, o.created_at, o.response, o.responded_at, responder.name AS responder_name
        FROM document_observations o
        JOIN users u ON u.id = o.user_id
        JOIN roles r ON r.id = u.role_id
        LEFT JOIN users responder ON responder.id = o.responded_by_user_id
        WHERE o.document_id = ?
        ORDER BY o.created_at
        """,
        (document_id,),
    ).fetchall()

    reader = PdfReader(document["stored_path"])
    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)

    packet = BytesIO()
    c = canvas.Canvas(packet, pagesize=letter)
    c.setFillColor(HexColor("#a93636"))
    c.setFont("Helvetica-Bold", 11)
    c.drawString(42, 735, "Observaciónes vigentes antes de firma")
    c.setFillColor(HexColor("#16202a"))
    c.setFont("Helvetica", 8)
    y = 718
    for idx, obs in enumerate(observations, start=1):
        for text in observation_stamp_lines(idx, obs):
            if y < 40:
                c.showPage()
                y = 735
                c.setFont("Helvetica", 8)
            c.drawString(42, y, text[:130])
            y -= 12
        y -= 4
    c.save()
    packet.seek(0)
    overlay = PdfReader(packet)
    if overlay.pages:
        writer.pages[0].merge_page(overlay.pages[0])
        for i in range(1, min(len(writer.pages), len(overlay.pages))):
            writer.pages[i].merge_page(overlay.pages[i])

    output = BytesIO()
    writer.write(output)
    return output.getvalue()


def document_signature_rows(document_id: int) -> list[sqlite3.Row]:
    return db().execute(
        """
        SELECT u.name, u.signature_order, r.name AS role_name, s.signature_code, s.signed_at
        FROM signatures s
        JOIN users u ON u.id = s.user_id
        JOIN roles r ON r.id = u.role_id
        WHERE s.document_id = ?
        ORDER BY u.signature_order, u.name, s.signed_at
        """,
        (document_id,),
    ).fetchall()


def document_observation_rows(document_id: int) -> list[sqlite3.Row]:
    return db().execute(
        """
        SELECT u.name, r.name AS role_name, o.note, o.created_at, o.response, o.responded_at, responder.name AS responder_name
        FROM document_observations o
        JOIN users u ON u.id = o.user_id
        JOIN roles r ON r.id = u.role_id
        LEFT JOIN users responder ON responder.id = o.responded_by_user_id
        WHERE o.document_id = ?
        ORDER BY o.created_at
        """,
        (document_id,),
    ).fetchall()


def signature_stamp_lines(index: int, signature: sqlite3.Row) -> tuple[str, str, str]:
    signed_at = signature["signed_at"] or ""
    date_part, time_part = (signed_at.split(" ", 1) + [""])[:2] if " " in signed_at else (signed_at, "")
    return signature["name"], signature["role_name"], f"{signature['signature_code']} {date_part} {time_part}".strip()


def observation_stamp_lines(index: int, observation: sqlite3.Row) -> list[str]:
    lines = [
        f"{index}. {observation['name']} | Cargo {observation['role_name']} | {observation['created_at']}",
        f"Observación: {observation['note']}",
    ]
    if observation["response"]:
        lines.append(f"Respuesta: {observation['response']}")
        lines.append(f"Respondio: {observation['responder_name']} | {observation['responded_at']}")
    else:
        lines.append("Respuesta: pendiente de administrador")
    return lines


def stamp_document(document_id: int) -> str:
    document = db().execute("SELECT * FROM documents WHERE id = ?", (document_id,)).fetchone()
    if document_file_format(document) == "word":
        return stamp_word_to_pdf(document_id)
    return stamp_pdf(document_id)


def draw_wrapped_text(c: canvas.Canvas, text: str, x: int, y: int, max_chars: int = 95, line_height: int = 12) -> int:
    words = text.split()
    line = ""
    for word in words or [""]:
        candidate = f"{line} {word}".strip()
        if len(candidate) > max_chars and line:
            if y < 45:
                c.showPage()
                c.setFont("Helvetica", 9)
                y = 735
            c.drawString(x, y, line)
            y -= line_height
            line = word
        else:
            line = candidate
    if line:
        if y < 45:
            c.showPage()
            c.setFont("Helvetica", 9)
            y = 735
        c.drawString(x, y, line)
        y -= line_height
    return y


def split_pdf_words(text: str, c: canvas.Canvas, max_width: int, font_name: str = "Helvetica", font_size: int = 9) -> list[list[str]]:
    lines: list[list[str]] = []
    current: list[str] = []
    for word in text.split():
        candidate = current + [word]
        if c.stringWidth(" ".join(candidate), font_name, font_size) > max_width and current:
            lines.append(current)
            current = [word]
        else:
            current = candidate
    if current:
        lines.append(current)
    return lines or [[]]


def draw_aligned_text(c: canvas.Canvas, text: str, y: int, font_name: str, font_size: int, alignment: str, x: int = 42, max_width: int = 528) -> None:
    width = c.stringWidth(text, font_name, font_size)
    if alignment == "right":
        draw_x = x + max_width - width
    elif alignment == "center":
        draw_x = x + (max_width - width) / 2
    else:
        draw_x = x
    c.drawString(draw_x, y, text)


def draw_config_image(c: canvas.Canvas, image_path: str, position: str) -> None:
    if not image_path:
        return
    path = Path(image_path)
    if not path.exists():
        return
    width, height = 88, 58
    positions = {
        "top_left": (42, 700),
        "top_center": ((letter[0] - width) / 2, 700),
        "top_right": (letter[0] - width - 42, 700),
        "center_left": (42, 360),
        "center": ((letter[0] - width) / 2, 360),
        "center_right": (letter[0] - width - 42, 360),
        "bottom_left": (42, 55),
        "bottom_center": ((letter[0] - width) / 2, 55),
        "bottom_right": (letter[0] - width - 42, 55),
    }
    x, y = positions.get(position, positions["top_right"])
    c.drawImage(str(path), x, y, width=width, height=height, preserveAspectRatio=True, mask="auto")


def draw_document_watermark(c: canvas.Canvas, image_path: str, size: int, opacity: float = 0.15) -> None:
    if not image_path:
        return
    path = Path(image_path)
    if not path.exists():
        return
    image_size = min(600, max(40, int(size)))
    x = (letter[0] - image_size) / 2
    y = (letter[1] - image_size) / 2
    c.saveState()
    try:
        c.setFillAlpha(opacity)
        c.setStrokeAlpha(opacity)
    except AttributeError:
        pass
    c.drawImage(str(path), x, y, width=image_size, height=image_size, preserveAspectRatio=True, mask="auto")
    c.restoreState()


def draw_justified_paragraph(
    c: canvas.Canvas,
    text: str,
    x: int,
    y: int,
    max_width: int,
    line_height: int = 13,
    font_name: str = "Helvetica",
    font_size: int = 9,
    on_new_page=None,
) -> int:
    lines = split_pdf_words(text, c, max_width, font_name, font_size)
    for index, words in enumerate(lines):
        if y < 45:
            c.showPage()
            c.setFillColor(HexColor("#16202a"))
            if on_new_page:
                on_new_page()
            c.setFont(font_name, font_size)
            y = 735
        if not words:
            y -= line_height
            continue
        is_last = index == len(lines) - 1
        if is_last or len(words) == 1:
            c.drawString(x, y, " ".join(words))
        else:
            words_width = sum(c.stringWidth(word, font_name, font_size) for word in words)
            extra_space = (max_width - words_width) / (len(words) - 1)
            cursor_x = x
            for word in words:
                c.drawString(cursor_x, y, word)
                cursor_x += c.stringWidth(word, font_name, font_size) + extra_space
        y -= line_height
    return y


def create_acta_pdf(acta_id: int) -> str:
    acta = db().execute("SELECT * FROM acta_documents WHERE id = ?", (acta_id,)).fetchone()
    if not acta:
        abort(404)
    filename = safe_display_filename(f"{acta['title']}.pdf")
    pdf_path = SIGNED_DIR / filename
    settings = system_settings()
    font_name = settings["acta_font_family"]
    body_font_size = int(settings["acta_body_font_size"])
    title_font_size = int(settings["acta_title_font_size"])
    title_font_name = f"{font_name}-Bold" if font_name in {"Helvetica", "Courier"} else "Times-Bold"
    line_height = max(body_font_size + 4, 12)
    c = canvas.Canvas(str(pdf_path), pagesize=letter)
    c.setFillColor(HexColor("#16202a"))
    def draw_acta_watermark() -> None:
        if settings.get("acta_watermark_enabled") == "1":
            draw_document_watermark(
                c,
                settings.get("acta_watermark_image_path", ""),
                int(settings["acta_watermark_size"]),
                int(settings["acta_watermark_opacity"]) / 100,
            )

    draw_acta_watermark()
    if settings.get("acta_image_enabled") == "1":
        draw_config_image(
            c,
            settings.get("acta_image_path", ""),
            settings["acta_image_position"],
        )
    c.setFont(title_font_name, title_font_size)
    y = 735
    for title_line in (settings.get("acta_title_line_1", "").strip(), settings.get("acta_title_line_2", "").strip(), acta["title"][:95]):
        if title_line:
            draw_aligned_text(c, title_line, y, title_font_name, title_font_size, settings["acta_title_alignment"])
            y -= title_font_size + 5
    y -= 8
    c.setFont(font_name, body_font_size)
    for paragraph in acta["body"].splitlines():
        if paragraph.strip():
            y = draw_justified_paragraph(
                c,
                paragraph,
                42,
                y,
                528,
                line_height=line_height,
                font_name=font_name,
                font_size=body_font_size,
                on_new_page=draw_acta_watermark,
            )
            y -= 4
        else:
            y -= line_height
            if y < 45:
                c.showPage()
                c.setFillColor(HexColor("#16202a"))
                draw_acta_watermark()
                c.setFont(font_name, body_font_size)
                y = 735
    c.save()
    return str(pdf_path)


def stamp_word_to_pdf(document_id: int) -> str:
    document = db().execute("SELECT * FROM documents WHERE id = ?", (document_id,)).fetchone()
    signatures = document_signature_rows(document_id)
    observations = document_observation_rows(document_id)
    signed_path = SIGNED_DIR / f"signed-{document_id}-{Path(document['original_filename']).stem}.pdf"

    c = canvas.Canvas(str(signed_path), pagesize=letter)
    c.setFillColor(HexColor("#16202a"))
    c.setFont("Helvetica-Bold", 14)
    c.drawString(42, 735, document["title"][:80])
    c.setFont("Helvetica", 9)
    y = 710
    for paragraph in docx_paragraphs(document["stored_path"]):
        y = draw_wrapped_text(c, paragraph, 42, y)
        y -= 4

    if observations:
        if y < 70:
            c.showPage()
            y = 735
        c.setFillColor(HexColor("#a93636"))
        c.setFont("Helvetica-Bold", 10)
        c.drawString(42, y, "Observaciónes registradas")
        y -= 14
        c.setFillColor(HexColor("#16202a"))
        c.setFont("Helvetica", 8)
        for idx, obs in enumerate(observations, start=1):
            for text in observation_stamp_lines(idx, obs):
                y = draw_wrapped_text(c, text, 42, y, max_chars=120)
            y -= 4

    if y < 90:
        c.showPage()
        y = 735
    c.setFillColor(HexColor("#146c5f"))
    c.setFont("Helvetica-Bold", 11)
    c.drawString(42, y, "Registro de firmas digitales internas")
    y -= 16
    c.setFillColor(HexColor("#16202a"))
    c.setFont("Helvetica", 8)
    for idx, sig in enumerate(signatures, start=1):
        for line in signature_stamp_lines(idx, sig):
            y = draw_wrapped_text(c, line, 42, y, max_chars=120)
        y -= 5

    c.save()
    return str(signed_path)


def stamp_pdf(document_id: int) -> str:
    document = db().execute("SELECT * FROM documents WHERE id = ?", (document_id,)).fetchone()
    signatures = document_signature_rows(document_id)
    observations = document_observation_rows(document_id)

    reader = PdfReader(document["stored_path"])
    writer = PdfWriter()
    for page in reader.pages:
        writer.add_page(page)

    signature_page = build_pdf_signature_page(signatures, observations)
    for page in PdfReader(BytesIO(signature_page)).pages:
        writer.add_page(page)

    signed_path = SIGNED_DIR / f"signed-{document_id}-{Path(document['original_filename']).name}"
    with signed_path.open("wb") as fh:
        writer.write(fh)
    return str(signed_path)


def build_pdf_signature_page(signatures: list[sqlite3.Row], observations: list[sqlite3.Row]) -> bytes:
    packet = BytesIO()
    c = canvas.Canvas(packet, pagesize=letter)
    y = 735
    if observations:
        if y < 70:
            c.showPage()
            y = 735
        c.setFillColor(HexColor("#a93636"))
        c.setFont("Helvetica-Bold", 10)
        c.drawString(42, y, "Observaciónes registradas")
        y -= 14
        c.setFillColor(HexColor("#16202a"))
        c.setFont("Helvetica", 8)
        for idx, obs in enumerate(observations, start=1):
            for text in observation_stamp_lines(idx, obs):
                if y < 40:
                    c.showPage()
                    y = 735
                    c.setFont("Helvetica", 8)
                c.drawString(42, y, text[:130])
                y -= 12
            y -= 4
    if y < 80:
        c.showPage()
        y = 735
    c.setFillColor(HexColor("#146c5f"))
    c.setFont("Helvetica-Bold", 11)
    c.drawString(42, y, "Registro de firmas digitales internas")
    y -= 17
    c.setFillColor(HexColor("#16202a"))
    c.setFont("Helvetica", 8)
    for idx, sig in enumerate(signatures, start=1):
        column = (idx - 1) % 2
        x = 42 if column == 0 else 320
        if column == 0 and y < 60:
            c.showPage()
            y = 735
            c.setFont("Helvetica", 8)
        name_line, role_line, code_line = signature_stamp_lines(idx, sig)
        c.setFont("Helvetica-Bold", 9)
        c.drawString(x, y, name_line[:52])
        c.setFont("Helvetica", 8)
        c.drawString(x, y - 13, role_line[:54])
        c.drawString(x, y - 26, code_line[:64])
        if column == 1:
            y -= 48
    if len(signatures) % 2 == 1:
        y -= 48
    c.save()
    return packet.getvalue()


def word_paragraph(text: str) -> ET.Element:
    w = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    paragraph = ET.Element(w + "p")
    run = ET.SubElement(paragraph, w + "r")
    text_node = ET.SubElement(run, w + "t")
    text_node.text = text
    return paragraph


def append_word_audit_section(document_xml: bytes, document_id: int) -> bytes:
    w = "{http://schemas.openxmlformats.org/wordprocessingml/2006/main}"
    ET.register_namespace("w", "http://schemas.openxmlformats.org/wordprocessingml/2006/main")
    root = ET.fromstring(document_xml)
    body = root.find(w + "body")
    if body is None:
        return document_xml

    insert_index = len(body)
    if insert_index and body[insert_index - 1].tag == w + "sectPr":
        insert_index -= 1

    signatures = document_signature_rows(document_id)
    observations = document_observation_rows(document_id)
    lines = []
    if observations:
        lines.extend(["", "Observaciónes registradas"])
        for idx, obs in enumerate(observations, start=1):
            lines.extend(observation_stamp_lines(idx, obs))
    lines.extend(["", "Registro de firmas digitales internas"])
    for idx, sig in enumerate(signatures, start=1):
        lines.extend(signature_stamp_lines(idx, sig))

    for offset, line in enumerate(lines):
        body.insert(insert_index + offset, word_paragraph(line))
    return ET.tostring(root, encoding="utf-8", xml_declaration=True)


def stamp_word(document_id: int) -> str:
    document = db().execute("SELECT * FROM documents WHERE id = ?", (document_id,)).fetchone()
    signed_path = SIGNED_DIR / f"signed-{document_id}-{Path(document['original_filename']).name}"
    with zipfile.ZipFile(document["stored_path"], "r") as source, zipfile.ZipFile(signed_path, "w", zipfile.ZIP_DEFLATED) as target:
        for item in source.infolist():
            data = source.read(item.filename)
            if item.filename == "word/document.xml":
                data = append_word_audit_section(data, document_id)
            target.writestr(item, data)
    return str(signed_path)


@app.route("/admin")
@login_required
def admin_home():
    return html_page(
        "Administración",
        render_template_string(
            """
        <h1>Administración</h1>
        <div class="grid">
          {% if g.user.is_admin or g.user.is_super_admin %}
            <a class="btn" href="/admin/users">Usuarios</a>
            <a class="btn" href="/admin/roles">Cargos</a>
            <a class="btn" href="/admin/document-types">Tipos de documentos</a>
            <a class="btn" href="/admin/history">Informes de usuarios</a>
            <a class="btn" href="/admin/emails">Bandeja de correos</a>
          {% endif %}
            <a class="btn" href="/admin/system-config">Configuración</a>
        </div>
        """
        ),
    )


@app.route("/admin/system-config", methods=["GET", "POST"])
@login_required
def system_config():
    if request.method == "POST":
        action = request.form.get("action", "")
        if action == "change_password":
            current_password = request.form.get("current_password", "")
            new_password = request.form.get("new_password", "")
            confirm_password = request.form.get("confirm_password", "")
            if not current_password:
                flash("Debes ingresar la clave anterior.")
                return redirect(url_for("system_config"))
            if not check_password_hash(g.user["password_hash"], current_password):
                attempts = int(session.get("system_config_password_attempts", 0)) + 1
                session["system_config_password_attempts"] = attempts
                remaining = max(0, 4 - attempts)
                if attempts >= 4:
                    sent, failed = notify_admins_password_change_alert(g.user)
                    audit(
                        "password_change_blocked",
                        f"Cuarto intento fallido de cambio de clave. Alertas enviadas: {sent}. Fallidas: {failed}",
                        g.user["id"],
                    )
                    session.clear()
                    flash("Se agotaron los intentos. Se aviso a los administradores y la sesion fue cerrada.")
                    return redirect(url_for("login"))
                flash(f"Clave anterior incorrecta. Te quedan {remaining} oportunidades.")
                return redirect(url_for("system_config"))
            session.pop("system_config_password_attempts", None)
            if not new_password or not confirm_password:
                flash("Debes ingresar y confirmar la nueva clave.")
                return redirect(url_for("system_config"))
            if not valid_password(new_password):
                flash(password_help())
                return redirect(url_for("system_config"))
            if new_password != confirm_password:
                flash("Claves no coinciden")
                return redirect(url_for("system_config"))
            db().execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(new_password), g.user["id"]))
            db().commit()
            audit("password_changed", "Clave cambiada desde configuración del sistema", g.user["id"])
            flash("Se guardaron los cambios.")
            return redirect(url_for("system_config"))

        if action not in {"aesthetic_settings", "smtp_settings", "smtp_test", "document_options"}:
            flash("Acción no válida.")
            return redirect(url_for("system_config"))
        if not (g.user["is_admin"] or g.user["is_super_admin"]):
            abort(403)
        if action == "document_options":
            ensure_document_type_options()
            for doc_type in db().execute("SELECT id, name FROM document_types ORDER BY name").fetchall():
                value = request.form.get(f"next_correlative_{doc_type['id']}", "1").strip() or "1"
                if len(value) > 10 or not value.isdigit() or int(value) < 1:
                    flash(f"El correlativo de {doc_type['name']} debe ser un número de 1 a 10 dígitos mayor a cero.")
                    return redirect(url_for("system_config"))
                db().execute(
                    """
                    INSERT INTO document_type_options(document_type_id, next_correlative)
                    VALUES (?, ?)
                    ON CONFLICT(document_type_id) DO UPDATE SET next_correlative = excluded.next_correlative
                    """,
                    (doc_type["id"], int(value)),
                )
            title_alignment = request.form.get("acta_title_alignment", "")
            acta_image_position = request.form.get("acta_image_position", "")
            acta_font_family = request.form.get("acta_font_family", "")
            body_font_size = request.form.get("acta_body_font_size", "9").strip()
            title_font_size = request.form.get("acta_title_font_size", "13").strip()
            watermark_size = request.form.get("acta_watermark_size", "260").strip()
            watermark_opacity = request.form.get("acta_watermark_opacity", "15").strip()
            title_line_1 = request.form.get("acta_title_line_1", "").strip()
            title_line_2 = request.form.get("acta_title_line_2", "").strip()
            if len(title_line_1) > 70 or len(title_line_2) > 70:
                flash("Los títulos del documento deben tener como máximo 70 caracteres.")
                return redirect(url_for("system_config"))
            if len(watermark_size) > 7 or len(watermark_opacity) > 7:
                flash("Los campos de tamaño de imágenes deben tener como máximo 7 dígitos.")
                return redirect(url_for("system_config"))
            if (
                title_alignment not in ACTA_TITLE_ALIGNMENTS
                or acta_image_position not in SCREEN_IMAGE_POSITIONS
                or acta_font_family not in ACTA_FONT_FAMILIES
                or not body_font_size.isdigit()
                or not title_font_size.isdigit()
                or not watermark_size.isdigit()
                or not watermark_opacity.isdigit()
            ):
                flash("Opciones de documento no validas.")
                return redirect(url_for("system_config"))
            body_size = min(28, max(7, int(body_font_size)))
            title_size = min(28, max(7, int(title_font_size)))
            watermark_size_value = min(600, max(40, int(watermark_size)))
            watermark_opacity_value = min(100, max(1, int(watermark_opacity)))
            acta_values = {
                "acta_title_line_1": title_line_1,
                "acta_title_line_2": title_line_2,
                "acta_title_alignment": title_alignment,
                "acta_image_position": acta_image_position,
                "acta_image_enabled": "1" if request.form.get("acta_image_enabled") else "0",
                "acta_font_family": acta_font_family,
                "acta_body_font_size": str(body_size),
                "acta_title_font_size": str(title_size),
                "acta_watermark_enabled": "1" if request.form.get("acta_watermark_enabled") else "0",
                "acta_watermark_size": str(watermark_size_value),
                "acta_watermark_opacity": str(watermark_opacity_value),
            }
            if request.form.get("delete_acta_image"):
                acta_values["acta_image_path"] = ""
                acta_values["acta_image_filename"] = ""
            if request.form.get("delete_acta_watermark_image"):
                acta_values["acta_watermark_image_path"] = ""
                acta_values["acta_watermark_image_filename"] = ""
            acta_image = request.files.get("acta_image")
            if acta_image and acta_image.filename:
                extension = Path(acta_image.filename).suffix.lower()
                if extension not in ALLOWED_CONFIG_IMAGE_EXTENSIONS:
                    flash("La imagen del documento debe ser JPG o PNG.")
                    return redirect(url_for("system_config"))
                filename = secure_filename(acta_image.filename)
                stored_path = CONFIG_IMAGE_DIR / f"acta-{secrets.token_hex(8)}-{filename}"
                acta_image.save(stored_path)
                acta_values["acta_image_path"] = str(stored_path)
                acta_values["acta_image_filename"] = Path(acta_image.filename).name
            watermark_image = request.files.get("acta_watermark_image")
            if watermark_image and watermark_image.filename:
                extension = Path(watermark_image.filename).suffix.lower()
                if extension not in ALLOWED_CONFIG_IMAGE_EXTENSIONS:
                    flash("La imagen de sello de agua debe ser JPG o PNG.")
                    return redirect(url_for("system_config"))
                filename = secure_filename(watermark_image.filename)
                stored_path = CONFIG_IMAGE_DIR / f"watermark-{secrets.token_hex(8)}-{filename}"
                watermark_image.save(stored_path)
                acta_values["acta_watermark_image_path"] = str(stored_path)
                acta_values["acta_watermark_image_filename"] = Path(watermark_image.filename).name
            for key, value in acta_values.items():
                db().execute(
                    "INSERT INTO system_settings(key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                    (key, value),
                )
            db().commit()
            audit("document_options_updated", "Opciones de documentos actualizadas")
            flash("Se guardaron los cambios.")
            return redirect(url_for("system_config"))
        if action in {"smtp_settings", "smtp_test"}:
            smtp_port = request.form.get("smtp_port", "587").strip() or "587"
            if not smtp_port.isdigit():
                flash("El puerto SMTP debe ser numerico.")
                return redirect(url_for("system_config"))
            test_email = request.form.get("smtp_test_email", "").strip()
            if action == "smtp_test" and not test_email:
                flash("Debes ingresar un correo para hacer la prueba de envío.")
                return redirect(url_for("system_config"))
            ensure_system_settings_table()
            smtp_values = {
                "smtp_host": request.form.get("smtp_host", "").strip(),
                "smtp_port": smtp_port,
                "smtp_user": request.form.get("smtp_user", "").strip(),
                "smtp_password": request.form.get("smtp_password", ""),
                "smtp_from": request.form.get("smtp_from", "").strip(),
                "smtp_tls": "1" if request.form.get("smtp_tls") else "0",
            }
            for key, value in smtp_values.items():
                db().execute(
                    "INSERT INTO system_settings(key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                    (key, value),
                )
            db().commit()
            audit("smtp_config_updated", f"Configuración SMTP actualizada: {'configurado' if smtp_values['smtp_host'] else 'sin configurar'}")
            if action == "smtp_test":
                ok = send_email(
                    test_email,
                    "Prueba SMTP E-Signum",
                    f"Este es un correo de prueba enviado desde E-Signum.\n\nUsuario: {g.user['name']}\nFecha: {now()}",
                )
                audit("smtp_test_sent", f"Prueba SMTP enviada a {test_email}. Resultado: {'ok' if ok else 'fallo'}")
                if ok:
                    flash("Se guardaron los cambios y el correo de prueba fue enviado.")
                else:
                    flash("Se guardaron los cambios, pero no se pudo enviar el correo de prueba. Revisa la bandeja de correos.")
                return redirect(url_for("system_config"))
            flash("Se guardaron los cambios.")
            return redirect(url_for("system_config"))

        menu_position = request.form.get("menu_position", "")
        button_color = request.form.get("button_color", "")
        button_text_color = request.form.get("button_text_color", "")
        field_border_color = request.form.get("field_border_color", "")
        page_background = request.form.get("page_background", "")
        screen_values: dict[str, str] = {}
        for image_index in ("", "_2"):
            position = request.form.get(f"screen_image{image_index}_position", "")
            size = request.form.get(f"screen_image{image_index}_size", "180").strip()
            if position not in SCREEN_IMAGE_POSITIONS or not size.isdigit():
                flash("Configuración de imagen no valida.")
                return redirect(url_for("system_config"))
            screen_values[f"screen_image{image_index}_position"] = position
            screen_values[f"screen_image{image_index}_size"] = str(min(600, max(40, int(size))))
        if (
            menu_position not in MENU_POSITIONS
            or button_color not in BUTTON_COLORS
            or button_text_color not in BUTTON_TEXT_COLORS
            or field_border_color not in FIELD_BORDER_COLORS
            or page_background not in PAGE_BACKGROUNDS
        ):
            flash("Configuración no valida.")
            return redirect(url_for("system_config"))
        ensure_system_settings_table()
        values = {
            "menu_position": menu_position,
            "button_color": button_color,
            "button_text_color": button_text_color,
            "field_border_color": field_border_color,
            "page_background": page_background,
            **screen_values,
        }
        upload_map = {
            "screen_image": ("screen_image_path", "pantalla1"),
            "screen_image_2": ("screen_image_2_path", "pantalla2"),
        }
        try:
            for field_name, (setting_key, prefix) in upload_map.items():
                if request.form.get(f"delete_{field_name}"):
                    values[setting_key] = ""
                    continue
                stored_path = store_config_image(request.files.get(field_name), prefix)
                if stored_path is not None:
                    values[setting_key] = stored_path
        except ValueError as exc:
            flash(str(exc))
            return redirect(url_for("system_config"))
        for key, value in values.items():
            db().execute(
                "INSERT INTO system_settings(key, value) VALUES (?, ?) ON CONFLICT(key) DO UPDATE SET value = excluded.value",
                (key, value),
            )
        db().commit()
        audit(
            "aesthetic_config_updated",
            f"Configuración estética actualizada. Menú: {MENU_POSITIONS[menu_position]}. Color: {BUTTON_COLORS[button_color][0]}. Fondo: {PAGE_BACKGROUNDS[page_background][0]}.",
        )
        flash("Se guardaron los cambios.")
        return redirect(url_for("system_config"))
    settings = system_settings()
    doc_types = db().execute("SELECT * FROM document_types ORDER BY name").fetchall()
    correlative_options = document_type_options_map()
    screen_images_config = [
        {
            "label": "Imagen de pantalla 1",
            "file_field": "screen_image",
            "position_field": "screen_image_position",
            "size_field": "screen_image_size",
            "position": settings["screen_image_position"],
            "size": settings["screen_image_size"],
            "has_image": bool(settings["screen_image_url"]),
        },
        {
            "label": "Imagen de pantalla 2",
            "file_field": "screen_image_2",
            "position_field": "screen_image_2_position",
            "size_field": "screen_image_2_size",
            "position": settings["screen_image_2_position"],
            "size": settings["screen_image_2_size"],
            "has_image": bool(settings["screen_image_2_url"]),
        },
    ]
    return html_page(
        "Configuración",
        render_template_string(
            """
            <h1>Configuración</h1>
            <div class="grid">
              <button type="button" data-open-config-modal="password_modal">Cambio de clave</button>
              {% if g.user.is_admin or g.user.is_super_admin %}
                <button type="button" data-open-config-modal="aesthetic_modal">Opciones de estética</button>
                <button type="button" data-open-config-modal="smtp_modal">Ubicación de correos SMTP</button>
                <button type="button" data-open-config-modal="document_options_modal">Opciones de documentos</button>
              {% endif %}
            </div>

            <div class="modal-backdrop" id="password_modal">
              <div class="modal">
                <h2>Cambio de clave</h2>
                <form method="post">
                  <input type="hidden" name="action" value="change_password">
                  <label>Clave anterior</label><input name="current_password" type="password" required autocomplete="current-password">
                  <label>Nueva clave</label><input name="new_password" type="password" minlength="7" maxlength="15" required autocomplete="new-password">
                  <label>Confirmar nueva clave</label><input name="confirm_password" type="password" minlength="7" maxlength="15" required autocomplete="new-password">
                  <p><button>Grabar cambio de clave</button> <button type="button" class="danger" data-close-config-modal>Cerrar</button></p>
                </form>
              </div>
            </div>

            {% if g.user.is_admin or g.user.is_super_admin %}
              <div class="modal-backdrop" id="aesthetic_modal">
                <div class="modal modal-scroll" style="width:min(760px,96vw);">
                  <h2>Opciones de estética</h2>
                  <form method="post" enctype="multipart/form-data">
                    <input type="hidden" name="action" value="aesthetic_settings">
                    <div class="modal-scroll-body">
                    <div class="config-fields-grid">
                      <div>
                        <label>Ubicación del menú</label>
                        <select name="menu_position" required>
                          {% for value, label in menu_positions.items() %}
                            <option value="{{ value }}" {% if settings.menu_position == value %}selected{% endif %}>{{ label }}</option>
                          {% endfor %}
                        </select>
                      </div>
                      <div>
                        <label>Color de botones</label>
                        <select name="button_color" required>
                          {% for value, color in button_colors.items() %}
                            <option value="{{ value }}" {% if settings.button_color == value %}selected{% endif %}>{{ color[0] }}</option>
                          {% endfor %}
                        </select>
                      </div>
                      <div>
                        <label>Color del texto de botones</label>
                        <select name="button_text_color" required>
                          {% for value, color in button_text_colors.items() %}
                            <option value="{{ value }}" {% if settings.button_text_color == value %}selected{% endif %}>{{ color[0] }}</option>
                          {% endfor %}
                        </select>
                      </div>
                      <div>
                        <label>Color de bordes de campos</label>
                        <select name="field_border_color" required>
                          {% for value, color in field_border_colors.items() %}
                            <option value="{{ value }}" {% if settings.field_border_color == value %}selected{% endif %}>{{ color[0] }}</option>
                          {% endfor %}
                        </select>
                      </div>
                      <div>
                        <label>Fondo de pantalla</label>
                        <select name="page_background" required>
                          {% for value, color in page_backgrounds.items() %}
                            <option value="{{ value }}" {% if settings.page_background == value %}selected{% endif %}>{{ color[0] }}</option>
                          {% endfor %}
                        </select>
                      </div>
                    </div>
                    <h3>Imágenes de pantalla</h3>
                    <div class="config-fields-grid">
                      {% for image in screen_images_config %}
                        <div>
                          <label>{{ image.label }} JPG o PNG</label>
                          <input name="{{ image.file_field }}" type="file" accept="image/jpeg,image/png,.jpg,.jpeg,.png">
                        </div>
                        <div>
                          <label>Ubicación de {{ image.label|lower }}</label>
                          <select name="{{ image.position_field }}" required>
                            {% for value, label in screen_image_positions.items() %}
                              <option value="{{ value }}" {% if image.position == value %}selected{% endif %}>{{ label }}</option>
                            {% endfor %}
                          </select>
                        </div>
                        <div>
                          <label>Tamaño de {{ image.label|lower }}</label>
                          <input name="{{ image.size_field }}" type="number" min="40" max="600" value="{{ image.size }}" required>
                        </div>
                        {% if image.has_image %}
                          <div>
                            <label><input type="checkbox" name="delete_{{ image.file_field }}"> Eliminar {{ image.label|lower }}</label>
                          </div>
                        {% endif %}
                      {% endfor %}
                    </div>
                    </div>
                    <p class="modal-actions-sticky"><button>Grabar cambio</button> <button type="button" class="danger" data-close-config-modal>Cerrar</button></p>
                  </form>
                </div>
              </div>

              <div class="modal-backdrop" id="smtp_modal">
                <div class="modal">
                  <h2>Ubicación de correos SMTP</h2>
                  <form method="post">
                    <label>Servidor SMTP</label><input name="smtp_host" value="{{ settings.smtp_host }}" placeholder="smtp.ejemplo.cl">
                    <label>Puerto SMTP</label><input name="smtp_port" value="{{ settings.smtp_port }}" inputmode="numeric" pattern="\\d+">
                    <label>Usuario SMTP</label><input name="smtp_user" value="{{ settings.smtp_user }}">
                    <label>Clave SMTP</label><input name="smtp_password" type="password" value="{{ settings.smtp_password }}" autocomplete="new-password">
                    <label>Correo remitente</label><input name="smtp_from" type="email" value="{{ settings.smtp_from }}" placeholder="avisos@ejemplo.cl">
                    <label><input type="checkbox" name="smtp_tls" {% if settings.smtp_tls == "1" %}checked{% endif %}> Usar TLS</label>
                    <label>Correo para prueba</label><input name="smtp_test_email" type="email" placeholder="destino@ejemplo.cl">
                    <p>
                      <button name="action" value="smtp_settings">Grabar cambio</button>
                      <button name="action" value="smtp_test">Probar envío</button>
                      <button type="button" class="danger" data-close-config-modal>Cerrar</button>
                    </p>
                  </form>
                </div>
              </div>

              <div class="modal-backdrop" id="document_options_modal">
                <div class="modal document-options-modal" style="width:min(920px,96vw); max-height:92vh; overflow:auto;">
                  <h2>Opciones de documentos</h2>
                  <form method="post" enctype="multipart/form-data">
                    <input type="hidden" name="action" value="document_options">
                    <div class="config-fields-grid">
                      <h3 class="full-row">Correlativos por tipo</h3>
                      {% for t in doc_types %}
                        <div class="correlative-row">
                          <input class="correlative-input" name="next_correlative_{{ t.id }}" type="number" min="1" max="9999999999" maxlength="10" value="{{ correlative_options.get(t.id, 1) }}" required>
                          <label>{{ t.name }}</label>
                        </div>
                      {% endfor %}

                      <h3 class="full-row">Títulos del documento</h3>
                      <div class="full-row">
                        <label>Título línea 1</label><input class="title-line-input" name="acta_title_line_1" maxlength="70" value="{{ settings.acta_title_line_1 }}">
                      </div>
                      <div class="full-row config-inline-pair">
                        <div><label>Título línea 2</label><input class="title-line-input" name="acta_title_line_2" maxlength="70" value="{{ settings.acta_title_line_2 }}"></div>
                        <div>
                          <label>Ubicación de títulos</label>
                          <select name="acta_title_alignment" required>
                            {% for value, label in acta_title_alignments.items() %}
                              <option value="{{ value }}" {% if settings.acta_title_alignment == value %}selected{% endif %}>{{ label }}</option>
                            {% endfor %}
                          </select>
                        </div>
                      </div>

                      <h3 class="full-row">Imagen del documento</h3>
                      <div class="full-row config-inline-pair">
                        <div>
                          <label>Imagen del documento JPG o PNG</label>
                          <input name="acta_image" id="acta_image_input" type="file" accept="image/jpeg,image/png,.jpg,.jpeg,.png">
                          <p class="muted">Archivo seleccionado: <span id="acta_image_filename">{{ settings.acta_image_filename or "Sin archivo" }}</span></p>
                        </div>
                        <div>
                          <label>Ubicación de imagen</label>
                          <select name="acta_image_position" required>
                            {% for value, label in screen_image_positions.items() %}
                              <option value="{{ value }}" {% if settings.acta_image_position == value %}selected{% endif %}>{{ label }}</option>
                            {% endfor %}
                          </select>
                        </div>
                      </div>
                      {% if settings.acta_image_url %}
                        <div><label><input type="checkbox" name="acta_image_enabled" {% if settings.acta_image_enabled == "1" %}checked{% endif %}> Activar imagen del documento</label></div>
                      {% else %}
                        <div><label><input type="checkbox" name="acta_image_enabled" checked> Activar imagen del documento</label></div>
                      {% endif %}

                      <h3 class="full-row">Imagen como sello de agua</h3>
                      <div>
                        <label>Imagen para documentos JPG o PNG</label>
                        <input name="acta_watermark_image" id="acta_watermark_image_input" type="file" accept="image/jpeg,image/png,.jpg,.jpeg,.png">
                        <p class="muted">Archivo seleccionado: <span id="acta_watermark_image_filename">{{ settings.acta_watermark_image_filename or "Sin archivo" }}</span></p>
                      </div>
                      <div>
                        <label>Tamaño de sello de agua</label>
                        <input class="image-size-input" name="acta_watermark_size" type="number" min="40" max="600" maxlength="7" value="{{ settings.acta_watermark_size }}" required>
                      </div>
                      <div>
                        <label>Transparencia del sello de agua</label>
                        <input class="image-size-input" name="acta_watermark_opacity" type="number" min="1" max="100" maxlength="7" value="{{ settings.acta_watermark_opacity }}" required>
                      </div>
                      {% if settings.acta_watermark_image_url %}
                        <div><label><input type="checkbox" name="acta_watermark_enabled" {% if settings.acta_watermark_enabled == "1" %}checked{% endif %}> Activar sello de agua en documentos</label></div>
                      {% else %}
                        <div><label><input type="checkbox" name="acta_watermark_enabled" checked> Activar sello de agua en documentos</label></div>
                      {% endif %}

                      <h3 class="full-row">Texto del documento</h3>
                      <div>
                        <label>Tipo de letra</label>
                        <select name="acta_font_family" required>
                          {% for value, label in acta_font_families.items() %}
                            <option value="{{ value }}" {% if settings.acta_font_family == value %}selected{% endif %}>{{ label }}</option>
                          {% endfor %}
                        </select>
                      </div>
                      <div><label>Tamaño de títulos</label><input name="acta_title_font_size" type="number" min="7" max="28" value="{{ settings.acta_title_font_size }}" required></div>
                      <div><label>Tamaño del texto</label><input name="acta_body_font_size" type="number" min="7" max="28" value="{{ settings.acta_body_font_size }}" required></div>
                    </div>
                    <p><button>Grabar cambio</button> <button type="button" class="danger" data-close-config-modal>Cerrar</button></p>
                  </form>
                </div>
              </div>
            {% endif %}
            <p><a class="btn" href="{{ url_for('admin_home') }}">Volver</a></p>
            <script>
              document.querySelectorAll("[data-open-config-modal]").forEach((button) => {
                button.addEventListener("click", () => {
                  const modal = document.getElementById(button.dataset.openConfigModal);
                  if (modal) modal.classList.add("active");
                });
              });
              document.querySelectorAll("[data-close-config-modal]").forEach((button) => {
                button.addEventListener("click", () => {
                  button.closest(".modal-backdrop").classList.remove("active");
                });
              });
              document.getElementById("acta_image_input")?.addEventListener("change", (event) => {
                const fileName = event.target.files?.[0]?.name || "{{ settings.acta_image_filename or 'Sin archivo' }}";
                const label = document.getElementById("acta_image_filename");
                if (label) label.textContent = fileName;
              });
              document.getElementById("acta_watermark_image_input")?.addEventListener("change", (event) => {
                const fileName = event.target.files?.[0]?.name || "{{ settings.acta_watermark_image_filename or 'Sin archivo' }}";
                const label = document.getElementById("acta_watermark_image_filename");
                if (label) label.textContent = fileName;
              });
            </script>
            """,
            settings=settings,
            menu_positions=MENU_POSITIONS,
            button_colors=BUTTON_COLORS,
            button_text_colors=BUTTON_TEXT_COLORS,
            field_border_colors=FIELD_BORDER_COLORS,
            page_backgrounds=PAGE_BACKGROUNDS,
            screen_image_positions=SCREEN_IMAGE_POSITIONS,
            screen_images_config=screen_images_config,
            doc_types=doc_types,
            correlative_options=correlative_options,
            acta_title_alignments=ACTA_TITLE_ALIGNMENTS,
            acta_font_families=ACTA_FONT_FAMILIES,
        ),
    )


@app.route("/system-screen-image")
@login_required
def system_screen_image():
    settings = system_settings()
    image_path = settings.get("screen_image_path", "")
    if not image_path:
        abort(404)
    path = Path(image_path)
    try:
        path.resolve().relative_to(CONFIG_IMAGE_DIR.resolve())
    except ValueError:
        abort(404)
    if not path.exists():
        abort(404)
    return send_file(path)


@app.route("/system-screen-image/<int:image_index>")
@login_required
def system_extra_screen_image(image_index: int):
    if image_index not in (2, 3):
        abort(404)
    settings = system_settings()
    image_path = settings.get(f"screen_image_{image_index}_path", "")
    if not image_path:
        abort(404)
    path = Path(image_path)
    try:
        path.resolve().relative_to(CONFIG_IMAGE_DIR.resolve())
    except ValueError:
        abort(404)
    if not path.exists():
        abort(404)
    return send_file(path)


@app.route("/acta-config-image")
@login_required
def acta_config_image():
    settings = system_settings()
    image_path = settings.get("acta_image_path", "")
    if not image_path:
        abort(404)
    path = Path(image_path)
    try:
        path.resolve().relative_to(CONFIG_IMAGE_DIR.resolve())
    except ValueError:
        abort(404)
    if not path.exists():
        abort(404)
    return send_file(path)


@app.route("/acta-watermark-config-image")
@login_required
def acta_watermark_config_image():
    settings = system_settings()
    image_path = settings.get("acta_watermark_image_path", "")
    if not image_path:
        abort(404)
    path = Path(image_path)
    try:
        path.resolve().relative_to(CONFIG_IMAGE_DIR.resolve())
    except ValueError:
        abort(404)
    if not path.exists():
        abort(404)
    return send_file(path)


@app.route("/admin/upload", methods=["GET", "POST"])
@admin_required
def upload_document():
    doc_types = db().execute("SELECT * FROM document_types ORDER BY name").fetchall()
    users = db().execute(
        """
        SELECT u.*, r.name AS role_name
        FROM users u JOIN roles r ON r.id = u.role_id
        WHERE u.active = 1 AND u.can_sign = 1 AND u.is_super_admin = 0
        ORDER BY r.name, u.name
        """
    ).fetchall()
    doc_type_rules = {
        str(doc_type["id"]): [
            row["role_id"]
            for row in db().execute(
                "SELECT role_id FROM document_type_signer_roles WHERE document_type_id = ?",
                (doc_type["id"],),
            ).fetchall()
        ]
        for doc_type in doc_types
    }
    if request.method == "POST":
        uploaded = request.files.get("document_file") or request.files.get("pdf")
        title = request.form["title"].strip()
        title_key = normalize_document_title(title)
        doc_type_id = int(request.form["document_type_id"])
        signer_ids = [int(x) for x in request.form.getlist("signers")]
        doc_type = db().execute("SELECT * FROM document_types WHERE id = ?", (doc_type_id,)).fetchone()
        allowed_role_ids = {
            row["role_id"]
            for row in db().execute(
                "SELECT role_id FROM document_type_signer_roles WHERE document_type_id = ?",
                (doc_type_id,),
            ).fetchall()
        }
        selected_users = db().execute(
            f"SELECT id, name, role_id, active, can_sign, is_super_admin FROM users WHERE id IN ({','.join('?' for _ in signer_ids)})",
            signer_ids,
        ).fetchall() if signer_ids else []
        if not uploaded or Path(uploaded.filename).suffix.lower() not in ALLOWED_EXTENSIONS:
            flash("Debes subir un archivo PDF.")
            return redirect(url_for("upload_document"))
        if not title_key:
            flash("Debes ingresar un título válido para el documento.")
            return redirect(url_for("upload_document"))
        duplicate = db().execute(
            """
            SELECT id, title, status
            FROM documents
            WHERE document_type_id = ?
              AND title_key = ?
              AND status != 'annulled'
            ORDER BY created_at DESC
            LIMIT 1
            """,
            (doc_type_id, title_key),
        ).fetchone()
        if duplicate:
            flash(f"Ya existe un documento de este tipo con el mismo título: {duplicate['title']}. No se subió el archivo.")
            return redirect(url_for("upload_document"))
        if not signer_ids:
            flash("Debes seleccionar al menos un firmante.")
            return redirect(url_for("upload_document"))
        if doc_type and len(signer_ids) > doc_type["required_signatures"]:
            flash(f"Este tipo de documento permite máximo {doc_type['required_signatures']} firmantes.")
            return redirect(url_for("upload_document"))
        if len(selected_users) != len(set(signer_ids)):
            flash("Uno o mas firmantes seleccionados no existen.")
            return redirect(url_for("upload_document"))
        if allowed_role_ids and any(user["role_id"] not in allowed_role_ids for user in selected_users):
            flash("Uno o mas firmantes seleccionados no tienen un cargo autorizado para este tipo de documento.")
            return redirect(url_for("upload_document"))
        if any(not user["active"] or not user["can_sign"] or user["is_super_admin"] for user in selected_users):
            flash("Uno o mas firmantes seleccionados estan desactivados o no tienen permiso para firmar.")
            return redirect(url_for("upload_document"))
        filename = secure_filename(uploaded.filename)
        stored_path = UPLOAD_DIR / f"{secrets.token_hex(8)}-{filename}"
        uploaded.save(stored_path)
        cur = db().execute(
            """
            INSERT INTO documents(title, title_key, document_type_id, original_filename, stored_path, uploaded_by_user_id, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (title, title_key, doc_type_id, filename, str(stored_path), g.user["id"], now()),
        )
        document_id = cur.lastrowid
        for signer_id in signer_ids:
            db().execute("INSERT INTO document_required_signers(document_id, user_id) VALUES (?, ?)", (document_id, signer_id))
        db().commit()
        audit("document_uploaded", f"Documento subido: {title}", document_id=document_id)
        sent, failed = notify_document_signers(document_id, signer_ids)
        audit("document_signers_notified", f"Correos enviados: {sent}. Pendientes/fallidos: {failed}", document_id=document_id)
        if failed:
            flash(f"Documento creado. Correos enviados: {sent}. Correos no enviados o sin SMTP: {failed}. Revisa la bandeja de salida.")
        else:
            flash(f"Documento creado y correos enviados a {sent} firmantes.")
        return redirect(url_for("document_detail", document_id=document_id))
    return html_page(
        "Subir archivo",
        render_template_string(
            """
            <h1>Subir archivo</h1>
            <form class="panel" method="post" enctype="multipart/form-data">
              <label>Título</label><input name="title" required>
              <label>Tipo de documento</label>
              <select name="document_type_id" id="document_type_id" required>
                {% for t in doc_types %}<option value="{{ t.id }}">{{ t.name }} - máximo {{ t.required_signatures }} firmantes</option>{% endfor %}
              </select>
              <label>Archivo PDF</label><input name="document_file" type="file" accept="application/pdf,.pdf" required>
              <h2>Firmantes</h2>
              <p id="signer_count" class="muted"></p>
              {% for u in users %}
                <label class="signer-option" data-role-id="{{ u.role_id }}">
                  <input type="checkbox" name="signers" value="{{ u.id }}"> {{ u.name }} - {{ u.role_name }}
                </label>
              {% endfor %}
              <p><button>Crear documento</button></p>
            </form>
            <script>
              const signerRules = {{ doc_type_rules|tojson }};
              const typeSelect = document.getElementById("document_type_id");
              const signerCount = document.getElementById("signer_count");
              const signerOptions = Array.from(document.querySelectorAll(".signer-option"));

              function filterSigners() {
                const allowedRoles = new Set((signerRules[typeSelect.value] || []).map(String));
                let visibleCount = 0;
                signerOptions.forEach((option) => {
                  const allowed = allowedRoles.size === 0 || allowedRoles.has(option.dataset.roleId);
                  option.style.display = allowed ? "block" : "none";
                  if (!allowed) {
                    option.querySelector("input").checked = false;
                  } else {
                    visibleCount += 1;
                  }
                });
                signerCount.textContent = visibleCount + " usuarios disponibles para firmar este tipo de documento.";
              }

              typeSelect.addEventListener("change", filterSigners);
              filterSigners();
            </script>
            """,
            doc_types=doc_types,
            users=users,
            doc_type_rules=doc_type_rules,
        ),
    )


@app.route("/admin/users", methods=["GET", "POST"])
@admin_required
def admin_users():
    roles = db().execute("SELECT * FROM roles ORDER BY name").fetchall()
    if request.method == "POST":
        registration_number = request.form["registration_number"].strip()
        fire_department = request.form.get("fire_department", "").strip().upper()
        company = request.form.get("company", "").strip()
        name = request.form["name"].strip().upper()
        rut = normalize_rut(request.form["rut"])
        email = request.form["email"].strip().lower()
        password = request.form["password"].strip()
        signature_order = request.form.get("signature_order", "99").strip()
        role_id = request.form.get("role_id", "").strip()
        session["admin_user_form_data"] = {
            "name": name,
            "rut": rut,
            "registration_number": registration_number,
            "fire_department": fire_department,
            "company": company,
            "email": email,
            "signature_order": signature_order,
            "role_id": role_id,
            "can_sign": bool(request.form.get("can_sign")),
            "can_view": bool(request.form.get("can_view")),
            "can_print": bool(request.form.get("can_print")),
            "can_download": bool(request.form.get("can_download")),
            "can_view_user_history": bool(request.form.get("can_view_user_history")),
            "is_admin": bool(request.form.get("is_admin")),
            "active": bool(request.form.get("active", "1")),
            "password_error": False,
        }
        if not registration_number or not name or not rut or not email or not password:
            flash("Número de registro, nombre, RUT, correo y clave son obligatorios.")
            return redirect(url_for("admin_users"))
        if len(name) > 50:
            flash("El nombre debe tener como máximo 50 caracteres.")
            return redirect(url_for("admin_users"))
        if len(registration_number) > 10:
            flash("El número de registro debe tener como máximo 10 caracteres.")
            return redirect(url_for("admin_users"))
        if len(company) > 7:
            flash("La compañía debe tener como máximo 7 caracteres.")
            return redirect(url_for("admin_users"))
        if len(email) > 30:
            flash("El correo recuperación debe tener como máximo 30 caracteres.")
            return redirect(url_for("admin_users"))
        if not role_id.isdigit():
            flash("Debes seleccionar un cargo.")
            return redirect(url_for("admin_users"))
        if len(signature_order) > 2 or not signature_order.isdigit() or not 0 <= int(signature_order) <= 99:
            flash("El orden de firma debe ser numerico entre 0 y 99.")
            return redirect(url_for("admin_users"))
        signature_order_value = int(signature_order)
        if not valid_rut(rut):
            flash("El RUT no es válido. Revisa el número y su dígito verificador.")
            return redirect(url_for("admin_users"))
        if not valid_password(password):
            session["admin_user_form_data"]["password_error"] = True
            flash(password_help())
            return redirect(url_for("admin_users"))
        existing_user = find_user_by_rut(rut)
        if existing_user:
            flash("usuario ya existe")
            return redirect(url_for("admin_users"))
        try:
            cursor = db().execute(
                """
                INSERT INTO users(registration_number, fire_department, company, name, rut, recovery_email, internal_code, password_hash, role_id,
                                  signature_order, can_sign, can_view, can_print, can_download, can_view_user_history,
                                  is_admin, active, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    registration_number,
                    fire_department,
                    company,
                    name,
                    rut,
                    email,
                    random_internal_code(),
                    generate_password_hash(password),
                    int(role_id),
                    signature_order_value,
                    1 if request.form.get("can_sign") else 0,
                    1 if request.form.get("can_view") else 0,
                    1 if request.form.get("can_print") else 0,
                    1 if request.form.get("can_download") else 0,
                    1 if request.form.get("can_view_user_history") else 0,
                    1 if request.form.get("is_admin") and can_manage_admins() else 0,
                    1 if request.form.get("active") else 0,
                    now(),
                ),
            )
            new_user_id = cursor.lastrowid
            if request.form.get("is_admin") and can_manage_admins():
                db().execute(
                    "INSERT INTO admin_history(changed_by_user_id, target_user_id, action, created_at) VALUES (?, ?, 'grant', ?)",
                    (g.user["id"], new_user_id, now()),
                )
            db().commit()
        except sqlite3.IntegrityError:
            db().rollback()
            flash("usuario ya existe")
            return redirect(url_for("admin_users"))
        audit("user_created", f"Usuario creado: {name}")
        session.pop("admin_user_form_data", None)
        flash("Usuario creado correctamente.")
        return redirect(url_for("admin_users"))
    status = request.args.get("status", "")
    if status not in {"", "active", "inactive"}:
        status = ""
    search = request.args.get("search", "").strip()
    role_filter = request.args.get("role_id", "").strip()
    if role_filter and not any(str(role["id"]) == role_filter for role in roles):
        role_filter = ""
    users = []
    search_results = []
    if search or role_filter:
        search_query = """
            SELECT u.*, r.name AS role_name
            FROM users u JOIN roles r ON r.id = u.role_id
            WHERE u.is_super_admin = 0
        """
        search_params: list[object] = []
        if search:
            search_query += " AND u.name LIKE ?"
            search_params.append(f"%{search.upper()}%")
        if role_filter:
            search_query += " AND u.role_id = ?"
            search_params.append(int(role_filter))
        search_query += " ORDER BY u.active DESC, u.name"
        search_results = db().execute(search_query, search_params).fetchall()
    if status:
        query = """
            SELECT u.*, r.name AS role_name
            FROM users u JOIN roles r ON r.id = u.role_id
            WHERE u.active = ? AND u.is_super_admin = 0
        """
        params: list[object] = [1 if status == "active" else 0]
        if search:
            query += " AND u.name LIKE ?"
            params.append(f"%{search.upper()}%")
        if role_filter:
            query += " AND u.role_id = ?"
            params.append(int(role_filter))
        query += " ORDER BY u.name"
        users = db().execute(query, params).fetchall()
    saved_form_data = session.pop("admin_user_form_data", None)
    show_create_form = bool(saved_form_data)
    form_data = saved_form_data or {
            "name": "",
            "rut": "",
            "registration_number": "",
            "fire_department": "",
            "company": "",
            "email": "",
            "signature_order": "0",
            "role_id": str(roles[0]["id"]) if roles else "",
            "can_sign": False,
            "can_view": False,
            "can_print": False,
            "can_download": False,
            "can_view_user_history": False,
            "is_admin": False,
            "active": False,
            "password_error": False,
        }
    return html_page(
        "Usuarios",
        render_template_string(
            """
            <h1>Usuarios</h1>
            <p class="inline-actions">
              <button type="button" id="show_create_user">Nuevo usuario</button>
              <button type="button" id="show_import_user">Importar usuario</button>
              <a class="btn" href="{{ url_for('home') }}">Salir</a>
            </p>
            <div class="modal-backdrop {% if show_create_form %}active{% endif %}" id="create_user_modal">
              <div class="modal modal-scroll user-modal">
                <form method="post" autocomplete="off" id="create_user_form">
                  <div class="modal-scroll-body">
                    <h2>Nuevo usuario</h2>
                    <div class="user-form-grid">
                      <div class="user-fields-grid">
                        <div class="full-row"><label><input type="checkbox" name="active" {% if form_data.active %}checked{% endif %}> Cuenta activa</label></div>
                        <div class="full-row"><label>Nombre</label><input class="user-name-input uppercase-input" name="name" value="{{ form_data.name }}" maxlength="50" autocomplete="off" required></div>
                        <div><label>RUT</label><input name="rut" value="{{ form_data.rut }}" data-rut-input placeholder="9.999.999-5 o 99.999.999-9" pattern="\\d{1,2}\\.\\d{3}\\.\\d{3}-[\\dKk]" title="Formato: 9.999.999-5 o 99.999.999-9 con dígito verificador válido" autocomplete="off" required></div>
                        <div><label>Número de registro</label><input class="field-registry" name="registration_number" value="{{ form_data.registration_number }}" maxlength="10" autocomplete="off" required></div>
                        <div><label>Cuerpo de Bomberos</label><input class="uppercase-input" name="fire_department" value="{{ form_data.fire_department }}" autocomplete="off"></div>
                        <div><label>Compañía</label><input class="field-company" name="company" value="{{ form_data.company }}" maxlength="7" autocomplete="off"></div>
                        <div><label>Orden de firma</label><input class="field-signature-order" name="signature_order" inputmode="numeric" pattern="\\d{1,2}" maxlength="2" value="{{ form_data.signature_order }}" autocomplete="off" required></div>
                        <div><label>Correo recuperación</label><input class="field-email lowercase-input" name="email" type="email" value="{{ form_data.email }}" maxlength="30" autocomplete="off" required></div>
                        <div><label>Clave</label><input name="password" type="password" minlength="7" maxlength="15" title="Debe tener entre 7 y 15 caracteres. Puede incluir letras, números y símbolos." autocomplete="new-password" required></div>
                        <div><label>Cargo</label>
                        <select class="field-role" name="role_id" required>{% for r in roles %}<option value="{{ r.id }}" {% if form_data.role_id|int == r.id %}selected{% endif %}>{{ r.name }}</option>{% endfor %}</select></div>
                      </div>
                      <div class="user-permissions">
                        <h2>Permisos</h2>
                        <label><input type="checkbox" name="can_sign" data-sign-toggle {% if form_data.can_sign %}checked{% endif %}> Puede firmar</label>
                        <label><input type="checkbox" name="can_view" {% if form_data.can_view %}checked{% endif %}> Puede ver documentos cerrados permitidos</label>
                        <label><input type="checkbox" name="can_print" {% if form_data.can_print %}checked{% endif %}> Puede imprimir</label>
                        <label><input type="checkbox" name="can_download" {% if form_data.can_download %}checked{% endif %}> Puede descargar</label>
                        <label><input type="checkbox" name="can_view_user_history" {% if form_data.can_view_user_history %}checked{% endif %}> Puede ver informe historial</label>
                        {% if can_manage_admins %}
                          <label><input type="checkbox" name="is_admin" {% if form_data.is_admin %}checked{% endif %}> Administrador</label>
                        {% endif %}
                      </div>
                    </div>
                    {% if form_data.password_error %}
                      <div class="flash" style="border-color:#a93636;background:#ffe4e4;color:#7a1f1f;">
                        <strong>Clave no cumple</strong><br>
                        Debe tener 7 caracteres mínimo y 15 máximo; puede tener símbolos.<br>
                        Ejemplo: <strong>Cecom1115!</strong>
                      </div>
                    {% endif %}
                  </div>
                  <p class="modal-actions-sticky"><button>Guardar usuario</button> <button type="button" id="hide_create_user">Cancelar</button></p>
                </form>
              </div>
            </div>
            <div class="modal-backdrop" id="import_user_modal">
              <div class="modal">
                <form method="post" action="{{ url_for('import_users_excel') }}" enctype="multipart/form-data" id="import_user_form">
                  <h2>Importar usuarios desde Excel</h2>
                  <p class="muted">El archivo debe ser .xlsx y usar las columnas de la plantilla. Si un RUT ya existe, esa fila no se crea.</p>
                  <p><a class="btn" href="{{ url_for('download_users_template') }}">Descargar plantilla Excel</a></p>
                  <label>Archivo Excel</label><input name="excel_file" type="file" accept=".xlsx" required>
                  <p><button>Importar usuarios</button> <button type="button" id="hide_import_user">Cancelar</button></p>
                </form>
              </div>
            </div>
            <form class="panel" method="get">
              <div class="inline-actions">
                <div>
                  <label>Buscar por nombre</label><input name="search" value="{{ search }}" placeholder="Nombre del usuario">
                </div>
                <div>
                  <label>Buscar por cargo</label>
                  <select name="role_id">
                    <option value="">Todos</option>
                    {% for r in roles %}
                      <option value="{{ r.id }}" {% if role_filter|int == r.id %}selected{% endif %}>{{ r.name }}</option>
                    {% endfor %}
                  </select>
                </div>
              </div>
              <input type="hidden" name="status" value="{{ status }}">
              <p><button>Buscar</button> <a class="btn" href="{{ url_for('admin_users', status=status) }}">Limpiar</a></p>
              {% if search or role_filter %}
                {% if search_results %}
                  <table>
                    <tr><th>Orden firma</th><th>Nombre</th><th>Cargo</th><th>Estado</th><th>Acciones</th></tr>
                    {% for u in search_results %}
                      <tr>
                        <td>{{ "%02d"|format(u.signature_order or 99) }}</td>
                        <td>{{ u.name }}</td>
                        <td>{{ u.role_name }}</td>
                        <td>{{ "Activo" if u.active else "Inactivo" }}</td>
                        <td><a class="btn" href="{{ url_for('edit_user', user_id=u.id) }}">Editar</a></td>
                      </tr>
                    {% endfor %}
                  </table>
                {% else %}
                  <p class="flash">Usuario no existe.</p>
                {% endif %}
              {% endif %}
            </form>
            <div class="panel">
              <a class="btn" href="{{ url_for('admin_users', status='active', search=search, role_id=role_filter) }}">Ver activos</a>
              <a class="btn" href="{{ url_for('admin_users', status='inactive', search=search, role_id=role_filter) }}">Ver desactivados</a>
              {% if status %}
                <p class="muted">Mostrando usuarios {{ "activos" if status == "active" else "desactivados" }}.</p>
              {% else %}
                <p class="muted">Elige Ver activos o Ver desactivados para mostrar el listado.</p>
              {% endif %}
            </div>
            {% if status %}
            <table>
              <tr><th>Orden firma</th><th>Nombre</th><th>Cargo</th><th>Permisos</th><th>Admin</th><th>Acciones</th></tr>
              {% for u in users %}
                <tr>
                  <td>{{ "%02d"|format(u.signature_order or 99) }}</td><td>{{ u.name }}</td><td>{{ u.role_name }}</td>
                  <td>Firmar: {{ u.can_sign }} | Ver: {{ u.can_view }} | Imprimir: {{ u.can_print }} | Descargar: {{ u.can_download }} | Historial: {{ u.can_view_user_history }}</td>
                  <td>{{ "Super" if u.is_super_admin else ("Admin" if u.is_admin else "No") }} | {{ "Activo" if u.active else "Inactivo" }}</td>
                  <td><a class="btn" href="{{ url_for('edit_user', user_id=u.id) }}">Editar</a></td>
                </tr>
              {% endfor %}
            </table>
            {% endif %}
            <script>
              const createUserModal = document.getElementById("create_user_modal");
              document.getElementById("show_create_user")?.addEventListener("click", () => {
                createUserModal?.classList.add("active");
              });
              document.getElementById("hide_create_user")?.addEventListener("click", () => {
                createUserModal?.classList.remove("active");
              });
              const importUserModal = document.getElementById("import_user_modal");
              document.getElementById("show_import_user")?.addEventListener("click", () => {
                importUserModal?.classList.add("active");
              });
              document.getElementById("hide_import_user")?.addEventListener("click", () => {
                importUserModal?.classList.remove("active");
              });
            </script>
            """,
            users=users,
            roles=roles,
            status=status,
            search=search,
            role_filter=role_filter,
            search_results=search_results,
            form_data=form_data,
            show_create_form=show_create_form,
            can_manage_admins=can_manage_admins(),
        ),
    )


@app.route("/admin/users/template.xlsx")
@admin_required
def download_users_template():
    from openpyxl import Workbook

    headers = [
        "nombre",
        "rut",
        "numero_registro",
        "cuerpo_de_bomberos",
        "compania",
        "correo_recuperacion",
        "clave",
        "cargo",
        "orden_firma",
        "puede_firmar",
        "puede_ver",
        "puede_imprimir",
        "puede_descargar",
        "puede_ver_historial",
        "administrador",
        "activo",
    ]
    wb = Workbook()
    ws = wb.active
    ws.title = "Usuarios"
    ws.append(headers)
    ws.append([
        "JUAN PEREZ",
        "12.345.678-5",
        "1001",
        "CUERPO DE BOMBEROS DE EJEMPLO",
        "Primera Compañía",
        "juan.perez@correo.cl",
        "Clave123",
        "Bombero",
        "1",
        "SI",
        "SI",
        "NO",
        "NO",
        "NO",
        "NO",
        "SI",
    ])
    for column in ws.columns:
        width = max(len(str(cell.value or "")) for cell in column) + 2
        ws.column_dimensions[column[0].column_letter].width = min(max(width, 14), 28)
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="plantilla-usuarios.xlsx",
    )


@app.post("/admin/users/import")
@admin_required
def import_users_excel():
    from openpyxl import load_workbook

    uploaded = request.files.get("excel_file")
    if not uploaded or not uploaded.filename:
        flash("Debes seleccionar un archivo Excel.")
        return redirect(url_for("admin_users"))
    if Path(uploaded.filename).suffix.lower() != ".xlsx":
        flash("El archivo debe estar en formato .xlsx.")
        return redirect(url_for("admin_users"))

    try:
        workbook = load_workbook(uploaded.stream, data_only=True)
        sheet = workbook.active
    except Exception as exc:
        flash(f"No se pudo leer el Excel: {exc}")
        return redirect(url_for("admin_users"))

    raw_headers = [strip_accents(excel_text(cell.value).lower()).replace(" ", "_") for cell in sheet[1]]
    headers = {header: index for index, header in enumerate(raw_headers) if header}
    required_headers = ["nombre", "rut", "numero_registro", "correo_recuperacion", "clave", "cargo"]
    missing_headers = [header for header in required_headers if header not in headers]
    if missing_headers:
        flash("Faltan columnas obligatorias: " + ", ".join(missing_headers) + ".")
        return redirect(url_for("admin_users"))

    roles = db().execute("SELECT id, name FROM roles").fetchall()
    role_map = {strip_accents(row["name"].lower()): row["id"] for row in roles}
    created = 0
    skipped: list[str] = []

    def value(row: tuple, key: str) -> str:
        index = headers.get(key)
        if index is None or index >= len(row):
            return ""
        return excel_text(row[index])

    for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
        if not any(excel_text(cell) for cell in row):
            continue
        name = value(row, "nombre").upper()
        rut = normalize_rut(value(row, "rut"))
        registration_number = value(row, "numero_registro")
        fire_department = value(row, "cuerpo_de_bomberos").upper()
        company = value(row, "compania")
        email = value(row, "correo_recuperacion").lower()
        password = value(row, "clave")
        role_name = value(row, "cargo")
        role_id = role_map.get(strip_accents(role_name.lower()))
        signature_order = value(row, "orden_firma") or "99"

        row_errors = []
        if not name:
            row_errors.append("nombre")
        if not rut:
            row_errors.append("rut")
        if not registration_number:
            row_errors.append("numero_registro")
        if not email:
            row_errors.append("correo_recuperacion")
        if not password:
            row_errors.append("clave")
        if not role_name:
            row_errors.append("cargo")
        if row_errors:
            skipped.append(f"Fila {row_number}: faltan {', '.join(row_errors)}")
            continue
        if len(name) > 50:
            skipped.append(f"Fila {row_number}: nombre máximo 50 caracteres")
            continue
        if len(registration_number) > 10:
            skipped.append(f"Fila {row_number}: numero_registro máximo 10 caracteres")
            continue
        if len(company) > 7:
            skipped.append(f"Fila {row_number}: compania máximo 7 caracteres")
            continue
        if len(email) > 30:
            skipped.append(f"Fila {row_number}: correo_recuperacion máximo 30 caracteres")
            continue
        if len(role_name) > 50:
            skipped.append(f"Fila {row_number}: cargo máximo 50 caracteres")
            continue
        if not valid_rut(rut):
            skipped.append(f"Fila {row_number}: RUT inválido")
            continue
        if not valid_password(password):
            skipped.append(f"Fila {row_number}: clave no cumple")
            continue
        if not role_id:
            skipped.append(f"Fila {row_number}: cargo no existe")
            continue
        if len(signature_order) > 2 or not signature_order.isdigit() or not 0 <= int(signature_order) <= 99:
            skipped.append(f"Fila {row_number}: orden_firma debe ser 0 a 99")
            continue
        can_sign_value = excel_bool(value(row, "puede_firmar"), True)
        signature_order_value = int(signature_order)
        if find_user_by_rut(rut):
            skipped.append(f"Fila {row_number}: usuario ya existe")
            continue

        try:
            cursor = db().execute(
                """
                INSERT INTO users(registration_number, fire_department, company, name, rut, recovery_email, internal_code, password_hash, role_id,
                                  signature_order, can_sign, can_view, can_print, can_download, can_view_user_history,
                                  is_admin, active, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                """,
                (
                    registration_number,
                    fire_department,
                    company,
                    name,
                    rut,
                    email,
                    random_internal_code(),
                    generate_password_hash(password),
                    int(role_id),
                    signature_order_value,
                    1 if can_sign_value else 0,
                    1 if excel_bool(value(row, "puede_ver"), True) else 0,
                    1 if excel_bool(value(row, "puede_imprimir"), False) else 0,
                    1 if excel_bool(value(row, "puede_descargar"), False) else 0,
                    1 if excel_bool(value(row, "puede_ver_historial"), False) else 0,
                    1 if excel_bool(value(row, "administrador"), False) and can_manage_admins() else 0,
                    1 if excel_bool(value(row, "activo"), True) else 0,
                    now(),
                ),
            )
            if excel_bool(value(row, "administrador"), False) and can_manage_admins():
                db().execute(
                    "INSERT INTO admin_history(changed_by_user_id, target_user_id, action, created_at) VALUES (?, ?, 'grant', ?)",
                    (g.user["id"], cursor.lastrowid, now()),
                )
            created += 1
        except sqlite3.IntegrityError:
            skipped.append(f"Fila {row_number}: usuario ya existe")

    audit("users_imported_excel", f"Importación Excel usuarios. Creados: {created}. Omitidos: {len(skipped)}")
    db().commit()
    message = f"Importación terminada. Usuarios creados: {created}. Filas omitidas: {len(skipped)}."
    if skipped:
        message += " " + " | ".join(skipped[:10])
        if len(skipped) > 10:
            message += f" | y {len(skipped) - 10} errores mas."
    flash(message)
    return redirect(url_for("admin_users"))


@app.route("/admin/users/<int:user_id>", methods=["GET", "POST"])
@admin_required
def edit_user(user_id: int):
    user = db().execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        abort(404)
    if user["is_super_admin"]:
        flash("El super administrador es una cuenta técnica y no se administra como usuario normal.")
        return redirect(url_for("admin_users"))
    roles = db().execute("SELECT * FROM roles ORDER BY name").fetchall()
    if request.method == "POST":
        old_admin = user["is_admin"]
        registration_number = request.form["registration_number"].strip()
        fire_department = request.form.get("fire_department", "").strip().upper()
        company = request.form.get("company", "").strip()
        name = request.form["name"].strip().upper()
        email = request.form["email"].strip().lower()
        password = request.form.get("password", "").strip()
        rut = normalize_rut(request.form["rut"])
        signature_order = request.form.get("signature_order", "99").strip()
        if not registration_number or not name or not rut or not email:
            flash("Número de registro, nombre, RUT y correo son obligatorios.")
            return redirect(url_for("edit_user", user_id=user_id))
        if len(name) > 50:
            flash("El nombre debe tener como máximo 50 caracteres.")
            return redirect(url_for("edit_user", user_id=user_id))
        if len(registration_number) > 10:
            flash("El número de registro debe tener como máximo 10 caracteres.")
            return redirect(url_for("edit_user", user_id=user_id))
        if len(company) > 7:
            flash("La compañía debe tener como máximo 7 caracteres.")
            return redirect(url_for("edit_user", user_id=user_id))
        if len(email) > 30:
            flash("El correo recuperación debe tener como máximo 30 caracteres.")
            return redirect(url_for("edit_user", user_id=user_id))
        if len(signature_order) > 2 or not signature_order.isdigit() or not 0 <= int(signature_order) <= 99:
            flash("El orden de firma debe ser numerico entre 0 y 99.")
            return redirect(url_for("edit_user", user_id=user_id))
        if password and not valid_password(password):
            flash(password_help())
            return redirect(url_for("edit_user", user_id=user_id))
        signature_order_value = int(signature_order)
        if not valid_rut(rut):
            flash("El RUT no es válido. Revisa el número y su dígito verificador.")
            return redirect(url_for("edit_user", user_id=user_id))
        existing_user = find_user_by_rut(rut)
        if existing_user and existing_user["id"] != user_id:
            flash("usuario ya existe")
            return redirect(url_for("edit_user", user_id=user_id))
        is_admin = 1 if request.form.get("is_admin") and can_manage_admins() else old_admin
        if can_manage_admins() and not request.form.get("is_admin"):
            is_admin = 0
        active = 1 if request.form.get("active") else 0
        if user["is_super_admin"]:
            is_admin = 1
            active = 1
        db().execute(
            """
            UPDATE users
            SET registration_number=?, fire_department=?, company=?, name=?, rut=?, recovery_email=?, signature_order=?, role_id=?, can_sign=?, can_view=?, can_print=?,
                can_download=?, can_view_user_history=?, is_admin=?, active=?
            WHERE id=?
            """,
            (
                registration_number,
                fire_department,
                company,
                name,
                rut,
                email,
                signature_order_value,
                int(request.form["role_id"]),
                1 if request.form.get("can_sign") else 0,
                1 if request.form.get("can_view") else 0,
                1 if request.form.get("can_print") else 0,
                1 if request.form.get("can_download") else 0,
                1 if request.form.get("can_view_user_history") else 0,
                is_admin,
                active,
                user_id,
            ),
        )
        if password:
            db().execute("UPDATE users SET password_hash = ? WHERE id = ?", (generate_password_hash(password), user_id))
            audit("password_changed_by_admin", "Clave cambiada por administrador desde edición de usuario", user_id)
        if old_admin != is_admin:
            action = "grant" if is_admin else "revoke"
            db().execute(
                "INSERT INTO admin_history(changed_by_user_id, target_user_id, action, created_at) VALUES (?, ?, ?, ?)",
                (g.user["id"], user_id, action, now()),
            )
            audit(f"admin_{action}", f"Permiso administrador: {action}", user_id)
        db().commit()
        audit("user_updated", "Usuario actualizado", user_id)
        flash("Usuario actualizado.")
        return redirect(url_for("admin_users"))
    return html_page(
        "Editar usuario",
        render_template_string(
            """
            <h1>Editar usuario</h1>
            <div class="modal-backdrop active">
              <div class="modal modal-scroll user-modal">
                <form method="post" autocomplete="off" id="edit_user_form">
                  <div class="modal-scroll-body">
                    <h2>Editar usuario</h2>
                    <div class="user-form-grid">
                      <div class="user-fields-grid">
                        {% if not user.is_super_admin %}
                          <div class="full-row"><label><input type="checkbox" name="active" {% if user.active %}checked{% endif %}> Cuenta activa</label></div>
                        {% endif %}
                        <div class="full-row"><label>Nombre</label><input class="user-name-input uppercase-input" name="name" value="{{ user.name }}" maxlength="50" autocomplete="off" required></div>
                        <div><label>RUT</label><input name="rut" value="{{ user.rut }}" data-rut-input pattern="\\d{1,2}\\.\\d{3}\\.\\d{3}-[\\dKk]" title="Formato: 9.999.999-5 o 99.999.999-9 con dígito verificador válido" autocomplete="off" required></div>
                        <div><label>Número de registro</label><input class="field-registry" name="registration_number" value="{{ user.registration_number }}" maxlength="10" autocomplete="off" required></div>
                        <div><label>Cuerpo de Bomberos</label><input class="uppercase-input" name="fire_department" value="{{ user.fire_department }}" autocomplete="off"></div>
                        <div><label>Compañía</label><input class="field-company" name="company" value="{{ user.company }}" maxlength="7" autocomplete="off"></div>
                        <div><label>Orden de firma</label><input class="field-signature-order" name="signature_order" inputmode="numeric" pattern="\\d{1,2}" maxlength="2" value="{{ user.signature_order if user.signature_order is not none else 99 }}" autocomplete="off" required></div>
                        <div><label>Correo recuperación</label><input class="field-email lowercase-input" name="email" type="email" value="{{ user.recovery_email }}" maxlength="30" autocomplete="off" required></div>
                        <div><label>Clave</label><input name="password" type="password" minlength="7" maxlength="15" title="Dejar vacío para mantener la clave actual. Si la cambias, debe tener entre 7 y 15 caracteres." autocomplete="new-password"></div>
                        <div><label>Cargo</label>
                        <select class="field-role" name="role_id" required>{% for r in roles %}<option value="{{ r.id }}" {% if r.id == user.role_id %}selected{% endif %}>{{ r.name }}</option>{% endfor %}</select></div>
                      </div>
                      <div class="user-permissions">
                        <h2>Permisos</h2>
                        {% if can_manage_admins and not user.is_super_admin %}
                          <label><input type="checkbox" name="is_admin" {% if user.is_admin %}checked{% endif %}> Administrador</label>
                        {% endif %}
                        <label><input type="checkbox" name="can_sign" data-sign-toggle {% if user.can_sign %}checked{% endif %}> Puede firmar</label>
                        <label><input type="checkbox" name="can_view" {% if user.can_view %}checked{% endif %}> Puede ver documentos cerrados permitidos</label>
                        <label><input type="checkbox" name="can_print" {% if user.can_print %}checked{% endif %}> Puede imprimir</label>
                        <label><input type="checkbox" name="can_download" {% if user.can_download %}checked{% endif %}> Puede descargar</label>
                        <label><input type="checkbox" name="can_view_user_history" {% if user.can_view_user_history %}checked{% endif %}> Puede ver informe historial</label>
                      </div>
                    </div>
                  </div>
                  <p class="modal-actions-sticky"><button>Guardar</button> <a class="btn danger" href="{{ url_for('admin_users') }}" id="exit_edit_user">Salir</a></p>
                </form>
              </div>
            </div>
            <script>
              const editUserForm = document.getElementById("edit_user_form");
              const exitEditUser = document.getElementById("exit_edit_user");
              const initialEditUserData = editUserForm ? new FormData(editUserForm) : null;
              function formChanged(form, initialData) {
                if (!form || !initialData) return false;
                const current = new FormData(form);
                const keys = new Set([...initialData.keys(), ...current.keys()]);
                for (const key of keys) {
                  if (initialData.getAll(key).join("|") !== current.getAll(key).join("|")) return true;
                }
                return false;
              }
              exitEditUser?.addEventListener("click", (event) => {
                if (formChanged(editUserForm, initialEditUserData) && !confirm("No ha guardado los cambios ¿está seguro que quiere salir?")) {
                  event.preventDefault();
                }
              });
            </script>
            """,
            user=user,
            roles=roles,
            can_manage_admins=can_manage_admins(),
        ),
    )


@app.post("/admin/users/<int:user_id>/delete")
@admin_required
def delete_user(user_id: int):
    user = db().execute("SELECT * FROM users WHERE id = ?", (user_id,)).fetchone()
    if not user:
        abort(404)
    if user["is_super_admin"] or user["id"] == g.user["id"]:
        flash("No se puede eliminar este usuario.")
        return redirect(url_for("admin_users"))
    db().execute(
        "UPDATE users SET active = 0, can_sign = 0, is_admin = 0 WHERE id = ?",
        (user_id,),
    )
    db().commit()
    audit("user_deleted", f"Usuario desactivado: {user['name']} ({user['rut']})", user_id)
    flash("Usuario eliminado del uso activo. Su historial se conserva.")
    return redirect(url_for("admin_users"))


@app.route("/admin/roles", methods=["GET", "POST"])
@admin_required
def admin_roles():
    if request.method == "POST":
        name = request.form["name"].strip()
        if not name:
            flash("El nombre del cargo es obligatorio.")
            return redirect(url_for("admin_roles"))
        if len(name) > 50:
            flash("El cargo debe tener como máximo 50 caracteres.")
            return redirect(url_for("admin_roles"))
        db().execute("INSERT OR IGNORE INTO roles(name) VALUES (?)", (name,))
        db().commit()
        audit("role_created", f"Cargo creado: {name}")
        return redirect(url_for("admin_roles"))
    roles = db().execute("SELECT * FROM roles ORDER BY name").fetchall()
    return html_page(
        "Cargos",
        render_template_string(
            """
            <h1>Cargos</h1>
            <form class="panel" method="post">
              <label>Nuevo cargo</label><input class="field-role" name="name" maxlength="50" required>
              <p><button>Crear cargo</button></p>
            </form>
            <table>
              <tr><th>Cargo</th><th>Acciones</th></tr>
              {% for r in roles %}
                <tr>
                  <td>{{ r.name }}</td>
                  <td>
                    <a class="btn" href="{{ url_for('edit_role', role_id=r.id) }}">Editar</a>
                    {% if r.name != "Bombero" %}
                      <form method="post" action="{{ url_for('delete_role', role_id=r.id) }}" style="display:inline" onsubmit="return confirm('Eliminar este cargo?');">
                        <button class="danger">Eliminar</button>
                      </form>
                    {% endif %}
                  </td>
                </tr>
              {% endfor %}
            </table>
            """,
            roles=roles,
        ),
    )


@app.route("/admin/roles/<int:role_id>", methods=["GET", "POST"])
@admin_required
def edit_role(role_id: int):
    role = db().execute("SELECT * FROM roles WHERE id = ?", (role_id,)).fetchone()
    if not role:
        abort(404)
    if request.method == "POST":
        name = request.form["name"].strip()
        if not name:
            flash("El nombre del cargo es obligatorio.")
            return redirect(url_for("edit_role", role_id=role_id))
        if len(name) > 50:
            flash("El cargo debe tener como máximo 50 caracteres.")
            return redirect(url_for("edit_role", role_id=role_id))
        try:
            db().execute("UPDATE roles SET name = ? WHERE id = ?", (name, role_id))
            db().commit()
        except sqlite3.IntegrityError:
            db().rollback()
            flash("Ese cargo ya existe.")
            return redirect(url_for("edit_role", role_id=role_id))
        audit("role_updated", f"Cargo actualizado: {role['name']} -> {name}")
        flash("Cargo actualizado.")
        return redirect(url_for("admin_roles"))
    return html_page(
        "Editar cargo",
        render_template_string(
            """
            <h1>Editar cargo</h1>
            <form class="panel" method="post">
              <label>Nombre</label><input class="field-role" name="name" value="{{ role.name }}" maxlength="50" required>
              <p><button>Guardar cambios</button> <a class="btn" href="{{ url_for('admin_roles') }}">Volver</a></p>
            </form>
            """,
            role=role,
        ),
    )


@app.post("/admin/roles/<int:role_id>/delete")
@admin_required
def delete_role(role_id: int):
    role = db().execute("SELECT * FROM roles WHERE id = ?", (role_id,)).fetchone()
    if not role:
        abort(404)
    if role["name"] == "Bombero":
        flash("No se puede eliminar el cargo Bombero porque es el cargo por defecto.")
        return redirect(url_for("admin_roles"))
    users_count = db().execute("SELECT COUNT(*) AS total FROM users WHERE role_id = ?", (role_id,)).fetchone()["total"]
    rules_count = db().execute("SELECT COUNT(*) AS total FROM document_type_signer_roles WHERE role_id = ?", (role_id,)).fetchone()["total"]
    if users_count or rules_count:
        flash("No se puede eliminar el cargo porque esta asignado a usuarios o reglas de firma.")
        return redirect(url_for("admin_roles"))
    db().execute("DELETE FROM roles WHERE id = ?", (role_id,))
    db().commit()
    audit("role_deleted", f"Cargo eliminado: {role['name']}")
    flash("Cargo eliminado.")
    return redirect(url_for("admin_roles"))


@app.route("/admin/document-types", methods=["GET", "POST"])
@admin_required
def document_types():
    roles = db().execute("SELECT * FROM roles ORDER BY name").fetchall()
    if request.method == "POST":
        cur = db().execute(
            "INSERT INTO document_types(name, required_signatures, visible_to_completed_roles, created_at) VALUES (?, ?, ?, ?)",
            (
                request.form["name"].strip(),
                int(request.form["required_signatures"]),
                ",".join(request.form.getlist("visible_roles")),
                now(),
            ),
        )
        doc_type_id = cur.lastrowid
        for role_id in request.form.getlist("signer_roles"):
            db().execute("INSERT INTO document_type_signer_roles(document_type_id, role_id) VALUES (?, ?)", (doc_type_id, int(role_id)))
        db().commit()
        audit("document_type_created", f"Tipo documento creado: {request.form['name'].strip()}")
        return redirect(url_for("document_types"))
    types = db().execute("SELECT * FROM document_types ORDER BY name").fetchall()
    return html_page(
        "Tipos de documentos",
        render_template_string(
            """
            <h1>Tipos de documentos</h1>
            <form class="panel" method="post">
              <label>Nombre</label><input name="name" required>
              <label>Cantidad máxima de firmantes</label><input name="required_signatures" type="number" min="1" value="1" required>
              <h2>Cargos que pueden firmar</h2>
              {% for r in roles %}<label><input type="checkbox" name="signer_roles" value="{{ r.id }}"> {{ r.name }}</label>{% endfor %}
              <h2>Cargos que pueden ver cuando este cerrado</h2>
              {% for r in roles %}<label><input type="checkbox" name="visible_roles" value="{{ r.name }}"> {{ r.name }}</label>{% endfor %}
              <p><button>Crear tipo</button></p>
            </form>
            <table><tr><th>Tipo</th><th>Máximo firmantes</th><th>Visible cerrado para</th><th>Acciones</th></tr>
            {% for t in types %}
              <tr>
                <td>{{ t.name }}</td>
                <td>{{ t.required_signatures }}</td>
                <td>{{ t.visible_to_completed_roles }}</td>
                <td>
                  <a class="btn" href="{{ url_for('edit_document_type', document_type_id=t.id) }}">Editar</a>
                  <form method="post" action="{{ url_for('delete_document_type', document_type_id=t.id) }}" style="display:inline" onsubmit="return confirm('Eliminar este tipo de documento?');">
                    <button class="danger">Eliminar</button>
                  </form>
                </td>
              </tr>
            {% endfor %}</table>
            """,
            roles=roles,
            types=types,
        ),
    )


@app.route("/admin/document-types/<int:document_type_id>", methods=["GET", "POST"])
@admin_required
def edit_document_type(document_type_id: int):
    doc_type = db().execute("SELECT * FROM document_types WHERE id = ?", (document_type_id,)).fetchone()
    if not doc_type:
        abort(404)
    roles = db().execute("SELECT * FROM roles ORDER BY name").fetchall()
    selected_signer_roles = {
        row["role_id"]
        for row in db().execute(
            "SELECT role_id FROM document_type_signer_roles WHERE document_type_id = ?",
            (document_type_id,),
        ).fetchall()
    }
    selected_visible_roles = set(filter(None, doc_type["visible_to_completed_roles"].split(",")))

    if request.method == "POST":
        name = request.form["name"].strip()
        required_signatures = int(request.form["required_signatures"])
        signer_roles = [int(role_id) for role_id in request.form.getlist("signer_roles")]
        visible_roles = request.form.getlist("visible_roles")
        if required_signatures < 1:
            flash("La cantidad maxima de firmantes debe ser al menos 1.")
            return redirect(url_for("edit_document_type", document_type_id=document_type_id))
        if len(signer_roles) == 0:
            flash("Debes seleccionar al menos un cargo que pueda firmar.")
            return redirect(url_for("edit_document_type", document_type_id=document_type_id))

        db().execute(
            """
            UPDATE document_types
            SET name = ?, required_signatures = ?, visible_to_completed_roles = ?
            WHERE id = ?
            """,
            (name, required_signatures, ",".join(visible_roles), document_type_id),
        )
        db().execute("DELETE FROM document_type_signer_roles WHERE document_type_id = ?", (document_type_id,))
        for role_id in signer_roles:
            db().execute(
                "INSERT INTO document_type_signer_roles(document_type_id, role_id) VALUES (?, ?)",
                (document_type_id, role_id),
            )
        db().commit()
        audit("document_type_updated", f"Tipo documento actualizado: {name}")
        flash("Tipo de documento actualizado.")
        return redirect(url_for("document_types"))

    return html_page(
        "Editar tipo de documento",
        render_template_string(
            """
            <h1>Editar tipo de documento</h1>
            <form class="panel" method="post">
              <label>Nombre</label><input name="name" value="{{ doc_type.name }}" required>
              <label>Cantidad máxima de firmantes</label><input name="required_signatures" type="number" min="1" value="{{ doc_type.required_signatures }}" required>
              <h2>Cargos que pueden firmar</h2>
              {% for r in roles %}
                <label><input type="checkbox" name="signer_roles" value="{{ r.id }}" {% if r.id in selected_signer_roles %}checked{% endif %}> {{ r.name }}</label>
              {% endfor %}
              <h2>Cargos que pueden ver cuando este cerrado</h2>
              {% for r in roles %}
                <label><input type="checkbox" name="visible_roles" value="{{ r.name }}" {% if r.name in selected_visible_roles %}checked{% endif %}> {{ r.name }}</label>
              {% endfor %}
              <p><button>Guardar cambios</button> <a class="btn" href="{{ url_for('document_types') }}">Volver</a></p>
            </form>
            """,
            doc_type=doc_type,
            roles=roles,
            selected_signer_roles=selected_signer_roles,
            selected_visible_roles=selected_visible_roles,
        ),
    )


@app.post("/admin/document-types/<int:document_type_id>/delete")
@admin_required
def delete_document_type(document_type_id: int):
    doc_type = db().execute("SELECT * FROM document_types WHERE id = ?", (document_type_id,)).fetchone()
    if not doc_type:
        abort(404)
    documents_count = db().execute(
        "SELECT COUNT(*) AS total FROM documents WHERE document_type_id = ?",
        (document_type_id,),
    ).fetchone()["total"]
    if documents_count:
        flash("No se puede eliminar el tipo de documento porque ya tiene documentos asociados.")
        return redirect(url_for("document_types"))
    db().execute("DELETE FROM document_type_signer_roles WHERE document_type_id = ?", (document_type_id,))
    db().execute("DELETE FROM document_types WHERE id = ?", (document_type_id,))
    db().commit()
    audit("document_type_deleted", f"Tipo documento eliminado: {doc_type['name']}")
    flash("Tipo de documento eliminado.")
    return redirect(url_for("document_types"))


@app.route("/admin/audit")
@admin_required
def audit_view():
    return redirect(url_for("user_history"))


@app.route("/admin/history")
@login_required
def user_history():
    if not (g.user["is_admin"] or g.user["is_super_admin"] or g.user["can_view_user_history"]):
        abort(403)
    users = db().execute("SELECT id, name FROM users WHERE is_super_admin = 0 ORDER BY name").fetchall()
    selected_value = request.args.get("user_id", "general")
    selected_id = int(selected_value) if selected_value.isdigit() else None
    submitted = request.args.get("submitted") == "1"
    rows = get_history_rows(selected_id) if submitted else []
    can_export = bool(g.user["is_admin"] or g.user["is_super_admin"])
    return html_page(
        "Informes de usuarios",
        render_template_string(
            """
            <h1>Informes de usuarios</h1>
            <form class="panel" method="get">
              <input type="hidden" name="submitted" value="1">
              <label>Tipo de informe</label>
              <select name="user_id">
                <option value="general" {% if not selected_id %}selected{% endif %}>General</option>
                {% for u in users %}<option value="{{ u.id }}" {% if selected_id == u.id %}selected{% endif %}>{{ u.name }}</option>{% endfor %}
              </select>
              <p><button>Ver informe</button></p>
            </form>
            {% if submitted %}
              {% if can_export %}
                <div class="panel">
                  <button type="button" onclick="window.print()">Imprimir informe</button>
                  <a class="btn" href="{{ url_for('download_user_history', user_id=selected_value) }}">Descargar CSV</a>
                </div>
              {% endif %}
              <table>
                <tr><th>Fecha y hora</th><th>Actor</th><th>Acción</th><th>Detalle</th><th>Usuario afectado</th><th>Documento</th><th>IP</th></tr>
                {% for r in rows %}
                  <tr>
                    <td>{{ r.created_at }}</td>
                    <td>{{ r.actor_name or "Sistema" }}</td>
                    <td>{{ r.action }}</td>
                    <td>{{ r.details }}</td>
                    <td>{{ r.target_name or "" }}</td>
                    <td>{{ r.document_title or "" }}</td>
                    <td>{{ r.ip_address or "" }}</td>
                  </tr>
                {% endfor %}
              </table>
            {% endif %}
            """,
            users=users,
            selected_value=selected_value,
            selected_id=selected_id,
            submitted=submitted,
            can_export=can_export,
            rows=rows,
        ),
    )


def get_history_rows(selected_id: int | None):
    if selected_id:
        return db().execute(
            """
            SELECT a.*, actor.name AS actor_name, target.name AS target_name, d.title AS document_title
            FROM audit_log a
            LEFT JOIN users actor ON actor.id = a.actor_user_id
            LEFT JOIN users target ON target.id = a.target_user_id
            LEFT JOIN documents d ON d.id = a.document_id
            WHERE a.actor_user_id = ? OR a.target_user_id = ?
            ORDER BY a.created_at DESC
            """,
            (selected_id, selected_id),
        ).fetchall()
    return db().execute(
        """
        SELECT a.*, actor.name AS actor_name, target.name AS target_name, d.title AS document_title
        FROM audit_log a
        LEFT JOIN users actor ON actor.id = a.actor_user_id
        LEFT JOIN users target ON target.id = a.target_user_id
        LEFT JOIN documents d ON d.id = a.document_id
        ORDER BY a.created_at DESC
        LIMIT 500
        """
    ).fetchall()


@app.route("/admin/history/download")
@admin_required
def download_user_history():
    selected_value = request.args.get("user_id", "general")
    selected_id = int(selected_value) if selected_value.isdigit() else None
    rows = get_history_rows(selected_id)
    output = StringIO()
    writer = csv.writer(output)
    writer.writerow(["Fecha y hora", "Actor", "Acción", "Detalle", "Usuario afectado", "Documento", "IP"])
    for row in rows:
        writer.writerow(
            [
                row["created_at"],
                row["actor_name"] or "Sistema",
                row["action"],
                row["details"],
                row["target_name"] or "",
                row["document_title"] or "",
                row["ip_address"] or "",
            ]
        )
    filename = "informe-general.csv" if selected_id is None else f"informe-usuario-{selected_id}.csv"
    return Response(
        output.getvalue(),
        mimetype="text/csv",
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@app.route("/admin/emails")
@admin_required
def email_outbox():
    rows = db().execute("SELECT * FROM email_log ORDER BY created_at DESC LIMIT 300").fetchall()
    return html_page(
        "Bandeja de salida",
        render_template_string(
            """
            <h1>Bandeja de salida</h1>
            <div class="panel">
              <p class="muted">Si SMTP no está configurado, los avisos quedan aquí como not_configured.</p>
            </div>
            <table>
              <tr><th>Fecha</th><th>Para</th><th>Asunto</th><th>Estado</th><th>Error</th></tr>
              {% for r in rows %}
                <tr>
                  <td>{{ r.created_at }}</td>
                  <td>{{ r.to_email }}</td>
                  <td>{{ r.subject }}</td>
                  <td>{{ r.status }}</td>
                  <td>{{ r.error }}</td>
                </tr>
              {% endfor %}
            </table>
            """,
            rows=rows,
        ),
    )


if __name__ == "__main__":
    init_db()
    app.run(debug=True)


